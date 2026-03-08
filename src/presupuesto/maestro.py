"""Lector de las hojas 'Maestro' y 'Claves' de presupuesto.xlsx."""

from pathlib import Path

import openpyxl

# Mapeo de nombre de campo → índice de columna (1-based) en la hoja Maestro
_COLUMNAS_MAESTRO = {
    "anos": 1,
    "meses": 2,
    "categorias1": 3,
    "categorias2": 4,
    "categorias3": 5,
    "entidades": 6,
    "proveedores": 7,
    "tipos_gasto": 8,
    "cuentas": 9,
    "bancos": 10,
    "tipos_cuenta": 11,
}


def _leer_columna(hoja, col: int) -> list:
    """Lee todos los valores no vacíos de una columna, ignorando la fila 1 (cabecera)."""
    valores = []
    for fila in range(2, hoja.max_row + 1):
        valor = hoja.cell(fila, col).value
        if valor is not None and str(valor).strip() != "":
            valores.append(valor)
    return valores


class DatosMaestros:
    """Carga y expone los valores maestros y las claves de cuentas del archivo xlsx."""

    def __init__(self, ruta_archivo: str | Path):
        ruta = Path(ruta_archivo)
        if not ruta.exists():
            raise FileNotFoundError(f"No se encontró el archivo: {ruta}")

        wb = openpyxl.load_workbook(ruta, data_only=True)

        # --- Hoja Maestro ---
        hoja_maestro = wb["Maestro"]
        self._datos: dict[str, list] = {
            campo: _leer_columna(hoja_maestro, col)
            for campo, col in _COLUMNAS_MAESTRO.items()
        }
        # Los años vienen como int desde Excel; los normalizamos explícitamente
        self._datos["anos"] = [int(a) for a in self._datos["anos"]]

        # --- Hoja Claves: cuenta → (banco, tipo_cuenta) ---
        hoja_claves = wb["Claves"]
        self._claves: dict[str, tuple[str | None, str | None]] = {}
        for fila in range(2, hoja_claves.max_row + 1):
            cuenta = hoja_claves.cell(fila, 1).value
            banco = hoja_claves.cell(fila, 2).value
            tipo_cuenta = hoja_claves.cell(fila, 3).value
            if cuenta:
                self._claves[str(cuenta).strip()] = (
                    str(banco).strip() if banco else None,
                    str(tipo_cuenta).strip() if tipo_cuenta else None,
                )

        wb.close()

    # --- Propiedades de acceso a las listas ---

    @property
    def anos(self) -> list[int]:
        return self._datos["anos"]

    @property
    def meses(self) -> list[str]:
        return self._datos["meses"]

    @property
    def categorias1(self) -> list[str]:
        return self._datos["categorias1"]

    @property
    def categorias2(self) -> list[str]:
        return self._datos["categorias2"]

    @property
    def categorias3(self) -> list[str]:
        return self._datos["categorias3"]

    @property
    def entidades(self) -> list[str]:
        return self._datos["entidades"]

    @property
    def proveedores(self) -> list[str]:
        return self._datos["proveedores"]

    @property
    def tipos_gasto(self) -> list[str]:
        return self._datos["tipos_gasto"]

    @property
    def cuentas(self) -> list[str]:
        return self._datos["cuentas"]

    @property
    def bancos(self) -> list[str]:
        return self._datos["bancos"]

    @property
    def tipos_cuenta(self) -> list[str]:
        return self._datos["tipos_cuenta"]

    # --- Métodos de validación ---

    def validar(self, campo: str, valor) -> bool:
        """Devuelve True si el valor está en la lista del campo dado."""
        lista = self._datos.get(campo)
        if lista is None:
            raise ValueError(f"Campo desconocido: '{campo}'. Campos válidos: {list(_COLUMNAS_MAESTRO)}")
        return valor in lista

    def valores_validos(self, campo: str) -> list:
        """Devuelve la lista de valores válidos para un campo."""
        lista = self._datos.get(campo)
        if lista is None:
            raise ValueError(f"Campo desconocido: '{campo}'.")
        return list(lista)

    # --- Autocompletado de cuenta ---

    def autocompletar_cuenta(self, nombre_cuenta: str) -> tuple[str | None, str | None]:
        """Devuelve (banco, tipo_cuenta) para una cuenta dada usando la tabla Claves.

        Devuelve (None, None) si la cuenta no se encuentra.
        """
        return self._claves.get(nombre_cuenta.strip(), (None, None))

    def claves_cuentas(self) -> dict[str, tuple[str | None, str | None]]:
        """Devuelve el diccionario completo cuenta → (banco, tipo_cuenta)."""
        return dict(self._claves)
