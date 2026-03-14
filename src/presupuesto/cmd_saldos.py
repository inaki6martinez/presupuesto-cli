"""Comando 'saldos': evolución mes a mes del saldo de las cuentas.

Toma el saldo actual (suma de todas las entradas Real) y aplica los movimientos
Presupuesto futuros mes a mes, mostrando cómo evolucionará cada cuenta.
Por defecto las cuentas de tipo Pasivo se omiten; con --neto se incluyen.
"""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from pathlib import Path

import click

_MESES_ORD = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
              "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

_COL_AÑO     = 0
_COL_MES     = 1
_COL_IMPORTE = 6
_COL_CUENTA  = 9
_COL_ESTADO  = 12


# ---------------------------------------------------------------------------
# Lectura de datos
# ---------------------------------------------------------------------------

def _leer_datos(
    ruta_xlsx: Path,
    neto: bool = False,
) -> tuple[dict[str, Decimal], dict, list[tuple[str, str]]]:
    """Devuelve (saldo_actual, presupuesto, cuentas_info).

    - saldo_actual[cuenta]: suma acumulada de entradas Real.
    - presupuesto[año][mes][cuenta]: suma de entradas Presupuesto.
    - cuentas_info: lista de (cuenta, tipo_cuenta). Con neto=False se
                    excluyen las cuentas Pasivo; con neto=True se incluyen.
    """
    import openpyxl
    from presupuesto.escritor import leer_numero

    # Sin data_only: leemos fórmulas como texto y las evaluamos nosotros.
    # Necesario porque openpyxl borra el caché de fórmulas al guardar.
    wb = openpyxl.load_workbook(str(ruta_xlsx), read_only=True)

    # --- Cuentas válidas (no Pasivo) desde la hoja Claves ---
    cuentas_info: list[tuple[str, str]] = []
    try:
        ws_claves = wb["Claves"]
        for row in ws_claves.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            cuenta      = str(row[0]).strip()
            tipo_cuenta = str(row[2] or "").strip()
            if cuenta and (neto or tipo_cuenta.lower() != "pasivo"):
                cuentas_info.append((cuenta, tipo_cuenta))
    except KeyError:
        pass

    cuentas_validas = {c for c, _ in cuentas_info}

    # --- Datos ---
    saldo_actual: dict[str, Decimal] = defaultdict(Decimal)
    presupuesto: dict[int, dict[str, dict[str, Decimal]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(Decimal))
    )

    try:
        ws = wb["Datos"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[_COL_AÑO] is None:
                continue
            cuenta = str(row[_COL_CUENTA] or "").strip()
            if cuenta not in cuentas_validas:
                continue
            try:
                año = int(row[_COL_AÑO])
            except (TypeError, ValueError):
                continue
            mes    = str(row[_COL_MES]    or "").strip()
            estado = str(row[_COL_ESTADO] or "").strip()
            imp_raw = leer_numero(row[_COL_IMPORTE])
            if imp_raw is None:
                continue
            importe = Decimal(str(imp_raw))

            if estado == "Real":
                saldo_actual[cuenta] += importe
            elif estado == "Presupuesto":
                presupuesto[año][mes][cuenta] += importe
    except KeyError:
        pass

    wb.close()
    return dict(saldo_actual), presupuesto, cuentas_info


# ---------------------------------------------------------------------------
# Helpers de tabla
# ---------------------------------------------------------------------------

def _celda(valor: Decimal) -> str:
    if valor < 0:
        return f"[bold red]{valor:,.0f}[/bold red]"
    return f"{valor:,.0f}"


def _subtotal_row(
    tabla, label: str, cuentas: list[str],
    datos_col: dict[str, list[Decimal]],
    n_cols: int, bold: bool = True,
) -> list[Decimal]:
    """Añade fila de subtotal y devuelve la lista de valores."""
    vals_actual = [datos_col[c][0] for c in cuentas]
    vals_meses  = [
        [datos_col[c][i + 1] for c in cuentas]
        for i in range(n_cols - 1)
    ]
    total_actual = sum(vals_actual)
    totales_mes  = [sum(col) for col in vals_meses]
    fmt = "[bold]{}[/bold]" if bold else "{}"
    tabla.add_row(
        fmt.format(label),
        _celda(total_actual),
        *[_celda(t) for t in totales_mes],
    )
    return [total_actual] + totales_mes


# ---------------------------------------------------------------------------
# Comando click
# ---------------------------------------------------------------------------

@click.command("saldos")
@click.option("--neto", is_flag=True, default=False,
              help="Incluye cuentas Pasivo y muestra el saldo neto total.")
@click.option("--act", is_flag=True, default=False,
              help="Muestra solo el saldo actual, sin proyección de meses futuros.")
@click.option("--liquidez", is_flag=True, default=False,
              help="Agrupa las cuentas por tipo de liquidez con subtotales por grupo.")
def cmd_saldos(neto: bool, act: bool, liquidez: bool):
    """Muestra la evolución mensual del saldo de cada cuenta según el presupuesto."""
    from datetime import date
    from rich.console import Console
    from rich.table import Table
    from rich import box
    from presupuesto.config import cargar_config

    consola = Console()

    config   = cargar_config()
    ruta_str = config.get("archivo_presupuesto", "")
    if not ruta_str:
        consola.print("[red]No hay ruta al xlsx configurada.[/red]")
        raise SystemExit(1)
    ruta_xlsx = Path(ruta_str).expanduser()
    if not ruta_xlsx.exists():
        consola.print(f"[red]No se encuentra:[/red] {ruta_xlsx}")
        raise SystemExit(1)

    consola.print("[dim]Leyendo datos del xlsx…[/dim]")
    saldo_actual, presupuesto, cuentas_info = _leer_datos(ruta_xlsx, neto)

    cuentas_activos = [c for c, t in cuentas_info if t.lower() != "pasivo"]
    cuentas_pasivos = [c for c, t in cuentas_info if t.lower() == "pasivo"]
    cuentas         = [c for c, _ in cuentas_info]

    if not cuentas:
        consola.print("[yellow]No se encontraron cuentas en la hoja Claves.[/yellow]")
        return

    # --- Determinar rango de meses a mostrar ---
    hoy     = date.today()
    año_ini = hoy.year
    mes_ini = hoy.month

    if act:
        meses_rango: list[tuple[int, str]] = []
    else:
        año_max, mes_max_idx = año_ini, mes_ini
        for año, meses_dict in presupuesto.items():
            for i, mes in enumerate(_MESES_ORD, start=1):
                if mes in meses_dict:
                    if (año, i) > (año_max, mes_max_idx):
                        año_max, mes_max_idx = año, i

        if (año_max, mes_max_idx) < (año_ini, mes_ini):
            consola.print("[yellow]No hay entradas de Presupuesto futuras.[/yellow]")
            return

        meses_rango = []
        a, m = año_ini, mes_ini
        while (a, m) <= (año_max, mes_max_idx):
            meses_rango.append((a, _MESES_ORD[m - 1]))
            m += 1
            if m > 12:
                m = 1
                a += 1

    # --- Calcular saldos por cuenta y mes ---
    saldos: dict[str, Decimal] = {c: saldo_actual.get(c, Decimal(0)) for c in cuentas}
    columnas: list[tuple[int, str]] = []
    datos_col: dict[str, list[Decimal]] = {c: [] for c in cuentas}

    for c in cuentas:
        datos_col[c].append(saldos[c])

    for año, mes in meses_rango:
        mes_data = presupuesto.get(año, {}).get(mes, {})
        for cuenta in cuentas:
            saldos[cuenta] += mes_data.get(cuenta, Decimal(0))
            datos_col[cuenta].append(saldos[cuenta])
        columnas.append((año, mes))

    n_cols = 1 + len(columnas)   # Actual + meses proyectados

    # --- Construir tabla ---
    partes = []
    if neto:
        partes.append("netos")
    if liquidez:
        partes.append("por liquidez")
    if act:
        partes.append("actual")
    titulo = "Saldos" + (f" ({', '.join(partes)})" if partes else " — evolución (€)")

    tabla = Table(
        title=titulo,
        box=box.SIMPLE_HEAD,
        show_lines=False,
        padding=(0, 1),
    )
    tabla.add_column("Cuenta", no_wrap=True, min_width=22)
    tabla.add_column("Actual", justify="right", no_wrap=True, style="bold")
    for año, mes in columnas:
        tabla.add_column(f"{mes} {año}", justify="right", no_wrap=True)

    def _sep_row(label: str = "") -> None:
        tabla.add_row(f"[dim]{label}[/dim]", *[""] * n_cols)

    if liquidez:
        # Agrupar por tipo_cuenta preservando el orden de Claves
        tipo_a_cuentas: dict[str, list[str]] = {}
        for cuenta, tipo in cuentas_info:
            tipo_a_cuentas.setdefault(tipo, []).append(cuenta)

        totales_global: list[Decimal] = [Decimal(0)] * n_cols

        for tipo, grupo in tipo_a_cuentas.items():
            es_pasivo = tipo.lower() == "pasivo"
            # Cabecera del grupo
            tabla.add_row(
                f"[bold {'red' if es_pasivo else 'cyan'}]{tipo}[/bold {'red' if es_pasivo else 'cyan'}]",
                *[""] * n_cols,
            )
            for cuenta in grupo:
                estilo = "[dim]" if es_pasivo else ""
                cierre = "[/dim]" if es_pasivo else ""
                tabla.add_row(
                    f"  {estilo}{cuenta}{cierre}",
                    *[_celda(v) for v in datos_col[cuenta]],
                )
            # Subtotal del grupo
            sub = _subtotal_row(tabla, f"  Subtotal", grupo, datos_col, n_cols)
            for i, v in enumerate(sub):
                totales_global[i] += v
            _sep_row()

        # Total global
        tabla.add_row(
            "[bold]Total[/bold]",
            *[_celda(v) for v in totales_global],
        )

    else:
        # Vista plana (comportamiento original)
        for cuenta in cuentas_activos:
            tabla.add_row(cuenta, *[_celda(v) for v in datos_col[cuenta]])

        if neto and cuentas_pasivos:
            act_vals = _subtotal_row(tabla, "Activos", cuentas_activos, datos_col, n_cols)
            _sep_row()

            for cuenta in cuentas_pasivos:
                tabla.add_row(f"[dim]{cuenta}[/dim]",
                              *[_celda(v) for v in datos_col[cuenta]])

            pas_vals = _subtotal_row(tabla, "Pasivos", cuentas_pasivos, datos_col, n_cols)
            _sep_row()

            neto_vals = [a + p for a, p in zip(act_vals, pas_vals)]
            tabla.add_row("[bold]Neto[/bold]", *[_celda(v) for v in neto_vals])
        else:
            _subtotal_row(tabla, "Total", cuentas, datos_col, n_cols)

    consola.print()
    consola.print(tabla)
    consola.print("  [dim]Saldos en €. [bold red]Rojo[/bold red] = negativo.[/dim]")
    consola.print()
