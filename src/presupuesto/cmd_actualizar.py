"""Comando 'actualizar': ajusta el balance de una cuenta con una entrada Finanzas/Balance.

Flujo:
1. Lee los balances actuales de cada cuenta sumando la hoja 'Datos' (Estado=Real).
2. TUI para seleccionar la cuenta.
3. Prompt para introducir el valor real actual.
4. Calcula la diferencia y escribe una entrada Finanzas/Balance en el xlsx.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import TYPE_CHECKING

import click

if TYPE_CHECKING:
    pass

_MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
          "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


# ---------------------------------------------------------------------------
# Lectura de balances desde la hoja Datos
# ---------------------------------------------------------------------------

def leer_balances(ruta_xlsx: str | Path) -> dict[str, Decimal]:
    """Calcula el balance actual de cada cuenta sumando la hoja 'Datos' (Estado=Real).

    Columnas: A=Año(0) B=Mes(1) C=Cat1(2) D=Cat2(3) E=Cat3(4) F=Entidad(5)
              G=Importe(6) H=Proveedor(7) I=TipoGasto(8) J=Cuenta(9)
              K=Banco(10) L=TipoCuenta(11) M=Estado(12)
    """
    import openpyxl

    ruta = Path(ruta_xlsx)
    balances: dict[str, Decimal] = {}

    if not ruta.exists():
        return balances

    wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=True)
    try:
        ws = wb["Datos"]
    except KeyError:
        wb.close()
        return balances

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        estado = str(row[12] or "").strip().lower() if len(row) > 12 else ""
        if estado != "real":
            continue
        cuenta = str(row[9] or "").strip() if len(row) > 9 else ""
        if not cuenta:
            continue
        try:
            importe = Decimal(str(row[6] or 0))
        except InvalidOperation:
            continue
        balances[cuenta] = balances.get(cuenta, Decimal(0)) + importe

    wb.close()
    return balances


def leer_cuentas(ruta_xlsx: str | Path) -> list[tuple[str, str, str]]:
    """Devuelve lista de (cuenta, banco, tipo_cuenta) desde la hoja Claves."""
    import openpyxl

    ruta = Path(ruta_xlsx)
    if not ruta.exists():
        return []

    wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=True)
    try:
        ws = wb["Claves"]
    except KeyError:
        wb.close()
        return []

    cuentas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        cuenta     = str(row[0]).strip()
        banco      = str(row[1]).strip() if row[1] else ""
        tipo_cuenta = str(row[2]).strip() if row[2] else ""
        if cuenta:
            cuentas.append((cuenta, banco, tipo_cuenta))

    wb.close()
    return cuentas


# ---------------------------------------------------------------------------
# TUI selección de cuenta
# ---------------------------------------------------------------------------

def _tui_seleccionar_cuenta(
    cuentas: list[tuple[str, str, str]],
    balances: dict[str, Decimal],
) -> tuple[str, str, str] | None:
    """TUI full-screen para seleccionar una cuenta. Devuelve (cuenta, banco, tipo_cuenta) o None."""
    from prompt_toolkit import Application
    from prompt_toolkit.application import get_app
    from prompt_toolkit.formatted_text import FormattedText
    from prompt_toolkit.key_binding import KeyBindings
    from prompt_toolkit.layout import Layout
    from prompt_toolkit.layout.containers import Window
    from prompt_toolkit.layout.controls import FormattedTextControl
    from prompt_toolkit.styles import Style

    style = Style.from_dict({
        "titulo":  "bold",
        "cursor":  "reverse bold",
        "filtro":  "bold yellow",
        "dim":     "#666666",
        "neg":     "#ff5555",
        "pos":     "#55ff55",
        "footer":  "#666666",
        "fkey":    "#aaaaaa bold",
    })

    state: dict = {"cursor": 0, "filtro": "", "resultado": None}

    def _filtradas() -> list[tuple[str, str, str]]:
        f = state["filtro"].lower()
        if not f:
            return cuentas
        return [c for c in cuentas if f in c[0].lower()]

    def _clamp() -> None:
        vis = _filtradas()
        state["cursor"] = max(0, min(state["cursor"], len(vis) - 1))

    def _render() -> FormattedText:
        try:
            size = get_app().output.get_size()
            w, h = size.columns, size.rows
        except Exception:
            w, h = 120, 40

        buf: list[tuple[str, str]] = []
        def t(st: str, s: str) -> None: buf.append((st, s))
        def nl() -> None: buf.append(("", "\n"))

        t("class:titulo", "  Actualizar balance de cuenta")
        nl()
        filtro_txt = (state["filtro"] + "▌") if state["filtro"] else "▌"
        t("class:dim",    "  Filtro: ")
        t("class:filtro", filtro_txt)
        nl()
        t("class:dim", "─" * w)
        nl()

        t("class:dim", f"  {'Cuenta':<30}  {'Banco':<20}  {'Balance actual':>15}")
        nl()
        t("class:dim", "─" * w)
        nl()

        vis = _filtradas()
        list_h = max(3, h - 9)
        cur = state["cursor"]
        ws_start = max(0, cur - list_h // 2)
        ws_end   = min(len(vis), ws_start + list_h)
        ws_start = max(0, ws_end - list_h)

        for i in range(ws_start, ws_end):
            cuenta, banco, _ = vis[i]
            balance = balances.get(cuenta, Decimal(0))
            es_cur  = i == cur
            arrow   = "►" if es_cur else " "
            imp_st  = "class:neg" if balance < 0 else "class:pos"
            row_st  = "class:cursor" if es_cur else ""

            t("class:dim", f"  {arrow} ")
            t(row_st, f"{cuenta:<30}  {banco:<20}  ")
            t(row_st if es_cur else imp_st, f"{balance:>+14.2f}€")
            nl()

        t("class:dim", "─" * w)
        nl()
        for k, desc in [("↑↓", "Navegar"), ("Enter", "Seleccionar"),
                        ("^U", "Borrar filtro"), ("Esc", "Salir")]:
            t("class:fkey",   f" {k} ")
            t("class:footer", f"{desc}  ")

        return FormattedText(buf)

    kb = KeyBindings()

    @kb.add("up")
    def _(e):
        state["cursor"] = max(0, state["cursor"] - 1)

    @kb.add("down")
    def _(e):
        vis = _filtradas()
        state["cursor"] = min(max(0, len(vis) - 1), state["cursor"] + 1)

    @kb.add("enter")
    def _(e):
        vis = _filtradas()
        if vis:
            state["resultado"] = vis[state["cursor"]]
        e.app.exit()

    @kb.add("escape")
    @kb.add("c-c")
    def _(e): e.app.exit()  # noqa: E704

    @kb.add("backspace")
    @kb.add("c-h")
    def _(e):
        state["filtro"] = state["filtro"][:-1]
        _clamp()

    @kb.add("c-u")
    def _(e):
        state["filtro"] = ""
        _clamp()

    @kb.add("<any>")
    def _(e):
        key = e.key_sequence[0].key
        if isinstance(key, str) and len(key) == 1 and key.isprintable():
            state["filtro"] += key
            state["cursor"] = 0

    app = Application(
        layout=Layout(Window(content=FormattedTextControl(text=_render, focusable=True))),
        key_bindings=kb,
        style=style,
        full_screen=True,
    )
    app.run()
    return state["resultado"]


# ---------------------------------------------------------------------------
# Comando click
# ---------------------------------------------------------------------------

@click.command("actualizar")
def cmd_actualizar():
    """Ajusta el balance de una cuenta introduciendo su valor real actual."""
    from rich.console import Console
    from presupuesto.config import cargar_config
    from presupuesto.escritor import EscritorDatos

    consola = Console()

    config = cargar_config()
    ruta_xlsx = config.get("archivo_presupuesto", "")
    if not ruta_xlsx:
        consola.print("[red]No hay ruta al xlsx configurada. Ejecuta 'presupuesto config'.[/red]")
        raise SystemExit(1)
    ruta_xlsx = Path(ruta_xlsx).expanduser()
    if not ruta_xlsx.exists():
        consola.print(f"[red]No se encuentra el archivo:[/red] {ruta_xlsx}")
        raise SystemExit(1)

    # --- Leer datos del xlsx ---
    consola.print("[dim]Leyendo datos del xlsx…[/dim]")
    cuentas  = leer_cuentas(ruta_xlsx)
    balances = leer_balances(ruta_xlsx)

    if not cuentas:
        consola.print("[red]No se encontraron cuentas en la hoja Claves.[/red]")
        raise SystemExit(1)

    from presupuesto.categorizar import MovimientoCategorizado
    from presupuesto.duplicados import GestorMarcadores
    ruta_marcadores = Path("~/.config/presupuesto/marcadores.json").expanduser()
    gestor_marcadores = GestorMarcadores(ruta_marcadores)

    # --- Bucle: TUI → actualizar → volver a TUI ---
    while True:
        seleccion = _tui_seleccionar_cuenta(cuentas, balances)
        if seleccion is None:
            consola.print("[dim]Saliendo.[/dim]")
            return

        cuenta, banco, tipo_cuenta = seleccion
        balance_actual = balances.get(cuenta, Decimal(0))

        consola.print(f"\n  Cuenta:         [bold]{cuenta}[/bold]")
        consola.print(f"  Balance actual: [cyan]{balance_actual:+.2f}€[/cyan]")

        # Pedir nuevo valor
        while True:
            raw = click.prompt("\n  Nuevo valor real").strip().replace(",", ".")
            try:
                nuevo_valor = Decimal(raw).quantize(Decimal("0.01"))
                break
            except InvalidOperation:
                consola.print("  [red]Valor no válido. Usa formato numérico (ej: 1234.56)[/red]")

        diferencia = nuevo_valor - balance_actual

        consola.print(f"\n  Nuevo valor:    [bold]{nuevo_valor:+.2f}€[/bold]")
        color_dif = "green" if diferencia >= 0 else "red"
        consola.print(f"  Diferencia:     [{color_dif}]{diferencia:+.2f}€[/{color_dif}]")

        if diferencia == 0:
            consola.print("\n  [yellow]Diferencia 0, no hay nada que ajustar.[/yellow]")
            continue

        hoy = date.today()
        mes = _MESES[hoy.month - 1]

        consola.print(f"\n  Se escribirá:  {hoy.year} {mes}  Finanzas / Balance  "
                      f"{diferencia:+.2f}€  {cuenta}")

        if not click.confirm("\n  ¿Confirmar?", default=True):
            consola.print("  [dim]Cancelado, volviendo a la lista.[/dim]")
            continue

        # Escribir
        mov = MovimientoCategorizado(
            año=hoy.year,
            mes=mes,
            categoria1="Finanzas",
            categoria2="Balance",
            categoria3="",
            entidad="",
            importe=diferencia,
            proveedor="",
            tipo_gasto="",
            cuenta=cuenta,
            banco=banco or None,
            tipo_cuenta=tipo_cuenta or None,
            estado="Real",
            confianza="alta",
            fuente="actualizar",
            requiere_confirmacion=False,
            concepto_original=f"Ajuste balance {cuenta} → {nuevo_valor:+.2f}€",
        )

        try:
            n = EscritorDatos(ruta_xlsx).escribir([mov])
            consola.print(f"  [green]✓ {n} entrada(s) escritas.[/green]")
        except Exception as e:
            consola.print(f"  [red]Error al escribir:[/red] {e}")
            continue

        # Actualizar marcador y balance local para la próxima iteración
        gestor_marcadores.actualizar_marcador(cuenta, hoy)
        consola.print(f"  [dim]Marcador actualizado: {cuenta} → {hoy}[/dim]")
        balances[cuenta] = nuevo_valor
