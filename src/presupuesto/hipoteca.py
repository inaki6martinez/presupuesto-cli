"""Expansión de cuotas hipotecarias en movimientos separados.

Cuando se detecta un movimiento de cuota hipotecaria (Vivienda/Hipoteca/Piso),
se reemplaza por tres movimientos consultando la hoja "Cuadro hipteca":

1. Intereses (negativo)   → Vivienda / Hipoteca  (gasto real)
2. Amortización (negativo)→ Ahorro / Hipoteca    (reducción de deuda = ahorro)
3. Balance (positivo)     → Finanzas / Balance   (contrapartida en cuenta Hipoteca Piso)
"""

from __future__ import annotations

import dataclasses
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from presupuesto.categorizar import MovimientoCategorizado
    from presupuesto.maestro import DatosMaestros

# Nombres de las columnas en la hoja (fila 6 = cabecera, datos desde fila 7)
_HOJA = "Cuadro hipteca"
_COL_FECHA       = 1   # A
_COL_INTERESES   = 4   # D
_COL_AMORTIZACION = 5  # E

# Criterio para identificar movimientos de cuota hipotecaria
_CATEGORIA1_HIPOTECA = "Vivienda"
_CATEGORIA2_HIPOTECA = "Hipoteca"
_ENTIDAD_HIPOTECA    = "Piso"

# Cuenta donde se registra la reducción de capital
_CUENTA_HIPOTECA_PISO = "Hipoteca Piso"


def es_cuota_hipoteca(mov: MovimientoCategorizado) -> bool:
    """True si el movimiento corresponde a una cuota hipotecaria a expandir."""
    return (
        mov.categoria1 == _CATEGORIA1_HIPOTECA
        and mov.categoria2 == _CATEGORIA2_HIPOTECA
        and mov.entidad == _ENTIDAD_HIPOTECA
    )


def buscar_cuota(año: int, mes_abr: str, ruta_xlsx: str | Path) -> tuple[Decimal, Decimal] | None:
    """Devuelve (intereses, amortización) para el año/mes dado.

    Primero intenta leer los valores de las celdas de la hoja (si no son None/0).
    Si son None (fórmulas sin caché), calcula desde los parámetros estáticos:
      - Fila 2 col B: Capital inicial
      - Fila 3 col B: Tasa anual
      - Fila 4 col B: Plazo en meses
      - Fila 7 col A: Fecha del primer pago

    mes_abr es la abreviatura española de 3 letras (Ene, Feb, …, Dic).
    Devuelve None si no se encuentra la hoja o la fecha no pertenece al cuadro.
    """
    import openpyxl

    _MES_NUM = {
        "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
        "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
    }
    mes_num = _MES_NUM.get(mes_abr.lower())
    if mes_num is None:
        return None

    ruta = Path(ruta_xlsx)
    if not ruta.exists():
        return None

    try:
        wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=True)
        if _HOJA not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[_HOJA]

        # Leer parámetros estáticos (filas 2-4, col B = índice 1)
        filas_param = list(ws.iter_rows(min_row=2, max_row=4, values_only=True))
        capital_inicial = filas_param[0][1] if filas_param[0][1] else None
        tasa_anual      = filas_param[1][1] if filas_param[1][1] else None
        plazo_meses     = filas_param[2][1] if filas_param[2][1] else None

        # Buscar la fila del mes solicitado y leer valores o número de cuota
        fecha_primer_pago = None
        intereses_celda   = None
        amortiz_celda     = None
        num_cuota         = None

        for row in ws.iter_rows(min_row=7, values_only=True):
            fecha_val = row[_COL_FECHA - 1]
            if fecha_val is None:
                continue
            try:
                fila_año = fecha_val.year
                fila_mes = fecha_val.month
            except AttributeError:
                continue
            if fecha_primer_pago is None:
                fecha_primer_pago = fecha_val
            if fila_año == año and fila_mes == mes_num:
                intereses_celda = row[_COL_INTERESES    - 1]
                amortiz_celda   = row[_COL_AMORTIZACION - 1]
                # Número de cuota = diferencia de meses desde el primer pago + 1
                meses_desde = (
                    (año - fecha_primer_pago.year) * 12
                    + (mes_num - fecha_primer_pago.month)
                )
                num_cuota = meses_desde + 1
                break

        wb.close()

        if num_cuota is None:
            return None  # el mes no está en el cuadro

        # Usar valores de celda si están disponibles y son positivos
        if intereses_celda and amortiz_celda:
            return (
                Decimal(str(intereses_celda)).quantize(Decimal("0.01")),
                Decimal(str(amortiz_celda)).quantize(Decimal("0.01")),
            )

        # Fallback: calcular desde parámetros estáticos
        if not (capital_inicial and tasa_anual and plazo_meses and num_cuota >= 1):
            return None

        return _calcular_cuota_anualidad(
            float(capital_inicial), float(tasa_anual),
            int(plazo_meses), num_cuota,
        )

    except Exception:
        return None


def _calcular_cuota_anualidad(
    capital: float,
    tasa_anual: float,
    plazo: int,
    num_cuota: int,
) -> tuple[Decimal, Decimal] | None:
    """Calcula (intereses, amortización) para la cuota num_cuota de una hipoteca de anualidad fija.

    Fórmula estándar de préstamo francés (cuota constante):
        r  = tasa_anual / 12
        C  = capital * r / (1 - (1+r)^(-plazo))
        capital_pendiente(k) = capital*(1+r)^k - C*((1+r)^k - 1)/r
        intereses(k+1) = capital_pendiente(k) * r
        amortización(k+1) = C - intereses(k+1)
    """
    if num_cuota < 1 or num_cuota > plazo:
        return None

    r = tasa_anual / 12
    if r == 0:
        cuota = capital / plazo
        intereses    = 0.0
        amortizacion = cuota
    else:
        cuota = capital * r / (1 - (1 + r) ** (-plazo))
        k = num_cuota - 1   # cuotas ya pagadas antes de esta
        cap_pendiente = capital * (1 + r) ** k - cuota * ((1 + r) ** k - 1) / r
        intereses    = cap_pendiente * r
        amortizacion = cuota - intereses

    return (
        Decimal(str(round(intereses,    2))).quantize(Decimal("0.01")),
        Decimal(str(round(amortizacion, 2))).quantize(Decimal("0.01")),
    )


def expandir_hipoteca(
    mov: MovimientoCategorizado,
    ruta_xlsx: str | Path,
    maestros: DatosMaestros,
) -> list[MovimientoCategorizado]:
    """Reemplaza una cuota hipotecaria por 3 movimientos desglosados.

    Si no se encuentra la cuota en el cuadro, devuelve [mov] sin cambios.
    """
    resultado = buscar_cuota(mov.año, mov.mes, ruta_xlsx)
    if resultado is None:
        return [mov]

    intereses, amortizacion = resultado

    banco_hip, tipo_hip = maestros.autocompletar_cuenta(_CUENTA_HIPOTECA_PISO)

    concepto_base = mov.concepto_original or ""

    mov_intereses = dataclasses.replace(
        mov,
        importe      = -intereses,
        categoria1   = "Vivienda",
        categoria2   = "Hipoteca",
        categoria3   = "",
        entidad      = "Piso",
        proveedor    = "",
        tipo_gasto   = "Fijos",
        confianza    = "alta",
        fuente       = "hipoteca:intereses",
        requiere_confirmacion = False,
        concepto_original = f"{concepto_base} [intereses]",
    )

    mov_amortizacion = dataclasses.replace(
        mov,
        importe      = -amortizacion,
        categoria1   = "Ahorro",
        categoria2   = "Hipoteca",
        categoria3   = "",
        entidad      = "Piso",
        proveedor    = "",
        tipo_gasto   = "Fijos",
        confianza    = "alta",
        fuente       = "hipoteca:amortizacion",
        requiere_confirmacion = False,
        concepto_original = f"{concepto_base} [amortizacion]",
    )

    mov_balance = dataclasses.replace(
        mov,
        importe      = amortizacion,
        categoria1   = "Finanzas",
        categoria2   = "Balance",
        categoria3   = "",
        entidad      = "",
        proveedor    = "",
        tipo_gasto   = "",
        cuenta       = _CUENTA_HIPOTECA_PISO,
        banco        = banco_hip,
        tipo_cuenta  = tipo_hip,
        confianza    = "alta",
        fuente       = "hipoteca:balance",
        requiere_confirmacion = False,
        concepto_original = f"{concepto_base} [balance]",
    )

    return [mov_intereses, mov_amortizacion, mov_balance]


def expandir_hipotecas(
    movimientos: list[MovimientoCategorizado],
    ruta_xlsx: str | Path,
    maestros: DatosMaestros,
) -> list[MovimientoCategorizado]:
    """Expande todos los movimientos hipotecarios de la lista."""
    resultado: list[MovimientoCategorizado] = []
    for mov in movimientos:
        if es_cuota_hipoteca(mov):
            resultado.extend(expandir_hipoteca(mov, ruta_xlsx, maestros))
        else:
            resultado.append(mov)
    return resultado
