"""Parser de extractos bancarios de BBVA.

BBVA exporta un XLSX con una hoja "Informe BBVA" en dos idiomas:

Inglés (formato americano):
- Fila 5: Eff. Date | Date | Item | Transaction | Amount | ... | Comments
- Fechas en MM/DD/YYYY

Español:
- Fila 5: F.Valor | Fecha | Concepto | Movimiento | Importe | ... | Observaciones
- Fechas en DD/MM/YYYY

Las columnas son las mismas en ambos formatos (B=fecha, D=item/concepto,
E=transaction/movimiento, F=importe, J=comments/observaciones).
"""

from __future__ import annotations

import re
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl

from presupuesto.parsers.base import MovimientoCrudo, ParserBase

# ── Patrones inglés ────────────────────────────────────────────────────────────

_ITEMS_GENERICOS_EN = re.compile(
    r"^(transfer (received|completed)|service company debit|"
    r"fee for |credit interest|debit interest)",
    re.IGNORECASE,
)
_ITEMS_USA_COMMENTS_EN = re.compile(
    r"^service company debit$",
    re.IGNORECASE,
)
_TRANSACTION_GENERICA_EN = re.compile(
    r"^(card payment|debit no \d+|payment of sepa direct debit)$",
    re.IGNORECASE,
)

# ── Patrones español ───────────────────────────────────────────────────────────

# Conceptos genéricos donde el Movimiento tiene la info real
_ITEMS_GENERICOS_ES = re.compile(
    r"^(transferencia (recibida|realizada)|cargo por )",
    re.IGNORECASE,
)
# Conceptos donde Observaciones tiene el nombre real (domiciliaciones)
_ITEMS_USA_COMMENTS_ES = re.compile(
    r"^adeudo",
    re.IGNORECASE,
)
# Movimientos genéricos que no aportan info (ignorar → usar Concepto)
_TRANSACTION_GENERICA_ES = re.compile(
    r"^(pago con tarjeta|adeudo n[oº]\s+\d+|pago de adeudo directo sepa)$",
    re.IGNORECASE,
)

# ── Común ──────────────────────────────────────────────────────────────────────

# Prefijo numérico en Comments/Observaciones: "N 2026061002269417 Aguas Municipales..."
_RE_PREFIJO_COMMENTS = re.compile(r"^[N]\s+\d+\s+", re.IGNORECASE)

# Headers que identifican cada idioma
_HEADERS_EN = {"eff. date", "item", "amount"}
_HEADERS_ES = {"f.valor", "concepto", "importe"}

# Nombres de columna por idioma: fecha, item/concepto, transaction/movimiento,
# importe, comments/observaciones
_COL_NOMBRES = {
    "en": ("eff. date", "item",     "transaction", "amount",  "comments"),
    "es": ("f.valor",   "concepto", "movimiento",  "importe", "observaciones"),
}


def _limpiar_comments(texto: str) -> str:
    """Elimina el prefijo 'N 123456789 ' de Comments/Observaciones de BBVA."""
    return _RE_PREFIJO_COMMENTS.sub("", texto).strip()


def _parsear_fecha(valor: str, lang: str) -> date:
    """Convierte fecha BBVA a date. Inglés: MM/DD/YYYY. Español: DD/MM/YYYY."""
    valor = valor.strip()
    try:
        a, b, c = valor.split("/")
        if lang == "es":
            dia, mes, año = a, b, c
        else:
            mes, dia, año = a, b, c
        return date(int(año), int(mes), int(dia))
    except ValueError as e:
        raise ValueError(f"Fecha de BBVA no reconocida: '{valor}'") from e


def _construir_concepto(
    item: str, transaction: str, comments: str, lang: str
) -> str:
    """Elige el texto más informativo como concepto según el idioma del extracto.

    Prioridad:
    1. Item/Concepto es una domiciliación → usar Comments/Observaciones (nombre real).
    2. Transaction/Movimiento es genérico o vacío → usar solo Item/Concepto.
    3. Item/Concepto es genérico (transferencia...) → usar Transaction/Movimiento.
    4. Ambos con info → "{Item} - {Transaction}".
    """
    item        = " ".join(item.split())
    transaction = " ".join(transaction.split())
    comments    = " ".join(comments.split())

    if not item:
        return transaction or comments

    if lang == "es":
        usa_comments  = _ITEMS_USA_COMMENTS_ES
        item_generico = _ITEMS_GENERICOS_ES
        trans_generia = _TRANSACTION_GENERICA_ES
    else:
        usa_comments  = _ITEMS_USA_COMMENTS_EN
        item_generico = _ITEMS_GENERICOS_EN
        trans_generia = _TRANSACTION_GENERICA_EN

    if usa_comments.match(item):
        comentario_limpio = _limpiar_comments(comments)
        if comentario_limpio:
            return comentario_limpio

    if not transaction or trans_generia.match(transaction):
        return item
    if item_generico.match(item):
        return transaction

    return f"{item} - {transaction}"


class ParserBBVA(ParserBase):
    """Parser para extractos XLSX de BBVA en inglés y español."""

    def puede_parsear(self, ruta_archivo: str) -> bool:
        ruta = Path(ruta_archivo)
        if ruta.suffix.lower() != ".xlsx":
            return False
        try:
            wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                vals = {str(v).strip().lower() for v in row if v is not None}
                if _HEADERS_EN <= vals or _HEADERS_ES <= vals:
                    return True
            return False
        except Exception:
            return False

    def parsear(self, ruta_archivo: str) -> list[MovimientoCrudo]:
        ruta = Path(ruta_archivo)
        wb   = openpyxl.load_workbook(str(ruta), data_only=True)
        ws   = wb.active

        col_fecha = col_item = col_trans = col_importe = col_comments = None
        fila_inicio = None
        lang = "en"

        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            vals = [str(v).strip().lower() if v is not None else "" for v in row]

            for idioma, (n_fecha, n_item, n_trans, n_imp, n_com) in _COL_NOMBRES.items():
                if n_fecha in vals and n_imp in vals:
                    col_fecha    = vals.index(n_fecha) + 1
                    col_item     = vals.index(n_item)  + 1
                    col_trans    = vals.index(n_trans) + 1
                    col_importe  = vals.index(n_imp)   + 1
                    col_comments = vals.index(n_com)   + 1 if n_com in vals else None
                    fila_inicio  = r_idx + 1
                    lang         = idioma
                    break
            if fila_inicio:
                break

        if fila_inicio is None:
            raise ValueError(f"No se encontró la fila de cabecera en {ruta.name}")

        movimientos: list[MovimientoCrudo] = []

        for row in ws.iter_rows(min_row=fila_inicio, values_only=True):
            fecha_val    = row[col_fecha - 1]
            item_val     = row[col_item  - 1]
            trans_val    = row[col_trans - 1]
            importe_val  = row[col_importe - 1]
            comments_val = row[col_comments - 1] if col_comments else None

            if fecha_val is None or importe_val is None:
                continue

            fecha_str    = str(fecha_val).strip()
            item_raw     = str(item_val).strip()    if item_val     else ""
            trans_raw    = str(trans_val).strip()   if trans_val    else ""
            comments_raw = str(comments_val).strip() if comments_val else ""
            importe_str  = str(importe_val).strip()

            if not fecha_str:
                continue

            fecha    = _parsear_fecha(fecha_str, lang)
            importe  = Decimal(importe_str).quantize(Decimal("0.01"))
            concepto = _construir_concepto(item_raw, trans_raw, comments_raw, lang)

            concepto_original = f"{fecha_str} | {item_raw}"
            if trans_raw:
                concepto_original += f" | {trans_raw}"

            movimientos.append(MovimientoCrudo(
                fecha=fecha,
                concepto=concepto,
                importe=importe,
                concepto_original=concepto_original,
            ))

        return movimientos
