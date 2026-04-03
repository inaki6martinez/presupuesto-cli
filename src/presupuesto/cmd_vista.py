"""Comando 'vista': presupuesto a un año vista (TUI interactivo).

Vistas:
  main   — tabla resumen categorías × meses
  detail — lista de entradas con multiselección
  edit   — formulario de campos (aplica a seleccionadas)
  input  — edición de texto libre (año, mes, importe)
  picker — selector filtrable de opciones (cat1/2/3, entidad, cuenta)
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

import click

_MESES_ORD = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
               "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

_COL_AÑO        = 0
_COL_MES        = 1
_COL_CAT1       = 2
_COL_CAT2       = 3
_COL_CAT3       = 4
_COL_ENTIDAD    = 5
_COL_IMPORTE    = 6
_COL_TIPO_GASTO = 8
_COL_CUENTA     = 9
_COL_ESTADO     = 12

# Campos del formulario de edición: (clave, etiqueta, usa_picker)
_CAMPOS = [
    ("año",        "Año",        False),
    ("mes",        "Mes",        False),
    ("importe",    "Importe",    False),
    ("cat1",       "Cat1",       True),
    ("cat2",       "Cat2",       True),
    ("cat3",       "Cat3",       True),
    ("entidad",    "Entidad",    True),
    ("cuenta",     "Cuenta",     True),
    ("tipo_gasto", "Tipo gasto", True),
]


# ---------------------------------------------------------------------------
# Estructuras de datos
# ---------------------------------------------------------------------------

@dataclass
class _Entrada:
    fila_xlsx: int
    año: int
    mes: str
    importe: Decimal
    cat1: str
    cat2: str
    cat3: str
    entidad: str
    cuenta: str
    banco: str
    tipo_gasto: str = ""


@dataclass
class _FilaDisplay:
    tipo: str        # "cat2" | "subtotal" | "total"
    label: str
    cat1: str
    cat2: str
    mes_vals: list[Decimal]
    total: Decimal
    nav: bool


# ---------------------------------------------------------------------------
# Lectura de datos
# ---------------------------------------------------------------------------

def _leer_datos_wb(
    wb,
    meses_rango: list[tuple[int, str]],
    incluir_balance: bool = False,
    modo_balance: bool = False,
    modo_gastos: bool = False,
    ajuste_vivienda: bool = False,
    cuentas_filtro: set[str] | None = None,
) -> tuple[
    list[_FilaDisplay],
    list[int],
    dict[tuple[str, str], list[_Entrada]],
    dict[str, tuple[str, str]],
    dict[str, list[str]],
]:
    """Construye el estado de display a partir de un workbook ya abierto."""
    from presupuesto.escritor import leer_numero

    meses_set = set(meses_rango)

    # ── Claves: cuenta → (banco, tipo_cuenta) ──────────────────────────
    claves: dict[str, tuple[str, str]] = {}
    try:
        for row in wb["Claves"].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            c, b, tc = (str(row[i] or "").strip() for i in (0, 1, 2))
            if c:
                claves[c] = (b, tc)
    except KeyError:
        pass

    # ── Opciones desde Maestro ──────────────────────────────────────────
    opciones: dict[str, list[str]] = {
        "cat1": [], "cat2": [], "cat3": [], "entidad": [],
        "cuenta": sorted(claves.keys()),
        "tipo_gasto": [],
    }
    _maestro_cols = {2: "cat1", 3: "cat2", 4: "cat3", 5: "entidad", 7: "tipo_gasto"}
    try:
        seen: dict[str, set[str]] = {k: set() for k in _maestro_cols.values()}
        for row in wb["Maestro"].iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            for ci, campo in _maestro_cols.items():
                if ci < len(row) and row[ci] is not None:
                    val = str(row[ci]).strip()
                    if val and val not in seen[campo]:
                        seen[campo].add(val)
                        opciones[campo].append(val)
    except KeyError:
        pass

    # ── Datos (Presupuesto) ────────────────────────────────────────────
    datos: dict[tuple[str, str], dict[tuple[int, str], Decimal]] = defaultdict(
        lambda: defaultdict(Decimal)
    )
    detalles: dict[tuple[str, str], list[_Entrada]] = defaultdict(list)
    orden_visto: list[tuple[str, str]] = []
    visto: set[tuple[str, str]] = set()

    try:
        ws = wb["Datos"]
        for row in ws.iter_rows(min_row=2, values_only=False):
            if not row or row[_COL_AÑO].value is None:
                continue
            if str(row[_COL_ESTADO].value or "").strip() != "Presupuesto":
                continue
            cat1 = str(row[_COL_CAT1].value or "").strip()
            cat2 = str(row[_COL_CAT2].value or "").strip()
            es_balance = cat1.lower() == "finanzas" and cat2.lower() == "balance"

            if modo_balance:
                if not es_balance:
                    continue
            else:
                # modo_gastos siempre excluye Finanzas/Balance
                if es_balance:
                    continue
                if not incluir_balance and es_balance:
                    continue
                if not cat1:
                    continue

            try:
                año = int(row[_COL_AÑO].value)
            except (TypeError, ValueError):
                continue
            mes = str(row[_COL_MES].value or "").strip()
            if (año, mes) not in meses_set:
                continue
            imp_raw = leer_numero(row[_COL_IMPORTE].value)
            if imp_raw is None:
                continue
            importe    = Decimal(str(imp_raw))
            cat3       = str(row[_COL_CAT3].value       or "").strip()
            entidad    = str(row[_COL_ENTIDAD].value    or "").strip()
            cuenta     = str(row[_COL_CUENTA].value     or "").strip()
            if cuentas_filtro and cuenta not in cuentas_filtro:
                continue
            tipo_gasto = str(row[_COL_TIPO_GASTO].value or "").strip()
            banco      = claves.get(cuenta, ("", ""))[0]
            fila_n     = row[_COL_AÑO].row

            # Ajuste vivienda compartida: ocultar ingresos por transferencia de pareja
            # y dividir al 50% los gastos de Vivienda
            if ajuste_vivienda:
                if cat1.lower() == "vivienda" and "transferencia" in cat2.lower():
                    continue
                if cat1.lower() == "vivienda":
                    importe = importe / 2
                if cat1.lower() == "ahorro" and "hipoteca" in cat2.lower():
                    importe = importe / 2

            # En modo_balance agrupamos por cuenta; si no, por (cat1, cat2)
            key = ("", cuenta) if modo_balance else (cat1, cat2)
            datos[key][(año, mes)] += importe
            detalles[key].append(_Entrada(
                fila_xlsx=fila_n, año=año, mes=mes, importe=importe,
                cat1=cat1, cat2=cat2, cat3=cat3,
                entidad=entidad, cuenta=cuenta, banco=banco,
                tipo_gasto=tipo_gasto,
            ))
            if key not in visto:
                orden_visto.append(key)
                visto.add(key)
    except KeyError:
        pass

    def _sk(e: _Entrada) -> tuple[int, int]:
        return (e.año, _MESES_ORD.index(e.mes) if e.mes in _MESES_ORD else 0)

    for key in detalles:
        detalles[key].sort(key=_sk)

    # Suplementar tipo_gasto con valores vistos en las entradas
    tipos_en_datos = {e.tipo_gasto for ents in detalles.values() for e in ents if e.tipo_gasto}
    seen_tg = set(opciones["tipo_gasto"])
    for tg in sorted(tipos_en_datos):
        if tg not in seen_tg:
            opciones["tipo_gasto"].append(tg)

    # tipo_gasto representativo por (cat1, cat2): el más frecuente en las entradas
    cat12_tipo: dict[tuple[str, str], str] = {}
    for key, ents in detalles.items():
        from collections import Counter
        cnt: Counter = Counter(e.tipo_gasto for e in ents)
        cat12_tipo[key] = cnt.most_common(1)[0][0] if cnt else ""

    # ── Construir filas de display ──────────────────────────────────────
    filas: list[_FilaDisplay] = []
    nav_indices: list[int] = []
    total_mes = [Decimal(0)] * len(meses_rango)
    total_sum = Decimal(0)

    if modo_balance:
        # Una fila por cuenta, sin subtotales por cat1
        for key in orden_visto:
            _, cuenta = key
            mes_vals   = [datos[key].get(mk, Decimal(0)) for mk in meses_rango]
            fila_total = sum(mes_vals)
            if all(v == 0 for v in mes_vals):
                continue
            idx = len(filas)
            filas.append(_FilaDisplay(
                tipo="cat2", label=cuenta,
                cat1="", cat2=cuenta,
                mes_vals=mes_vals, total=fila_total, nav=True,
            ))
            nav_indices.append(idx)
            for i, v in enumerate(mes_vals):
                total_mes[i] += v
            total_sum += fila_total
    elif modo_gastos:
        # Secciones: Ingresos (primero), gastos por tipo_gasto, Ahorro (último).
        tg_order: list[str] = []
        tg_seen:  set[str]  = set()
        ahorro_keys:   list[tuple[str, str]] = []
        ingresos_keys: list[tuple[str, str]] = []

        for c1, c2 in orden_visto:
            cl = c1.lower()
            if cl == "ahorro":
                ahorro_keys.append((c1, c2))
                continue
            if cl == "ingresos":
                ingresos_keys.append((c1, c2))
                continue
            tg = cat12_tipo.get((c1, c2), "")
            if tg not in tg_seen:
                tg_order.append(tg)
                tg_seen.add(tg)

        total_sum_box   = [Decimal(0)]
        # Para las filas de porcentaje: lista de (label, sub_mes, sub_sum)
        seccion_totales: list[tuple[str, list[Decimal], Decimal]] = []
        ingresos_ref:    list                                      = [None]  # [list[Decimal], Decimal]

        def _build_grupo(
            keys: list[tuple[str, str]], label: str, *, es_ingresos: bool = False
        ) -> None:
            sub_mes = [Decimal(0)] * len(meses_rango)
            sub_sum = Decimal(0)
            for c1, c2 in keys:
                mes_vals   = [datos[(c1, c2)].get(mk, Decimal(0)) for mk in meses_rango]
                fila_total = sum(mes_vals)
                if all(v == 0 for v in mes_vals):
                    continue
                idx = len(filas)
                filas.append(_FilaDisplay(
                    tipo="cat2",
                    label=f"  {c1} / {c2}" if c2 else f"  {c1}",
                    cat1=c1, cat2=c2,
                    mes_vals=mes_vals, total=fila_total, nav=True,
                ))
                nav_indices.append(idx)
                for i, v in enumerate(mes_vals):
                    sub_mes[i] += v
                sub_sum += fila_total
            if sub_sum == 0 and all(v == 0 for v in sub_mes):
                return
            filas.append(_FilaDisplay(
                tipo="subtotal", label=label or "Sin tipo",
                cat1=label, cat2="",
                mes_vals=sub_mes, total=sub_sum, nav=False,
            ))
            for i, v in enumerate(sub_mes):
                total_mes[i] += v
            total_sum_box[0] += sub_sum
            if es_ingresos:
                ingresos_ref[0] = (list(sub_mes), sub_sum)
            else:
                seccion_totales.append((label or "Sin tipo", list(sub_mes), sub_sum))

        # 1. Ingresos
        if ingresos_keys:
            _build_grupo(ingresos_keys, "Ingresos", es_ingresos=True)

        # 2. Gastos por tipo
        for tg in tg_order:
            keys_tg = [(c1, c2) for c1, c2 in orden_visto
                       if cat12_tipo.get((c1, c2), "") == tg
                       and c1.lower() not in ("ahorro", "ingresos")]
            _build_grupo(keys_tg, tg)

        # 3. Ahorro
        if ahorro_keys:
            _build_grupo(ahorro_keys, "Ahorro")

        total_sum = total_sum_box[0]

        # 4. Filas de porcentaje sobre ingresos (tipo="pct")
        if ingresos_ref[0] is not None:
            i_mes, i_total = ingresos_ref[0]
            for (lbl, s_mes, s_total) in seccion_totales:
                pct_mes = [
                    round(abs(sv) / abs(iv) * Decimal(100), 1) if iv != 0 else Decimal(0)
                    for sv, iv in zip(s_mes, i_mes)
                ]
                pct_total = (
                    round(abs(s_total) / abs(i_total) * Decimal(100), 1)
                    if i_total != 0 else Decimal(0)
                )
                filas.append(_FilaDisplay(
                    tipo="pct", label=lbl,
                    cat1="", cat2="",
                    mes_vals=pct_mes, total=pct_total, nav=False,
                ))

    else:
        cat1_order: list[str] = []
        cat1_seen:  set[str]  = set()
        for c1, _ in orden_visto:
            if c1 not in cat1_seen:
                cat1_order.append(c1)
                cat1_seen.add(c1)

        for cat1 in cat1_order:
            sub_mes = [Decimal(0)] * len(meses_rango)
            sub_sum = Decimal(0)
            for c1, c2 in [(c1, c2) for c1, c2 in orden_visto if c1 == cat1]:
                mes_vals   = [datos[(c1, c2)].get(mk, Decimal(0)) for mk in meses_rango]
                fila_total = sum(mes_vals)
                if all(v == 0 for v in mes_vals):
                    continue
                idx = len(filas)
                filas.append(_FilaDisplay(
                    tipo="cat2",
                    label=f"  {c2}" if c2 else f"  ({c1})",
                    cat1=c1, cat2=c2,
                    mes_vals=mes_vals, total=fila_total, nav=True,
                ))
                nav_indices.append(idx)
                for i, v in enumerate(mes_vals):
                    sub_mes[i] += v
                sub_sum += fila_total

            if sub_sum == 0 and all(v == 0 for v in sub_mes):
                continue
            filas.append(_FilaDisplay(
                tipo="subtotal", label=cat1, cat1=cat1, cat2="",
                mes_vals=sub_mes, total=sub_sum, nav=False,
            ))
            for i, v in enumerate(sub_mes):
                total_mes[i] += v
            total_sum += sub_sum

    filas.append(_FilaDisplay(
        tipo="total", label="TOTAL", cat1="", cat2="",
        mes_vals=total_mes, total=total_sum, nav=False,
    ))

    return filas, nav_indices, dict(detalles), claves, opciones


def _leer_datos(
    ruta_xlsx: Path,
    meses_rango: list[tuple[int, str]],
    incluir_balance: bool = False,
    modo_balance: bool = False,
    modo_gastos: bool = False,
    ajuste_vivienda: bool = False,
    cuentas_filtro: set[str] | None = None,
) -> tuple[
    list[_FilaDisplay],
    list[int],
    dict[tuple[str, str], list[_Entrada]],
    dict[str, tuple[str, str]],
    dict[str, list[str]],
]:
    import openpyxl
    wb = openpyxl.load_workbook(str(ruta_xlsx), read_only=True)
    try:
        return _leer_datos_wb(wb, meses_rango, incluir_balance, modo_balance, modo_gastos,
                              ajuste_vivienda, cuentas_filtro)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Escritura
# ---------------------------------------------------------------------------

def _backup(ruta_xlsx: Path) -> None:
    """Crea un backup del xlsx con timestamp en el mismo directorio."""
    import shutil
    from datetime import datetime
    ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ruta_xlsx.parent / f"{ruta_xlsx.stem}_backup_{ts}{ruta_xlsx.suffix}"
    shutil.copy2(str(ruta_xlsx), str(backup))


def _insertar_entrada(
    wb,
    claves: dict[str, tuple[str, str]],
    campos: dict,
) -> None:
    """Añade una nueva fila al final de la hoja Datos en memoria (sin guardar a disco)."""
    from presupuesto.escritor import detectar_formulas_cuenta, adaptar_formula_fila

    ws = wb["Datos"]

    primera_libre = 2
    for r in range(ws.max_row, 1, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, 14)):
            primera_libre = r + 1
            break

    formula_k, formula_l = detectar_formulas_cuenta(ws, primera_libre)
    banco, tipo = claves.get(campos.get("cuenta", ""), ("", ""))
    fila = primera_libre

    ws.cell(fila,  1).value = campos["año"]
    ws.cell(fila,  2).value = campos["mes"]
    ws.cell(fila,  3).value = campos["cat1"]
    ws.cell(fila,  4).value = campos["cat2"]
    ws.cell(fila,  5).value = campos.get("cat3", "")
    ws.cell(fila,  6).value = campos.get("entidad", "")
    ws.cell(fila,  7).value = float(campos["importe"])
    ws.cell(fila,  8).value = ""
    ws.cell(fila,  9).value = campos.get("tipo_gasto", "")
    ws.cell(fila, 10).value = campos["cuenta"]
    ws.cell(fila, 11).value = (
        adaptar_formula_fila(formula_k, fila) if formula_k else banco
    )
    ws.cell(fila, 12).value = (
        adaptar_formula_fila(formula_l, fila) if formula_l else tipo
    )
    ws.cell(fila, 13).value = "Presupuesto"


def _eliminar_entradas(
    wb,
    filas_xlsx: list[int],
) -> None:
    """Elimina las filas indicadas de la hoja Datos en memoria (sin guardar a disco)."""
    ws = wb["Datos"]
    for fila_n in sorted(filas_xlsx, reverse=True):
        ws.delete_rows(fila_n, 1)


def _guardar_entradas(
    wb,
    claves: dict[str, tuple[str, str]],
    entries: list[tuple[int, dict]],
) -> None:
    """Escribe los cambios de varias filas en memoria (sin guardar a disco)."""
    from presupuesto.escritor import adaptar_formula_fila

    ws = wb["Datos"]

    for fila_n, campos in entries:
        ws.cell(fila_n, 1).value  = campos["año"]
        ws.cell(fila_n, 2).value  = campos["mes"]
        ws.cell(fila_n, 3).value  = campos["cat1"]
        ws.cell(fila_n, 4).value  = campos["cat2"]
        ws.cell(fila_n, 5).value  = campos["cat3"]
        ws.cell(fila_n, 6).value  = campos["entidad"]
        ws.cell(fila_n, 7).value  = float(campos["importe"])
        ws.cell(fila_n, 9).value  = campos.get("tipo_gasto", "")
        ws.cell(fila_n, 10).value = campos["cuenta"]

        banco, tipo = claves.get(campos["cuenta"], ("", ""))
        v_k = ws.cell(fila_n, 11).value
        v_l = ws.cell(fila_n, 12).value
        ws.cell(fila_n, 11).value = (
            adaptar_formula_fila(v_k, fila_n)
            if isinstance(v_k, str) and v_k.startswith("=") else banco
        )
        ws.cell(fila_n, 12).value = (
            adaptar_formula_fila(v_l, fila_n)
            if isinstance(v_l, str) and v_l.startswith("=") else tipo
        )


# ---------------------------------------------------------------------------
# Helpers de renderizado
# ---------------------------------------------------------------------------

_NUM_W = 7


def _fmt_num(v: Decimal, w: int = _NUM_W) -> tuple[str, str]:
    if v == 0:
        return ("class:zero", f"{'—':>{w}}")
    txt = f"{v:,.0f}"
    return ("class:neg" if v < 0 else "class:pos", f"{txt[:w]:>{w}}")


def _fmt_imp(v: Decimal, w: int = 11) -> tuple[str, str]:
    txt = f"{v:>+.2f}"
    return ("class:neg" if v < 0 else "class:pos", f"{txt:>{w}}")


def _clip(s: str, w: int) -> str:
    return s if len(s) <= w else s[:w - 1] + "…"


def _fila_lines(filas: list[_FilaDisplay]) -> tuple[list[int], int]:
    pos, ln = [], 0
    for f in filas:
        pos.append(ln)
        ln += 2 if f.tipo == "subtotal" else 1
    return pos, ln


# ---------------------------------------------------------------------------
# TUI multi-selección de cuentas
# ---------------------------------------------------------------------------

def _tui_seleccionar_cuentas(cuentas: list[str]) -> set[str] | None:
    """Multi-selección de cuentas. Devuelve el conjunto seleccionado o None si cancela."""
    from prompt_toolkit import Application
    from prompt_toolkit.application import get_app
    from prompt_toolkit.formatted_text import FormattedText
    from prompt_toolkit.key_binding import KeyBindings
    from prompt_toolkit.layout import Layout
    from prompt_toolkit.layout.containers import Window
    from prompt_toolkit.layout.controls import FormattedTextControl
    from prompt_toolkit.styles import Style

    style = Style.from_dict({
        "titulo":  "bold",
        "cursor":  "reverse bold",
        "selec":   "bold #00cc44",
        "cur-sel": "reverse bold #00cc44",
        "filtro":  "bold yellow",
        "dim":     "#666666",
        "footer":  "#666666",
        "fkey":    "#aaaaaa bold",
    })

    state: dict = {"cursor": 0, "selec": set(), "filtro": "", "resultado": None}

    def _filtradas() -> list[str]:
        f = state["filtro"].lower()
        return [c for c in cuentas if f in c.lower()] if f else list(cuentas)

    def _clamp() -> None:
        vis = _filtradas()
        state["cursor"] = max(0, min(state["cursor"], len(vis) - 1))

    def _render() -> FormattedText:
        try:
            size = get_app().output.get_size()
            w, h = size.columns, size.rows
        except Exception:
            w, h = 80, 30

        buf: list[tuple[str, str]] = []
        def t(st, s): buf.append((st, s))
        def nl():     buf.append(("", "\n"))

        t("class:titulo", "  Filtrar por cuenta(s)")
        nl()
        filtro_txt = (state["filtro"] + "▌") if state["filtro"] else "▌"
        t("class:dim",    "  Filtro: ")
        t("class:filtro", filtro_txt)
        t("class:dim", f"   {len(state['selec'])} seleccionada(s)")
        nl()
        t("class:dim", "─" * w)
        nl()

        vis      = _filtradas()
        list_h   = max(3, h - 8)
        cur      = state["cursor"]
        ws_start = max(0, cur - list_h // 2)
        ws_end   = min(len(vis), ws_start + list_h)
        ws_start = max(0, ws_end - list_h)

        for i in range(ws_start, ws_end):
            c = vis[i]
            es_cur = i == cur
            es_sel = c in state["selec"]
            marca  = "[✓]" if es_sel else "[ ]"
            txt    = f"  {marca} {c}"

            if es_cur and es_sel:
                t("class:cur-sel", txt)
            elif es_cur:
                t("class:cursor",  txt)
            elif es_sel:
                t("class:selec",   txt)
            else:
                t("",              txt)
            nl()

        t("class:dim", "─" * w)
        nl()
        for k, desc in [("↑↓", "Navegar"), ("Spc", "Seleccionar"), ("a", "Todas"),
                         ("^U", "Borrar filtro"), ("Enter", "Confirmar"), ("Esc", "Sin filtro")]:
            t("class:fkey",   f" {k} ")
            t("class:footer", f"{desc}  ")

        return FormattedText(buf)

    kb = KeyBindings()

    @kb.add("up")
    def _(e): state["cursor"] = max(0, state["cursor"] - 1)

    @kb.add("down")
    def _(e): state["cursor"] = min(max(0, len(_filtradas()) - 1), state["cursor"] + 1)

    @kb.add("space")
    def _(e):
        vis = _filtradas()
        if vis:
            c = vis[state["cursor"]]
            if c in state["selec"]:
                state["selec"].discard(c)
            else:
                state["selec"].add(c)

    @kb.add("a")
    def _(e):
        vis = _filtradas()
        if state["selec"] >= set(vis):
            state["selec"] -= set(vis)
        else:
            state["selec"] |= set(vis)

    @kb.add("enter")
    def _(e):
        state["resultado"] = set(state["selec"]) or None
        e.app.exit()

    @kb.add("escape")
    @kb.add("c-c")
    def _(e):
        state["resultado"] = "cancelar"
        e.app.exit()

    @kb.add("c-u")
    def _(e):
        state["filtro"] = ""
        _clamp()

    @kb.add("backspace")
    @kb.add("c-h")
    def _(e):
        state["filtro"] = state["filtro"][:-1]
        _clamp()

    @kb.add("<any>")
    def _(e):
        key = e.key_sequence[0].key
        if isinstance(key, str) and len(key) == 1 and key.isprintable():
            state["filtro"] += key
            state["cursor"] = 0

    app = Application(
        layout=Layout(Window(content=FormattedTextControl(text=_render, focusable=True))),
        key_bindings=kb,
        style=style,
        full_screen=True,
    )
    app.run()

    resultado = state["resultado"]
    if resultado == "cancelar":
        return None   # señal de cancelación total
    return resultado  # set de cuentas o None (sin filtro)


# ---------------------------------------------------------------------------
# TUI
# ---------------------------------------------------------------------------

def _tui_vista(
    filas: list[_FilaDisplay],
    nav_indices: list[int],
    meses_rango: list[tuple[int, str]],
    detalles: dict[tuple[str, str], list[_Entrada]],
    ruta_xlsx: Path,
    claves: dict[str, tuple[str, str]],
    opciones: dict[str, list[str]],
    ruta_origen: Path | None = None,
    incluir_balance: bool = False,
    modo_balance: bool = False,
    modo_gastos: bool = False,
    ajuste_vivienda: bool = False,
) -> None:

    import openpyxl
    import threading

    # Workbook compartido para todas las lecturas/escrituras de la sesión.
    # Se abre una sola vez → elimina el load_workbook de cada save.
    wb_live = openpyxl.load_workbook(str(ruta_xlsx))
    from prompt_toolkit import Application
    from prompt_toolkit.application import get_app
    from prompt_toolkit.filters import Condition
    from prompt_toolkit.formatted_text import FormattedText
    from prompt_toolkit.key_binding import KeyBindings
    from prompt_toolkit.layout import Layout
    from prompt_toolkit.layout.containers import Window
    from prompt_toolkit.layout.controls import FormattedTextControl
    from prompt_toolkit.styles import Style

    style = Style.from_dict({
        "titulo":    "bold",
        "header":    "bold #888888",
        "sep":       "#444444",
        "cat1":      "bold",
        "cursor":    "reverse bold",
        "sel":       "bold #55ff55",
        "dim":       "#666666",
        "neg":       "#ff5555",
        "pos":       "#55ff55",
        "zero":      "#555555",
        "footer":    "#666666",
        "fkey":      "#aaaaaa bold",
        "total":     "bold",
        "detail_h":  "bold #aaaaaa",
        "field_lbl": "#aaaaaa",
        "field_cur": "reverse",
        "field_mod": "bold #ffff55",   # campo modificado
        "field_ro":  "#555555",
        "err":       "bold #ff5555",
        "ok":        "bold #55ff55",
        "input_val": "bold",
        "filtro":    "bold yellow",
    })

    n_meses = len(meses_rango)

    state: dict = {
        "view":       "main",
        # main
        "cursor":     0,
        "m_offset":   0,
        # detail
        "d_cat1":     "",
        "d_cat2":     "",
        "d_sel":      0,
        "d_offset":   0,
        "d_selected": set(),   # set de fila_xlsx seleccionados
        # edit
        "e_campos":   {},
        "e_original": {},
        "e_fila":     0,
        "e_cursor":   0,
        "e_msg":      "",
        # dup_meses (selección de meses para duplicar o nueva entrada)
        "dup_campos":   {},
        "dup_cursor":   0,
        "dup_selected": set(),   # set de índices en meses_rango
        "e_nueva":      False,   # True cuando el edit es para insertar nueva entrada
        # saving / syncing
        "saving":        False,
        "syncing":       False,
        # confirm (eliminar)
        "confirm_filas": [],   # list[int] de fila_xlsx a eliminar
        "confirm_msg":   "",   # texto descriptivo
        # input (texto libre)
        "i_campo":    "",
        "i_buf":      "",
        "i_error":    "",
        # picker
        "p_campo":    "",
        "p_options":  [],
        "p_cursor":   0,
        "p_filter":   "",
    }

    # ── filtros de vista para key bindings ───────────────────────────
    # Impiden que teclas de texto sean interceptadas en modos de edición
    _no_input        = Condition(lambda: state["view"] != "input")
    _no_text         = Condition(lambda: state["view"] not in ("input", "picker"))
    _only_main       = Condition(lambda: state["view"] == "main")
    _only_detail     = Condition(lambda: state["view"] == "detail")
    _only_confirm    = Condition(lambda: state["view"] == "confirm")
    _only_dup        = Condition(lambda: state["view"] == "dup_meses")
    _edit_or_confirm = Condition(lambda: state["view"] in ("edit", "confirm", "dup_meses"))

    # ── disk write background ─────────────────────────────────────────

    _save_thread: list = [None]   # [0] = último hilo de escritura a disco

    def _disk_write(app_ref) -> None:
        """Hace backup + wb.save() + sync a origen en segundo plano."""
        def _do() -> None:
            import shutil as _shutil
            state["syncing"] = True
            app_ref.invalidate()
            try:
                _backup(ruta_xlsx)
                wb_live.save(str(ruta_xlsx))
                if ruta_origen is not None and ruta_origen != ruta_xlsx:
                    _shutil.copy2(str(ruta_xlsx), str(ruta_origen))
            finally:
                state["syncing"] = False
                app_ref.invalidate()

        t = threading.Thread(target=_do, daemon=True)
        _save_thread[0] = t
        t.start()

    # ── helpers ───────────────────────────────────────────────────────

    def _recargar() -> None:
        nf, nn, nd, nc, nop = _leer_datos_wb(wb_live, meses_rango, incluir_balance, modo_balance, modo_gastos, ajuste_vivienda)
        filas[:] = nf
        nav_indices[:] = nn
        detalles.clear();  detalles.update(nd)
        claves.clear();    claves.update(nc)
        opciones.clear();  opciones.update(nop)

    def _entradas() -> list[_Entrada]:
        return detalles.get((state["d_cat1"], state["d_cat2"]), [])

    def _d_clamp(list_h: int) -> None:
        n = max(0, len(_entradas()) - 1)
        state["d_sel"]    = max(0, min(state["d_sel"], n))
        if state["d_sel"] < state["d_offset"]:
            state["d_offset"] = state["d_sel"]
        elif state["d_sel"] >= state["d_offset"] + list_h:
            state["d_offset"] = state["d_sel"] - list_h + 1
        state["d_offset"] = max(0, state["d_offset"])

    def _modified() -> set[str]:
        orig = state["e_original"]
        return {k for k, _, _ in _CAMPOS
                if str(state["e_campos"].get(k, "")) != str(orig.get(k, ""))}

    def _abrir_detalle(cat1: str, cat2: str) -> None:
        state.update(d_cat1=cat1, d_cat2=cat2,
                     d_sel=0, d_offset=0, view="detail")

    def _abrir_edit() -> None:
        ents = _entradas()
        if not ents:
            return
        e = ents[state["d_sel"]]
        campos = {
            "año": e.año, "mes": e.mes, "importe": e.importe,
            "cat1": e.cat1, "cat2": e.cat2, "cat3": e.cat3,
            "entidad": e.entidad, "cuenta": e.cuenta,
            "tipo_gasto": e.tipo_gasto,
        }
        state.update(e_campos=campos, e_original=dict(campos),
                     e_fila=e.fila_xlsx, e_cursor=0, e_msg="",
                     view="edit")

    def _abrir_dup() -> None:
        ents = _entradas()
        if not ents:
            return
        e = ents[state["d_sel"]]
        campos = {
            "año": e.año, "mes": e.mes, "importe": e.importe,
            "cat1": e.cat1, "cat2": e.cat2, "cat3": e.cat3,
            "entidad": e.entidad, "cuenta": e.cuenta,
            "tipo_gasto": e.tipo_gasto,
        }
        state.update(dup_campos=campos, dup_cursor=0,
                     dup_selected=set(), view="dup_meses")

    def _abrir_nueva() -> None:
        año_def, mes_def = meses_rango[0]
        if modo_balance:
            # Pre-rellena cat1/cat2 y sugiere la cuenta del cursor actual
            cuenta_def = ""
            if nav_indices:
                cuenta_def = filas[nav_indices[state["cursor"]]].cat2
            campos = {
                "año": año_def, "mes": mes_def, "importe": Decimal("0"),
                "cat1": "Finanzas", "cat2": "Balance",
                "cat3": "", "entidad": "", "cuenta": cuenta_def, "tipo_gasto": "",
            }
            e_cursor_def = 2   # Importe
        else:
            campos = {
                "año": año_def, "mes": mes_def, "importe": Decimal("0"),
                "cat1": "", "cat2": "", "cat3": "", "entidad": "", "cuenta": "",
                "tipo_gasto": "",
            }
            e_cursor_def = 2
        state.update(e_campos=campos, e_original=dict(campos),
                     e_fila=-1, e_cursor=e_cursor_def,
                     e_msg="", e_nueva=True, view="edit")

    def _abrir_input() -> None:
        campo = _CAMPOS[state["e_cursor"]][0]
        state.update(i_campo=campo,
                     i_buf=str(state["e_campos"][campo]),
                     i_error="", view="input")

    def _abrir_picker() -> None:
        campo = _CAMPOS[state["e_cursor"]][0]
        state.update(p_campo=campo,
                     p_options=list(opciones.get(campo, [])),
                     p_cursor=0, p_filter="", view="picker")

    def _p_filtered() -> list[str]:
        f = state["p_filter"].lower()
        return [o for o in state["p_options"] if f in o.lower()] if f else state["p_options"]

    def _validar_input() -> object | None:
        campo = state["i_campo"]
        raw   = state["i_buf"].strip()
        if campo == "año":
            try:
                return int(raw)
            except ValueError:
                state["i_error"] = "Debe ser un entero (ej: 2026)"
                return None
        if campo == "mes":
            m = next((m for m in _MESES_ORD if m.lower() == raw.lower()), None)
            if m is None:
                state["i_error"] = f"Válidos: {', '.join(_MESES_ORD)}"
                return None
            return m
        if campo == "importe":
            try:
                return Decimal(raw.replace(",", "."))
            except InvalidOperation:
                state["i_error"] = "Importe no válido (ej: -431.25)"
                return None
        return raw

    def _do_save() -> None:
        new_cat1 = state["e_campos"].get("cat1", state["d_cat1"])
        new_cat2 = state["e_campos"].get("cat2", state["d_cat2"])

        mod     = _modified()
        ents    = _entradas()
        sel_set = state["d_selected"]

        if sel_set:
            to_save = []
            for e in ents:
                if e.fila_xlsx not in sel_set:
                    continue
                campos = {
                    "año": e.año, "mes": e.mes, "importe": e.importe,
                    "cat1": e.cat1, "cat2": e.cat2, "cat3": e.cat3,
                    "entidad": e.entidad, "cuenta": e.cuenta,
                    "tipo_gasto": e.tipo_gasto,
                }
                for k in mod:
                    campos[k] = state["e_campos"][k]
                to_save.append((e.fila_xlsx, campos))
        else:
            to_save = [(state["e_fila"], dict(state["e_campos"]))]

        _guardar_entradas(wb_live, claves, to_save)
        _recargar()
        state["d_selected"].clear()
        state["d_cat1"] = new_cat1
        state["d_cat2"] = new_cat2
        state["d_sel"]  = max(0, state["d_sel"] - 1)
        state["e_msg"]  = ""
        state["view"]   = "detail"

    # ── render principal ──────────────────────────────────────────────

    def _render_main(w: int, h: int) -> list[tuple[str, str]]:
        buf: list[tuple[str, str]] = []
        t  = lambda st, s: buf.append((st, s))
        nl = lambda: buf.append(("", "\n"))

        titulo_main = (
            "  Balance de cuentas — Finanzas/Balance" if modo_balance else
            "  Presupuesto por tipo de gasto"          if modo_gastos  else
            "  Presupuesto — próximos 12 meses"
        )
        t("class:titulo", titulo_main); nl()
        t("class:sep", "─" * w); nl()

        lbl_w     = 36
        tot_w     = _NUM_W + 2
        available = max(0, w - lbl_w - tot_w - 3)
        n_vis     = min(n_meses, max(1, available // (_NUM_W + 1)))
        mes_slice = list(range(n_vis))

        hdr_lbl = "Cuenta" if modo_balance else "Categoría"
        t("class:header", f"  {hdr_lbl:<{lbl_w - 2}}")
        for mi in mes_slice:
            año, mes = meses_rango[mi]
            t("class:header", f" {mes + ' ' + str(año)[2:]:>{_NUM_W}}")
        if n_meses > n_vis:
            t("class:dim", f" +{n_meses - n_vis}m")
        t("class:header", f"  {'Total':>{_NUM_W}}"); nl()
        t("class:sep", "─" * w); nl()

        # Fila total y filas pct: siempre ancladas al fondo, fuera del scroll
        fila_total   = next((f for f in filas if f.tipo == "total"), None)
        filas_pct    = [f for f in filas if f.tipo == "pct"]
        filas_scroll = [f for f in filas if f.tipo not in ("total", "pct")]

        n_pct_lines = len(filas_pct) + (2 if filas_pct else 0)  # header + rows + sep
        list_h = max(1, h - 9 - n_pct_lines)
        fpos, _ = _fila_lines(filas_scroll)
        cur_fila_idx = nav_indices[state["cursor"]] if nav_indices else 0
        # nav_indices index filas; filas_scroll is filas minus last → same indices
        cur_line = fpos[cur_fila_idx] if cur_fila_idx < len(fpos) else 0
        offset   = state["m_offset"]
        if cur_line < offset:
            offset = cur_line
        elif cur_line >= offset + list_h:
            offset = cur_line - list_h + 1
        state["m_offset"] = max(0, offset)

        for fi, fila in enumerate(filas_scroll):
            fl      = fpos[fi]
            n_lines = 2 if fila.tipo == "subtotal" else 1
            if fl + n_lines <= offset: continue
            if fl >= offset + list_h:  break

            es_cur = fila.nav and nav_indices[state["cursor"]] == fi
            arrow  = "►" if es_cur else " "

            if fila.tipo == "subtotal":
                row_st, lbl = "class:cat1", f"{_clip(fila.label, lbl_w):<{lbl_w}}"
            else:
                row_st = "class:cursor" if es_cur else ""
                lbl    = f"{_clip(fila.label, lbl_w):<{lbl_w}}"

            if fl >= offset:
                t("class:dim", f" {arrow}"); t(row_st, lbl)
                for mi in mes_slice:
                    st, txt = _fmt_num(fila.mes_vals[mi])
                    t(row_st if es_cur else st, f" {txt}")
                st_t, txt_t = _fmt_num(fila.total, _NUM_W + 1)
                t(row_st if es_cur else st_t, f"  {txt_t}"); nl()

            if fila.tipo == "subtotal" and fl + 1 >= offset and fl + 1 < offset + list_h:
                t("class:sep", "─" * w); nl()

        # Total anclado: separador + fila
        t("class:sep", "─" * w); nl()
        if fila_total:
            t("class:dim", "  "); t("class:total", f"{'TOTAL':<{lbl_w}}")  # fijo, no se trunca
            for mi in mes_slice:
                st, txt = _fmt_num(fila_total.mes_vals[mi])
                t(st, f" {txt}")
            st_t, txt_t = _fmt_num(fila_total.total, _NUM_W + 1)
            t(st_t, f"  {txt_t}"); nl()

        # Filas de porcentaje sobre ingresos (solo modo_gastos)
        if filas_pct:
            t("class:sep", "─" * w); nl()
            t("class:dim", f"  {'% sobre ingresos':<{lbl_w}.{lbl_w}}")
            for mi in mes_slice:
                t("class:dim", f" {'':>{_NUM_W}}")
            t("class:dim", f"  {'':>{_NUM_W + 1}}"); nl()
            for pf in filas_pct:
                t("class:dim", f"  {_clip(pf.label, lbl_w):<{lbl_w}}")
                for mi in mes_slice:
                    v = pf.mes_vals[mi]
                    txt = f"{v:.1f}%" if v > 0 else "—"
                    t("class:dim", f" {txt:>{_NUM_W}}")
                tot_txt = f"{pf.total:.1f}%" if pf.total > 0 else "—"
                t("class:dim", f"  {tot_txt:>{_NUM_W + 1}}"); nl()

        t("class:sep", "─" * w); nl()
        if ajuste_vivienda:
            t("class:dim",
              "  * Vivienda ÷2 y Ahorro/Hipoteca ÷2 (gastos compartidos). "
              "Vivienda/Transferencia oculto. "
              "Usa --sin-ajuste para ver sin filtros."); nl()
        for k, d in [("j/k","Mover"),("gg/G","Ini/Fin"),
                     ("^d/^u","Med pág"),("Enter","Detalle"),("n","Nueva"),("q","Salir")]:
            t("class:fkey", f" {k} "); t("class:footer", f"{d}  ")
        if n_meses > n_vis:
            t("class:footer", f" │ {n_vis}/{n_meses} meses")
        return buf

    # ── render detalle ─────────────────────────────────────────────────

    def _render_detail(w: int, h: int) -> list[tuple[str, str]]:
        buf: list[tuple[str, str]] = []
        t  = lambda st, s: buf.append((st, s))
        nl = lambda: buf.append(("", "\n"))

        cat1, cat2 = state["d_cat1"], state["d_cat2"]
        sel_set    = state["d_selected"]
        n_sel      = len(sel_set)

        titulo = f"  Detalle: {cat1} / {cat2}" if cat2 else f"  Detalle: {cat1}"
        if n_sel:
            titulo += f"  [{n_sel} seleccionada{'s' if n_sel > 1 else ''}]"
        t("class:titulo", titulo); nl()
        t("class:sep", "─" * w); nl()

        W_F, W_I, W_C3, W_EN, W_CU, W_BA = 9, 11, 13, 13, 20, 14
        t("class:detail_h",
          f"    {'Fecha':<{W_F}} {'Importe':>{W_I}}  "
          f"{'Cat3':<{W_C3}}  {'Entidad':<{W_EN}}  "
          f"{'Cuenta':<{W_CU}}  {'Banco':<{W_BA}}")
        nl()
        t("class:sep", "─" * w); nl()

        entradas = _entradas()
        list_h   = max(1, h - 9)
        _d_clamp(list_h)
        offset   = state["d_offset"]

        for idx, e in enumerate(entradas[offset: offset + list_h]):
            abs_idx  = idx + offset
            es_cur   = abs_idx == state["d_sel"]
            es_sel   = e.fila_xlsx in sel_set
            fecha    = f"{e.mes} {e.año}"
            st, imp  = _fmt_imp(e.importe, W_I)

            arrow    = "►" if es_cur else " "
            chk      = "✓" if es_sel else " "
            chk_st   = "class:sel" if es_sel else "class:dim"
            row_st   = "class:cursor" if es_cur else "class:dim"

            t("class:dim",  f" {arrow}")
            t(chk_st,       f"{chk} ")
            t(row_st,       f"{fecha:<{W_F}} ")
            t("class:cursor" if es_cur else st, imp)
            t(row_st,
              f"  {_clip(e.cat3, W_C3):<{W_C3}}  "
              f"{_clip(e.entidad, W_EN):<{W_EN}}  "
              f"{_clip(e.cuenta, W_CU):<{W_CU}}  "
              f"{_clip(e.banco, W_BA):<{W_BA}}")
            nl()

        total = sum(e.importe for e in entradas)
        t("class:sep", "─" * w); nl()
        st_t, tot = _fmt_imp(total, W_I)
        t("class:dim", f"  {'Total':<{W_F}} "); t(st_t, tot)
        t("class:dim",
          f"  {len(entradas)} entradas"
          f"  [{offset+1}–{min(offset+list_h, len(entradas))} de {len(entradas)}]")
        nl()
        t("class:sep", "─" * w); nl()
        for k, d in [("j/k","Mover"),("Esp","Selec"),("a","Todos"),
                     ("Enter","Editar"),("d","Duplicar"),("x","Eliminar"),("q/Esc","Volver")]:
            t("class:fkey", f" {k} "); t("class:footer", f"{d}  ")
        return buf

    # ── render edición ─────────────────────────────────────────────────

    def _render_edit(w: int, h: int) -> list[tuple[str, str]]:
        buf: list[tuple[str, str]] = []
        t  = lambda st, s: buf.append((st, s))
        nl = lambda: buf.append(("", "\n"))

        campos  = state["e_campos"]
        sel_set = state["d_selected"]
        n_sel   = len(sel_set)
        mod     = _modified()
        banco   = claves.get(campos.get("cuenta", ""), ("", ""))[0]

        if state["e_nueva"]:
            titulo = "  Nueva entrada  [rellena los campos y pulsa s para seleccionar meses]"
        else:
            titulo = "  Editar entrada"
            if n_sel:
                titulo += f"  [{n_sel} seleccionada{'s' if n_sel > 1 else ''} · solo campos ★ se aplicarán a todas]"
        t("class:titulo", titulo); nl()
        t("class:sep", "─" * w); nl()

        LBL_W = 10
        VAL_W = max(20, w - LBL_W - 10)

        for ci, (clave, etiqueta, usa_picker) in enumerate(_CAMPOS):
            es_cur  = ci == state["e_cursor"]
            val     = str(campos.get(clave, ""))
            es_mod  = clave in mod
            arrow   = "►" if es_cur else " "
            marca   = "★ " if (es_mod and n_sel) else "  "
            lbl_st  = "class:field_cur" if es_cur else (
                       "class:field_mod" if es_mod else "class:field_lbl")
            val_st  = "class:cursor" if es_cur else (
                       "class:field_mod" if es_mod else "")
            picker_hint = " [↵=lista]" if usa_picker and es_cur else ""
            t("class:dim",   f"  {arrow} ")
            t("class:sel",    marca)
            t(lbl_st,        f"{etiqueta:<{LBL_W}}")
            t(val_st,        f"  {_clip(val, VAL_W)}")
            t("class:dim",    picker_hint)
            nl()

        t("class:field_ro", f"      {'Banco':<{LBL_W}}  {banco}"); nl()
        nl()

        msg = state["e_msg"]
        if msg.startswith("ok:"):
            t("class:ok",  f"  ✓ {msg[3:]}"); nl()
        elif msg.startswith("err:"):
            t("class:err", f"  ✗ {msg[4:]}"); nl()

        t("class:sep", "─" * w); nl()
        for k, d in [("j/k","Campo"),("Enter","Editar"),("s","Guardar"),("Esc/q","Cancelar")]:
            t("class:fkey", f" {k} "); t("class:footer", f"{d}  ")
        return buf

    # ── render input ───────────────────────────────────────────────────

    def _render_input(w: int, h: int) -> list[tuple[str, str]]:
        buf: list[tuple[str, str]] = []
        t  = lambda st, s: buf.append((st, s))
        nl = lambda: buf.append(("", "\n"))

        campo = state["i_campo"]
        etiq  = next((e for c, e, _ in _CAMPOS if c == campo), campo)
        t("class:titulo", f"  Editando: {etiq}"); nl()
        t("class:sep", "─" * w); nl(); nl()

        if campo == "mes":
            t("class:dim", f"  Válidos: {', '.join(_MESES_ORD)}"); nl()
        elif campo == "importe":
            t("class:dim",  "  Formato: -431.25"); nl()
        nl()

        t("class:field_lbl", f"  {etiq}: ")
        t("class:input_val", state["i_buf"])
        t("class:input_val", "▌"); nl(); nl()

        if state["i_error"]:
            t("class:err", f"  ✗ {state['i_error']}"); nl()

        t("class:sep", "─" * w); nl()
        for k, d in [("Enter","Confirmar"),("Esc","Cancelar"),("Bksp","Borrar")]:
            t("class:fkey", f" {k} "); t("class:footer", f"{d}  ")
        return buf

    # ── render picker ──────────────────────────────────────────────────

    def _render_picker(w: int, h: int) -> list[tuple[str, str]]:
        buf: list[tuple[str, str]] = []
        t  = lambda st, s: buf.append((st, s))
        nl = lambda: buf.append(("", "\n"))

        campo  = state["p_campo"]
        etiq   = next((e for c, e, _ in _CAMPOS if c == campo), campo)
        filtro = state["p_filter"]

        t("class:titulo", f"  Seleccionar {etiq}"); nl()
        filtro_txt = (filtro + "▌") if filtro else "▌"
        t("class:dim",    "  Filtro: ")
        t("class:filtro", filtro_txt); nl()
        t("class:sep", "─" * w); nl()

        opts = _p_filtered()
        n    = len(opts)
        list_h = max(1, h - 6)
        cur    = max(0, min(state["p_cursor"], n - 1))
        state["p_cursor"] = cur

        offset = max(0, min(cur - list_h // 2, max(0, n - list_h)))
        for i in range(offset, min(n, offset + list_h)):
            es_cur = i == cur
            arrow  = "►" if es_cur else " "
            row_st = "class:cursor" if es_cur else "class:dim"
            t("class:dim", f"  {arrow} ")
            t(row_st, opts[i]); nl()

        t("class:sep", "─" * w); nl()
        for k, d in [("j/k","Mover"),("Enter","Selec"),("^U","Limpiar"),("Esc","Cancelar")]:
            t("class:fkey", f" {k} "); t("class:footer", f"{d}  ")
        return buf

    # ── render confirmar eliminación ──────────────────────────────────

    def _render_confirm(w: int, h: int) -> list[tuple[str, str]]:
        buf: list[tuple[str, str]] = []
        t  = lambda st, s: buf.append((st, s))
        nl = lambda: buf.append(("", "\n"))

        filas_del = state["confirm_filas"]
        n = len(filas_del)
        t("class:titulo", f"  Eliminar {n} entrada{'s' if n > 1 else ''}"); nl()
        t("class:sep", "─" * w); nl(); nl()

        # Mostrar las entradas que se van a eliminar
        ents = _entradas()
        fx_set = set(filas_del)
        W_F, W_I, W_C1, W_C2 = 9, 11, 18, 18
        t("class:detail_h",
          f"  {'Fecha':<{W_F}} {'Importe':>{W_I}}  {'Cat1':<{W_C1}}  {'Cat2':<{W_C2}}"); nl()
        t("class:sep", "─" * w); nl()
        for e in ents:
            if e.fila_xlsx not in fx_set:
                continue
            fecha = f"{e.mes} {e.año}"
            st, imp = _fmt_imp(e.importe, W_I)
            t("class:dim", f"  {fecha:<{W_F}} ")
            t(st, imp)
            t("class:err", f"  {_clip(e.cat1, W_C1):<{W_C1}}  {_clip(e.cat2, W_C2):<{W_C2}}"); nl()

        nl()
        t("class:err",    "  ⚠ Esta acción no se puede deshacer (se crea backup automático)."); nl()
        nl()
        t("class:sep", "─" * w); nl()
        for k, d in [("s/y","Confirmar eliminar"),("n/Esc","Cancelar")]:
            t("class:fkey", f" {k} "); t("class:footer", f"{d}  ")
        return buf

    # ── render selección de meses para duplicar ───────────────────────

    def _render_dup_meses(w: int, h: int) -> list[tuple[str, str]]:
        buf: list[tuple[str, str]] = []
        t  = lambda st, s: buf.append((st, s))
        nl = lambda: buf.append(("", "\n"))

        campos  = state["dup_campos"]
        sel_set = state["dup_selected"]
        cursor  = state["dup_cursor"]
        n_sel   = len(sel_set)

        t("class:titulo", "  Duplicar entrada — selecciona meses de destino"); nl()
        t("class:sep", "─" * w); nl()

        # Info de la entrada original
        cat_txt = " / ".join(filter(None, [campos.get("cat1", ""),
                                           campos.get("cat2", ""),
                                           campos.get("cat3", "")]))
        t("class:dim", f"  {cat_txt}")
        imp = campos.get("importe", 0)
        st_i, txt_i = _fmt_imp(Decimal(str(imp)))
        t("class:dim", "   "); t(st_i, txt_i); nl()
        t("class:dim", f"  Cuenta: {campos.get('cuenta','')}   "
                       f"Origen: {campos.get('mes','')} {campos.get('año','')}"); nl()
        t("class:sep", "─" * w); nl()

        list_h  = max(1, h - 11)
        n       = len(meses_rango)
        ws_start = max(0, min(cursor - list_h // 2, n - list_h))
        ws_end   = min(n, ws_start + list_h)
        ws_start = max(0, ws_end - list_h)

        for i in range(ws_start, ws_end):
            año, mes = meses_rango[i]
            es_cur = i == cursor
            es_sel = i in sel_set
            arrow  = "►" if es_cur else " "
            check  = "✓" if es_sel else " "
            row_st = "class:cursor" if es_cur else ""
            chk_st = "class:sel" if es_sel else "class:dim"
            t("class:dim", f"  {arrow} ")
            t(chk_st, f"[{check}] ")
            t(row_st, f"{mes} {año}"); nl()

        t("class:sep", "─" * w); nl()
        if n_sel:
            t("class:ok", f"  {n_sel} mes{'es' if n_sel != 1 else ''} seleccionado{'s' if n_sel != 1 else ''}  ")
        else:
            t("class:dim", "  Sin meses seleccionados  ")
        nl()
        for k, d in [("j/k","Mover"),("Esp","Marcar"),("a","Todos"),
                     ("s","Insertar"),("Esc","Cancelar")]:
            t("class:fkey", f" {k} "); t("class:footer", f"{d}  ")
        return buf

    # ── render guardando ──────────────────────────────────────────────

    def _render_saving(w: int, h: int) -> list[tuple[str, str]]:
        buf: list[tuple[str, str]] = []
        t  = lambda st, s: buf.append((st, s))
        nl = lambda: buf.append(("", "\n"))
        for _ in range(max(0, h // 2 - 2)):
            nl()
        t("class:titulo", "  ⏳ Guardando…"); nl()
        t("class:dim",    "  Escribiendo en el xlsx, por favor espera."); nl()
        return buf

    # ── dispatcher ────────────────────────────────────────────────────

    def _render() -> FormattedText:
        try:
            size = get_app().output.get_size()
            w, h = size.columns, size.rows
        except Exception:
            w, h = 120, 40
        view = state["view"]
        if state["saving"]:
            return FormattedText(_render_saving(w, h))
        fn   = {"main": _render_main, "detail": _render_detail,
                "edit": _render_edit, "input": _render_input,
                "picker": _render_picker, "confirm": _render_confirm,
                "dup_meses": _render_dup_meses}
        buf = fn[view](w, h)
        if state["syncing"]:
            buf.append(("class:dim", "\n  ↑ Sincronizando…"))
        return FormattedText(buf)

    # ── key bindings ──────────────────────────────────────────────────

    kb = KeyBindings()

    # j/k activos en todos los modos excepto input (donde deben escribirse)
    @kb.add("j", filter=_no_input)
    @kb.add("down")
    def _mv_down(e):
        v = state["view"]
        if v == "main":
            state["cursor"] = min(max(0, len(nav_indices) - 1), state["cursor"] + 1)
        elif v == "detail":
            state["d_sel"] = min(max(0, len(_entradas()) - 1), state["d_sel"] + 1)
        elif v == "edit":
            state["e_cursor"] = (state["e_cursor"] + 1) % len(_CAMPOS)
            state["e_msg"] = ""
        elif v == "picker":
            state["p_cursor"] += 1
        elif v == "dup_meses":
            state["dup_cursor"] = min(len(meses_rango) - 1, state["dup_cursor"] + 1)

    @kb.add("k", filter=_no_input)
    @kb.add("up")
    def _mv_up(e):
        v = state["view"]
        if v == "main":
            state["cursor"] = max(0, state["cursor"] - 1)
        elif v == "detail":
            state["d_sel"] = max(0, state["d_sel"] - 1)
        elif v == "edit":
            state["e_cursor"] = (state["e_cursor"] - 1) % len(_CAMPOS)
            state["e_msg"] = ""
        elif v == "picker":
            state["p_cursor"] = max(0, state["p_cursor"] - 1)
        elif v == "dup_meses":
            state["dup_cursor"] = max(0, state["dup_cursor"] - 1)

    @kb.add("c-d")
    def _pg_down(e):
        try:
            h = get_app().output.get_size().rows
        except Exception:
            h = 40
        half = max(1, (h - 7) // 2)
        v = state["view"]
        if v == "main":
            state["cursor"] = min(max(0, len(nav_indices) - 1), state["cursor"] + half)
        elif v == "detail":
            state["d_sel"] = min(max(0, len(_entradas()) - 1), state["d_sel"] + half)
        elif v == "picker":
            state["p_cursor"] += half
        elif v == "dup_meses":
            state["dup_cursor"] = min(len(meses_rango) - 1, state["dup_cursor"] + half)

    @kb.add("c-u")
    def _pg_up_or_clear(e):
        v = state["view"]
        if v == "picker":
            state["p_filter"] = ""
            state["p_cursor"] = 0
            return
        try:
            h = get_app().output.get_size().rows
        except Exception:
            h = 40
        half = max(1, (h - 7) // 2)
        if v == "main":
            state["cursor"] = max(0, state["cursor"] - half)
        elif v == "detail":
            state["d_sel"] = max(0, state["d_sel"] - half)
        elif v == "dup_meses":
            state["dup_cursor"] = max(0, state["dup_cursor"] - half)

    @kb.add("g", "g", filter=_no_input)
    def _go_top(e):
        v = state["view"]
        if v == "main":       state["cursor"]     = 0
        elif v == "detail":   state["d_sel"]      = state["d_offset"] = 0
        elif v == "picker":   state["p_cursor"]   = 0
        elif v == "dup_meses": state["dup_cursor"] = 0

    @kb.add("G", filter=_no_input)
    def _go_bottom(e):
        v = state["view"]
        if v == "main":
            state["cursor"] = max(0, len(nav_indices) - 1)
        elif v == "detail":
            state["d_sel"] = max(0, len(_entradas()) - 1)
        elif v == "picker":
            state["p_cursor"] = max(0, len(_p_filtered()) - 1)
        elif v == "dup_meses":
            state["dup_cursor"] = max(0, len(meses_rango) - 1)

    # n: nueva entrada desde main; cancela confirmación desde confirm
    @kb.add("n", filter=_only_main)
    def _nueva(e):
        _abrir_nueva()

    @kb.add("n", filter=_only_confirm)
    def _confirm_no(e):
        state["view"] = "detail"

    # d y x: solo en detail (no deben capturarse en input/picker)
    @kb.add("d", filter=_only_detail)
    def _duplicate(e):
        _abrir_dup()

    @kb.add("x", filter=_only_detail)
    def _delete(e):
        ents    = _entradas()
        sel_set = state["d_selected"]
        if sel_set:
            filas_del = [e.fila_xlsx for e in ents if e.fila_xlsx in sel_set]
        else:
            if not ents:
                return
            filas_del = [ents[state["d_sel"]].fila_xlsx]
        state["confirm_filas"] = filas_del
        state["view"]          = "confirm"

    # s/y: guardar (edit/dup_meses) o confirmar eliminar (confirm)
    @kb.add("s", filter=_edit_or_confirm)
    @kb.add("y", filter=_edit_or_confirm)
    def _confirm_yes(e):
        app = e.app

        # Nueva entrada en edit: pasar a selección de meses sin thread
        if state["view"] == "edit" and state["e_nueva"]:
            state.update(dup_campos=dict(state["e_campos"]),
                         dup_cursor=0, dup_selected=set(), view="dup_meses")
            return

        # Dup/nueva sin meses seleccionados: volver al paso anterior
        if state["view"] == "dup_meses" and not state["dup_selected"]:
            state["view"] = "edit" if state["e_nueva"] else "detail"
            return

        def _run_op():
            v = state["view"]
            saved_ok = False
            try:
                if v == "confirm":
                    _eliminar_entradas(wb_live, state["confirm_filas"])
                    _recargar()
                    state["d_selected"].clear()
                    state["d_sel"] = max(0, state["d_sel"] - len(state["confirm_filas"]))
                    state["view"]  = "detail"
                    saved_ok = True
                elif v == "edit":
                    _do_save()
                    saved_ok = True
                elif v == "dup_meses":
                    for i in sorted(state["dup_selected"]):
                        año, mes = meses_rango[i]
                        campos = dict(state["dup_campos"])
                        campos["año"] = año
                        campos["mes"] = mes
                        _insertar_entrada(wb_live, claves, campos)
                    _recargar()
                    state["d_cat1"] = state["dup_campos"].get("cat1", state["d_cat1"])
                    state["d_cat2"] = state["dup_campos"].get("cat2", state["d_cat2"])
                    state["e_nueva"] = False
                    state["view"]   = "detail"
                    saved_ok = True
            except Exception as ex:
                state["e_msg"] = f"err:{ex}"
                state["view"]  = v
            finally:
                state["saving"] = False
                app.invalidate()
            if saved_ok:
                # Esperar a escritura anterior antes de lanzar la nueva
                if _save_thread[0] is not None:
                    _save_thread[0].join(timeout=60)
                _disk_write(app)

        state["saving"] = True
        threading.Thread(target=_run_op, daemon=True).start()


    # space: toggle selección en detail o dup_meses; en input/picker se escribe
    @kb.add("space", filter=_only_detail)
    def _toggle_sel(e):
        ents = _entradas()
        if not ents:
            return
        sel     = state["d_sel"]
        fxl     = ents[sel].fila_xlsx
        sel_set = state["d_selected"]
        if fxl in sel_set:
            sel_set.discard(fxl)
        else:
            sel_set.add(fxl)
        state["d_sel"] = min(sel, len(ents) - 1)

    @kb.add("space", filter=_only_dup)
    def _toggle_mes(e):
        i       = state["dup_cursor"]
        sel_set = state["dup_selected"]
        if i in sel_set:
            sel_set.discard(i)
        else:
            sel_set.add(i)
        state["dup_cursor"] = min(len(meses_rango) - 1, i + 1)

    # a: seleccionar todo en detail o dup_meses; en input/picker se escribe
    @kb.add("a", filter=_only_detail)
    def _select_all(e):
        ents    = _entradas()
        sel_set = state["d_selected"]
        all_fx  = {e.fila_xlsx for e in ents}
        if all_fx == sel_set:
            sel_set.clear()
        else:
            sel_set.update(all_fx)

    @kb.add("a", filter=_only_dup)
    def _select_all_meses(e):
        all_idx = set(range(len(meses_rango)))
        if all_idx == state["dup_selected"]:
            state["dup_selected"].clear()
        else:
            state["dup_selected"] = all_idx

    @kb.add("enter")
    @kb.add("l", filter=_no_input)   # l no se escribe en input, sí en picker (select)
    def _enter(e):
        v = state["view"]
        if v == "main" and nav_indices:
            fila = filas[nav_indices[state["cursor"]]]
            _abrir_detalle(fila.cat1, fila.cat2)
        elif v == "detail":
            _abrir_edit()
        elif v == "edit":
            _, _, usa_picker = _CAMPOS[state["e_cursor"]]
            if usa_picker:
                _abrir_picker()
            else:
                _abrir_input()
        elif v == "input":
            val = _validar_input()
            if val is not None:
                state["e_campos"][state["i_campo"]] = val
                state["view"] = "edit"
                state["e_msg"] = ""
        elif v == "picker":
            opts = _p_filtered()
            if opts:
                state["e_campos"][state["p_campo"]] = opts[state["p_cursor"]]
            state["p_filter"] = ""
            state["view"] = "edit"
            state["e_msg"] = ""

    def _go_back() -> str | None:
        """Devuelve la vista anterior o None si hay que salir."""
        v = state["view"]
        if v in ("input", "picker"):
            return "edit"
        if v == "dup_meses":
            return "edit" if state["e_nueva"] else "detail"
        if v == "edit":
            if state["e_nueva"]:
                state["e_nueva"] = False
                return "main"
            return "detail"
        if v == "confirm":
            return "detail"
        if v == "detail":
            return "main"
        return None

    # h: volver atrás excepto en input/picker donde se escribe
    @kb.add("escape")
    @kb.add("h", filter=_no_text)
    def _back(e):
        dest = _go_back()
        if dest:
            state["view"] = dest
        else:
            e.app.exit()

    @kb.add("q", filter=_no_text)
    def _quit(e):
        dest = _go_back()
        if dest:
            state["view"] = dest
        else:
            e.app.exit()

    @kb.add("c-c")
    def _force_quit(e): e.app.exit()

    @kb.add("<any>")
    def _any_key(e):
        v = state["view"]
        key = e.key_sequence[0].key
        if v == "input":
            if key in ("backspace", "c-h"):
                state["i_buf"]   = state["i_buf"][:-1]
                state["i_error"] = ""
            elif isinstance(key, str) and len(key) == 1 and key.isprintable():
                state["i_buf"]  += key
                state["i_error"] = ""
        elif v == "picker":
            if key in ("backspace", "c-h"):
                state["p_filter"]  = state["p_filter"][:-1]
                state["p_cursor"]  = 0
            elif isinstance(key, str) and len(key) == 1 and key.isprintable():
                state["p_filter"] += key
                state["p_cursor"]  = 0

    app = Application(
        layout=Layout(Window(content=FormattedTextControl(text=_render, focusable=True))),
        key_bindings=kb,
        style=style,
        full_screen=True,
    )
    app.run()

    # Esperar a que termine la última escritura a disco pendiente
    if _save_thread[0] is not None:
        _save_thread[0].join(timeout=60)

    wb_live.close()


# ---------------------------------------------------------------------------
# Comando click
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Vista mensual (Presupuesto vs Real)
# ---------------------------------------------------------------------------

def _cmd_vista_mes(consola, ruta_origen: Path, mes_opt: str | None, año_opt: int | None) -> None:
    """Muestra comparativa Presupuesto vs Real para un mes concreto."""
    from datetime import date
    from collections import defaultdict
    from decimal import Decimal
    from rich import box
    from rich.table import Table
    from rich.text import Text
    from presupuesto.escritor import leer_numero
    import openpyxl

    hoy = date.today()
    año = año_opt or hoy.year
    if mes_opt:
        mes_norm = mes_opt.strip().capitalize()
        if mes_norm not in _MESES_ORD:
            consola.print(f"[red]Mes inválido:[/red] '{mes_opt}'. Usa: {', '.join(_MESES_ORD)}")
            return
        mes = mes_norm
    else:
        mes = _MESES_ORD[hoy.month - 1]

    consola.print(f"[dim]Leyendo datos para {mes} {año}…[/dim]")

    wb = openpyxl.load_workbook(str(ruta_origen), data_only=True, read_only=True)
    presupuesto: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    real:        dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    orden: list[tuple[str, str]] = []
    visto: set[tuple[str, str]] = set()

    try:
        ws = wb["Datos"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[_COL_AÑO] is None:
                continue
            try:
                fila_año = int(row[_COL_AÑO])
            except (TypeError, ValueError):
                continue
            if fila_año != año or str(row[_COL_MES] or "").strip() != mes:
                continue
            estado = str(row[_COL_ESTADO] or "").strip()
            if estado not in ("Presupuesto", "Real"):
                continue
            cat1 = str(row[_COL_CAT1] or "").strip()
            cat2 = str(row[_COL_CAT2] or "").strip()
            if not cat1:
                continue
            # Excluir Finanzas/Balance del total
            if cat1.lower() == "finanzas" and cat2.lower() == "balance":
                continue
            imp_raw = leer_numero(row[_COL_IMPORTE])
            if imp_raw is None:
                continue
            importe = Decimal(str(imp_raw))
            key = (cat1, cat2)
            if key not in visto:
                orden.append(key)
                visto.add(key)
            if estado == "Presupuesto":
                presupuesto[key] += importe
            else:
                real[key] += importe
    finally:
        wb.close()

    if not orden:
        consola.print(f"[yellow]No hay datos para {mes} {año}.[/yellow]")
        return

    # Agrupar por Cat1 manteniendo orden de aparición
    grupos: dict[str, list[str]] = defaultdict(list)
    orden_cat1: list[str] = []
    visto_cat1: set[str] = set()
    cat2_visto: set[tuple[str, str]] = set()
    for cat1, cat2 in orden:
        if cat1 not in visto_cat1:
            orden_cat1.append(cat1)
            visto_cat1.add(cat1)
        if (cat1, cat2) not in cat2_visto:
            grupos[cat1].append(cat2)
            cat2_visto.add((cat1, cat2))

    def _fmt(v: Decimal) -> Text:
        if v == Decimal(0):
            return Text("—", style="dim", justify="right")
        return Text(f"{v:+.2f}€", style=("red" if v < 0 else "green"), justify="right")

    def _fmt_dif(v: Decimal) -> Text:
        if v == Decimal(0):
            return Text("0.00€", style="dim", justify="right")
        return Text(f"{v:+.2f}€", style=("green" if v > 0 else "red"), justify="right")

    tabla = Table(
        title=f"Presupuesto vs Real — {mes} {año}",
        box=box.SIMPLE_HEAD, show_lines=False, padding=(0, 1),
    )
    tabla.add_column("Cat. 1",      style="bold", no_wrap=True, min_width=18)
    tabla.add_column("Cat. 2",                    no_wrap=True, min_width=20)
    tabla.add_column("Presupuesto", justify="right", no_wrap=True, min_width=13)
    tabla.add_column("Real",        justify="right", no_wrap=True, min_width=13)
    tabla.add_column("Diferencia",  justify="right", no_wrap=True, min_width=13)

    total_pres = Decimal(0)
    total_real = Decimal(0)

    for cat1 in orden_cat1:
        subtotal_pres = Decimal(0)
        subtotal_real = Decimal(0)
        primera = True

        for cat2 in grupos[cat1]:
            key = (cat1, cat2)
            pres = presupuesto.get(key, Decimal(0))
            re   = real.get(key, Decimal(0))
            subtotal_pres += pres
            subtotal_real += re
            tabla.add_row(
                cat1 if primera else "",
                cat2 or "—",
                _fmt(pres),
                _fmt(re),
                _fmt_dif(re - pres),
            )
            primera = False

        sub_dif = subtotal_real - subtotal_pres
        tabla.add_row(
            "",
            Text(f"  Σ {cat1}", style="dim italic"),
            Text(f"{subtotal_pres:+.2f}€", style="dim", justify="right"),
            Text(f"{subtotal_real:+.2f}€", style="dim", justify="right"),
            _fmt_dif(sub_dif),
            end_section=True,
        )
        total_pres += subtotal_pres
        total_real += subtotal_real

    tabla.add_row(
        Text("TOTAL", style="bold"), "",
        Text(f"{total_pres:+.2f}€", style="bold", justify="right"),
        Text(f"{total_real:+.2f}€", style="bold", justify="right"),
        _fmt_dif(total_real - total_pres),
    )

    consola.print()
    consola.print(tabla)


@click.command("vista")
@click.option("--meses", default=12, show_default=True,
              help="Número de meses a mostrar (desde el actual).")
@click.option("--cat1", "filtro_cat1", default=None, metavar="CATEGORIA",
              help="Mostrar sólo esta Categoría 1.")
@click.option("--balance", "incluir_balance", is_flag=True, default=False,
              help="Incluir entradas Finanzas/Balance (excluidas por defecto).")
@click.option("--cuentas", "modo_balance", is_flag=True, default=False,
              help="Ver solo Finanzas/Balance con filas por cuenta (modo balance).")
@click.option("--gastos", "modo_gastos", is_flag=True, default=False,
              help="Agrupar por tipo de gasto (Fijos/Discrecionales/…) con Ahorro separado.")
@click.option("--sin-ajuste", "sin_ajuste", is_flag=True, default=False,
              help="En --gastos, desactiva el ajuste de Vivienda compartida (÷2 y sin ocultar transferencia).")
@click.option("--cuenta", "filtrar_cuenta", is_flag=True, default=False,
              help="Abrir selector para filtrar por una o varias cuentas.")
@click.option("--mes", "mes_opt", default=None, metavar="MES",
              help="Ver comparativa Presupuesto vs Real de un mes (Ene, Feb, …). Por defecto: mes actual.")
@click.option("--año", "año_opt", default=None, type=int, metavar="AÑO",
              help="Año para --mes. Por defecto: año actual.")
def cmd_vista(meses: int, filtro_cat1: str | None, incluir_balance: bool,
              modo_balance: bool, modo_gastos: bool, sin_ajuste: bool,
              filtrar_cuenta: bool, mes_opt: str | None, año_opt: int | None):
    """Presupuesto a un año vista (TUI interactivo)."""
    import shutil
    import tempfile
    from datetime import date
    from rich.console import Console
    from presupuesto.config import cargar_config

    consola = Console()
    config  = cargar_config()
    ruta_str = config.get("archivo_presupuesto", "")
    if not ruta_str:
        consola.print("[red]No hay ruta al xlsx configurada.[/red]"); raise SystemExit(1)
    ruta_origen = Path(ruta_str).expanduser()
    if not ruta_origen.exists():
        consola.print(f"[red]No se encuentra:[/red] {ruta_origen}"); raise SystemExit(1)

    # ── Modo mensual ──────────────────────────────────────────────────────────
    if mes_opt is not None or año_opt is not None:
        _cmd_vista_mes(consola, ruta_origen, mes_opt, año_opt)
        return

    # Copiar xlsx a directorio temporal en filesystem local (evita latencia WSL→Windows)
    consola.print("[dim]Copiando xlsx a directorio temporal…[/dim]")
    tmp_dir   = Path(tempfile.mkdtemp(prefix="presupuesto_"))
    ruta_xlsx = tmp_dir / ruta_origen.name
    shutil.copy2(str(ruta_origen), str(ruta_xlsx))

    hoy = date.today()
    meses_rango: list[tuple[int, str]] = []
    a, m = hoy.year, hoy.month
    for _ in range(meses):
        meses_rango.append((a, _MESES_ORD[m - 1]))
        m += 1
        if m > 12:
            m, a = 1, a + 1

    ajuste_vivienda = modo_gastos and not sin_ajuste

    # ── Filtro opcional por cuenta ────────────────────────────────────────────
    cuentas_filtro: set[str] | None = None
    if filtrar_cuenta:
        import openpyxl as _opx
        _wb_cl = _opx.load_workbook(str(ruta_xlsx), data_only=True, read_only=True)
        try:
            _cuentas_lista = []
            for _row in _wb_cl["Claves"].iter_rows(min_row=2, values_only=True):
                if _row and _row[0]:
                    _cuentas_lista.append(str(_row[0]).strip())
        except KeyError:
            _cuentas_lista = []
        finally:
            _wb_cl.close()

        if not _cuentas_lista:
            consola.print("[yellow]No se encontraron cuentas en la hoja Claves.[/yellow]")
            shutil.rmtree(tmp_dir, ignore_errors=True)
            return

        seleccion = _tui_seleccionar_cuentas(_cuentas_lista)
        if seleccion == "cancelar":
            # Esc → cancelar completamente
            consola.print("[dim]Cancelado.[/dim]")
            shutil.rmtree(tmp_dir, ignore_errors=True)
            return
        cuentas_filtro = seleccion  # None = sin filtro (Enter sin selección), set = filtrar

    consola.print("[dim]Leyendo datos del xlsx…[/dim]")
    filas, nav_indices, detalles, claves, opciones = _leer_datos(
        ruta_xlsx, meses_rango, incluir_balance, modo_balance, modo_gastos,
        ajuste_vivienda, cuentas_filtro)

    if not filas or not nav_indices:
        msg = ("No hay entradas de Finanzas/Balance para el rango seleccionado." if modo_balance else
               "No hay entradas de Presupuesto (tipo de gasto) para el rango seleccionado." if modo_gastos else
               "No hay entradas de Presupuesto para el rango seleccionado.")
        consola.print(f"[yellow]{msg}[/yellow]")
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return

    if filtro_cat1 and not modo_balance:
        fl = filtro_cat1.lower()
        nav_indices = [i for i in nav_indices if filas[i].cat1.lower() == fl]
        if not nav_indices:
            consola.print(f"[yellow]No se encontraron entradas para '{filtro_cat1}'.[/yellow]")
            shutil.rmtree(tmp_dir, ignore_errors=True)
            return

    try:
        _tui_vista(filas, nav_indices, meses_rango, detalles, ruta_xlsx, claves, opciones,
                   ruta_origen=ruta_origen, incluir_balance=incluir_balance,
                   modo_balance=modo_balance, modo_gastos=modo_gastos,
                   ajuste_vivienda=ajuste_vivienda)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
