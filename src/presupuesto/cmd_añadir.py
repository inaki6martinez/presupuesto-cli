"""Comandos 'añadir': añade entradas de Presupuesto o movimientos Reales.

Subcomandos:
    presupuesto — Copia entradas de Presupuesto para uno o varios meses.
    movimiento  — Añade un movimiento Real con TUI campo a campo.
"""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

import click

_MESES_ORD = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
              "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

_COL_AÑO       = 0
_COL_MES       = 1
_COL_CAT1      = 2
_COL_CAT2      = 3
_COL_CAT3      = 4
_COL_ENTIDAD   = 5
_COL_IMPORTE   = 6
_COL_PROVEEDOR = 7
_COL_TIPO_GASTO = 8
_COL_CUENTA    = 9
_COL_BANCO     = 10
_COL_TIPO_CUEN = 11
_COL_ESTADO    = 12


# ---------------------------------------------------------------------------
# Lectura de datos
# ---------------------------------------------------------------------------

def _leer_meses_presupuesto(ruta_xlsx: Path) -> list[tuple[int, str]]:
    """Devuelve lista ordenada de (año, mes) con entradas Presupuesto."""
    import openpyxl

    wb  = openpyxl.load_workbook(str(ruta_xlsx), data_only=True, read_only=True)
    try:
        ws = wb["Datos"]
    except KeyError:
        wb.close()
        return []

    meses: set[tuple[int, str]] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[_COL_AÑO] is None:
            continue
        if str(row[_COL_ESTADO] or "").strip() != "Presupuesto":
            continue
        try:
            año = int(row[_COL_AÑO])
        except (TypeError, ValueError):
            continue
        mes = str(row[_COL_MES] or "").strip()
        if mes in _MESES_ORD:
            meses.add((año, mes))

    wb.close()

    def _orden(t: tuple[int, str]) -> tuple[int, int]:
        return t[0], _MESES_ORD.index(t[1])

    return sorted(meses, key=_orden)


def _leer_cuentas(ruta_xlsx: Path) -> list[tuple[str, str, str]]:
    """Devuelve lista de (cuenta, banco, tipo_cuenta) desde la hoja Claves."""
    import openpyxl

    wb = openpyxl.load_workbook(str(ruta_xlsx), data_only=True, read_only=True)
    try:
        ws = wb["Claves"]
    except KeyError:
        wb.close()
        return []

    resultado = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        cuenta      = str(row[0]).strip()
        banco       = str(row[1]).strip() if row[1] else ""
        tipo_cuenta = str(row[2]).strip() if row[2] else ""
        if cuenta:
            resultado.append((cuenta, banco, tipo_cuenta))
    wb.close()
    return resultado


# ---------------------------------------------------------------------------
# TUI multi-selección de meses (para subcomando presupuesto)
# ---------------------------------------------------------------------------

def _tui_seleccionar_meses(
    meses: list[tuple[int, str]],
) -> list[tuple[int, str]] | None:
    """TUI para seleccionar uno o varios meses. Devuelve lista seleccionada o None."""
    from prompt_toolkit import Application
    from prompt_toolkit.application import get_app
    from prompt_toolkit.formatted_text import FormattedText
    from prompt_toolkit.key_binding import KeyBindings
    from prompt_toolkit.layout import Layout
    from prompt_toolkit.layout.containers import Window
    from prompt_toolkit.layout.controls import FormattedTextControl
    from prompt_toolkit.styles import Style

    if not meses:
        return None

    style = Style.from_dict({
        "titulo":  "bold",
        "cursor":  "reverse bold",
        "selec":   "bold #00cc44",
        "cur_sel": "reverse bold #00cc44",
        "dim":     "#666666",
        "footer":  "#666666",
        "fkey":    "#aaaaaa bold",
    })

    state: dict = {
        "cursor":    0,
        "selec":     set(),   # índices seleccionados
        "resultado": None,
    }

    def _clamp() -> None:
        state["cursor"] = max(0, min(state["cursor"], len(meses) - 1))

    def _render() -> FormattedText:
        try:
            size = get_app().output.get_size()
            w, h = size.columns, size.rows
        except Exception:
            w, h = 80, 30

        buf: list[tuple[str, str]] = []
        def t(st, s): buf.append((st, s))
        def nl():     buf.append(("", "\n"))

        t("class:titulo", "  Seleccionar meses")
        nl()
        t("class:dim", "─" * w)
        nl()

        list_h = max(3, h - 7)
        cur    = state["cursor"]
        ws_start = max(0, cur - list_h // 2)
        ws_end   = min(len(meses), ws_start + list_h)
        ws_start = max(0, ws_end - list_h)

        for i in range(ws_start, ws_end):
            año, mes = meses[i]
            es_cur   = i == cur
            es_sel   = i in state["selec"]
            marca    = "[✓]" if es_sel else "[ ]"
            etiqueta = f"  {marca} {mes} {año}"

            if es_cur and es_sel:
                t("class:cur_sel", etiqueta)
            elif es_cur:
                t("class:cursor",  etiqueta)
            elif es_sel:
                t("class:selec",   etiqueta)
            else:
                t("",              etiqueta)
            nl()

        t("class:dim", "─" * w)
        nl()
        for k, desc in [
            ("↑↓", "Navegar"),
            ("Esp", "Seleccionar"),
            ("a",   "Todos"),
            ("Enter", "Confirmar"),
            ("Esc", "Cancelar"),
        ]:
            t("class:fkey",   f" {k} ")
            t("class:footer", f"{desc}  ")

        return FormattedText(buf)

    kb = KeyBindings()

    @kb.add("up")
    def _(e): state["cursor"] = max(0, state["cursor"] - 1)

    @kb.add("down")
    def _(e): state["cursor"] = min(len(meses) - 1, state["cursor"] + 1)

    @kb.add("space")
    def _(e):
        i = state["cursor"]
        if i in state["selec"]:
            state["selec"].discard(i)
        else:
            state["selec"].add(i)

    @kb.add("a")
    def _(e):
        if len(state["selec"]) == len(meses):
            state["selec"] = set()
        else:
            state["selec"] = set(range(len(meses)))

    @kb.add("enter")
    def _(e):
        sel = sorted(state["selec"])
        if sel:
            state["resultado"] = [meses[i] for i in sel]
        e.app.exit()

    @kb.add("escape")
    @kb.add("c-c")
    def _(e): e.app.exit()

    app = Application(
        layout=Layout(Window(content=FormattedTextControl(text=_render, focusable=True))),
        key_bindings=kb,
        style=style,
        full_screen=True,
    )
    app.run()
    return state["resultado"]


# ---------------------------------------------------------------------------
# TUI selección de cuenta
# ---------------------------------------------------------------------------

def _tui_seleccionar_cuenta(
    cuentas: list[tuple[str, str, str]],
) -> tuple[str, str, str] | None:
    """TUI para seleccionar una cuenta. Devuelve (cuenta, banco, tipo_cuenta) o None."""
    from prompt_toolkit import Application
    from prompt_toolkit.application import get_app
    from prompt_toolkit.formatted_text import FormattedText
    from prompt_toolkit.key_binding import KeyBindings
    from prompt_toolkit.layout import Layout
    from prompt_toolkit.layout.containers import Window
    from prompt_toolkit.layout.controls import FormattedTextControl
    from prompt_toolkit.styles import Style

    style = Style.from_dict({
        "titulo": "bold",
        "cursor": "reverse bold",
        "filtro": "bold yellow",
        "dim":    "#666666",
        "footer": "#666666",
        "fkey":   "#aaaaaa bold",
    })

    state: dict = {"cursor": 0, "filtro": "", "resultado": None}

    def _filtradas() -> list[tuple[str, str, str]]:
        f = state["filtro"].lower()
        return [c for c in cuentas if f in c[0].lower()] if f else cuentas

    def _clamp() -> None:
        vis = _filtradas()
        state["cursor"] = max(0, min(state["cursor"], len(vis) - 1))

    def _render() -> FormattedText:
        try:
            size = get_app().output.get_size()
            w, h = size.columns, size.rows
        except Exception:
            w, h = 80, 30

        buf: list[tuple[str, str]] = []
        def t(st, s): buf.append((st, s))
        def nl():     buf.append(("", "\n"))

        t("class:titulo", "  Seleccionar cuenta")
        nl()
        filtro_txt = (state["filtro"] + "▌") if state["filtro"] else "▌"
        t("class:dim",    "  Filtro: ")
        t("class:filtro", filtro_txt)
        nl()
        t("class:dim", "─" * w)
        nl()

        vis      = _filtradas()
        list_h   = max(3, h - 8)
        cur      = state["cursor"]
        ws_start = max(0, cur - list_h // 2)
        ws_end   = min(len(vis), ws_start + list_h)
        ws_start = max(0, ws_end - list_h)

        for i in range(ws_start, ws_end):
            cuenta, banco, tipo = vis[i]
            es_cur = i == cur
            arrow  = "►" if es_cur else " "
            st     = "class:cursor" if es_cur else ""
            t("class:dim", f"  {arrow} ")
            t(st, f"{cuenta:<28}  {banco:<20}  {tipo}")
            nl()

        t("class:dim", "─" * w)
        nl()
        for k, desc in [("↑↓", "Navegar"), ("Enter", "Seleccionar"),
                        ("^U", "Borrar filtro"), ("Esc", "Cancelar")]:
            t("class:fkey",   f" {k} ")
            t("class:footer", f"{desc}  ")

        return FormattedText(buf)

    kb = KeyBindings()

    @kb.add("up")
    def _(e): state["cursor"] = max(0, state["cursor"] - 1)

    @kb.add("down")
    def _(e):
        state["cursor"] = min(max(0, len(_filtradas()) - 1), state["cursor"] + 1)

    @kb.add("enter")
    def _(e):
        vis = _filtradas()
        if vis:
            state["resultado"] = vis[state["cursor"]]
        e.app.exit()

    @kb.add("escape")
    @kb.add("c-c")
    def _(e): e.app.exit()

    @kb.add("backspace")
    @kb.add("c-h")
    def _(e):
        state["filtro"] = state["filtro"][:-1]
        _clamp()

    @kb.add("c-u")
    def _(e):
        state["filtro"] = ""
        _clamp()

    @kb.add("<any>")
    def _(e):
        key = e.key_sequence[0].key
        if isinstance(key, str) and len(key) == 1 and key.isprintable():
            state["filtro"] += key
            state["cursor"] = 0

    app = Application(
        layout=Layout(Window(content=FormattedTextControl(text=_render, focusable=True))),
        key_bindings=kb,
        style=style,
        full_screen=True,
    )
    app.run()
    return state["resultado"]


# ---------------------------------------------------------------------------
# TUI selección de fecha (año + mes) para movimiento Real
# ---------------------------------------------------------------------------

def _tui_seleccionar_fecha(año_default: int) -> tuple[int, str] | None:
    """TUI para seleccionar año y mes. Devuelve (año, mes) o None si se cancela."""
    from prompt_toolkit import Application
    from prompt_toolkit.application import get_app
    from prompt_toolkit.formatted_text import FormattedText
    from prompt_toolkit.key_binding import KeyBindings
    from prompt_toolkit.layout import Layout
    from prompt_toolkit.layout.containers import Window
    from prompt_toolkit.layout.controls import FormattedTextControl
    from prompt_toolkit.styles import Style

    style = Style.from_dict({
        "titulo":   "bold",
        "anio":     "bold cyan",
        "mes-cur":  "reverse bold",
        "dim":      "#666666",
        "footer":   "#666666",
        "fkey":     "#aaaaaa bold",
        "campo":    "bold",
        "campo-act":"bold yellow",
    })

    # foco: "año" o "mes"
    state: dict = {
        "año":       año_default,
        "año_buf":   "",         # buffer de escritura del año
        "mes_cur":   0,          # índice en _MESES_ORD
        "foco":      "año",      # campo activo
        "resultado": None,
    }

    def _render() -> FormattedText:
        try:
            size = get_app().output.get_size()
            w, h = size.columns, size.rows
        except Exception:
            w, h = 80, 30

        buf: list[tuple[str, str]] = []
        def t(st, s): buf.append((st, s))
        def nl():     buf.append(("", "\n"))

        t("class:titulo", "  Seleccionar fecha del movimiento")
        nl()
        t("class:dim", "─" * w)
        nl()
        nl()

        # ── Año ──
        año_st = "class:campo-act" if state["foco"] == "año" else "class:campo"
        año_txt = state["año_buf"] + "▌" if state["año_buf"] else str(state["año"])
        t(año_st, f"  Año:  {año_txt}")
        if state["foco"] == "año":
            t("class:dim", "   (← → o escribe el año, Tab para ir a Mes)")
        nl()
        nl()

        # ── Mes ──
        mes_st = "class:campo-act" if state["foco"] == "mes" else "class:campo"
        t(mes_st, "  Mes:")
        nl()

        list_h   = max(4, h - 12)
        cur      = state["mes_cur"]
        ws_start = max(0, cur - list_h // 2)
        ws_end   = min(len(_MESES_ORD), ws_start + list_h)
        ws_start = max(0, ws_end - list_h)

        for i in range(ws_start, ws_end):
            es_cur = i == cur
            arrow  = "►" if es_cur else " "
            if state["foco"] == "mes" and es_cur:
                t("class:mes-cur", f"    {arrow} {_MESES_ORD[i]}")
            elif es_cur:
                t("class:dim",     f"    {arrow} {_MESES_ORD[i]}")
            else:
                t("",              f"    {arrow} {_MESES_ORD[i]}")
            nl()

        t("class:dim", "─" * w)
        nl()
        atajos = [
            ("Tab", "Cambiar campo"),
            ("↑↓",  "Navegar mes"),
            ("← →", "Año ±1"),
            ("Enter", "Confirmar"),
            ("Esc",   "Cancelar"),
        ]
        for k, desc in atajos:
            t("class:fkey",   f" {k} ")
            t("class:footer", f"{desc}  ")

        return FormattedText(buf)

    kb = KeyBindings()

    @kb.add("tab")
    def _(e):
        state["año_buf"] = ""
        state["foco"] = "mes" if state["foco"] == "año" else "año"

    @kb.add("up")
    def _(e):
        if state["foco"] == "mes":
            state["mes_cur"] = max(0, state["mes_cur"] - 1)
        else:
            state["año"] += 1

    @kb.add("down")
    def _(e):
        if state["foco"] == "mes":
            state["mes_cur"] = min(len(_MESES_ORD) - 1, state["mes_cur"] + 1)
        else:
            state["año"] -= 1

    @kb.add("left")
    def _(e):
        if state["foco"] == "año":
            state["año"] -= 1

    @kb.add("right")
    def _(e):
        if state["foco"] == "año":
            state["año"] += 1

    @kb.add("enter")
    def _(e):
        if state["foco"] == "año" and state["año_buf"]:
            try:
                state["año"] = int(state["año_buf"])
                state["año_buf"] = ""
            except ValueError:
                state["año_buf"] = ""
                return
        state["resultado"] = (state["año"], _MESES_ORD[state["mes_cur"]])
        e.app.exit()

    @kb.add("escape")
    @kb.add("c-c")
    def _(e): e.app.exit()

    @kb.add("backspace")
    @kb.add("c-h")
    def _(e):
        if state["foco"] == "año":
            state["año_buf"] = state["año_buf"][:-1]

    @kb.add("<any>")
    def _(e):
        key = e.key_sequence[0].key
        if not (isinstance(key, str) and len(key) == 1):
            return
        if state["foco"] == "año" and key.isdigit():
            buf = state["año_buf"] + key
            if len(buf) <= 4:
                state["año_buf"] = buf
                if len(buf) == 4:
                    try:
                        state["año"] = int(buf)
                        state["año_buf"] = ""
                        state["foco"] = "mes"
                    except ValueError:
                        state["año_buf"] = ""

    app = Application(
        layout=Layout(Window(content=FormattedTextControl(text=_render, focusable=True))),
        key_bindings=kb,
        style=style,
        full_screen=True,
    )
    app.run()
    return state["resultado"]


# ---------------------------------------------------------------------------
# Helpers comunes
# ---------------------------------------------------------------------------

def _cargar_xlsx_config() -> Path:
    """Carga config, valida ruta xlsx y devuelve ruta_xlsx."""
    from rich.console import Console
    from presupuesto.config import cargar_config

    consola = Console()
    config   = cargar_config()
    ruta_str = config.get("archivo_presupuesto", "")
    if not ruta_str:
        consola.print("[red]No hay ruta al xlsx configurada.[/red]")
        raise SystemExit(1)
    ruta_xlsx = Path(ruta_str).expanduser()
    if not ruta_xlsx.exists():
        consola.print(f"[red]No se encuentra:[/red] {ruta_xlsx}")
        raise SystemExit(1)
    return ruta_xlsx


def _registrar_revision_interactivo(consola, cuenta: str) -> None:
    """Ofrece al usuario registrar la revisión manual de la cuenta en revisiones.json."""
    from datetime import date
    from presupuesto.duplicados import GestorRevisiones

    gestor = GestorRevisiones()
    revision_actual = gestor.obtener_revision(cuenta)
    hoy = date.today()

    consola.print()
    if revision_actual:
        consola.print(
            f"  Última revisión de [bold]{cuenta}[/bold]: "
            f"[cyan]{revision_actual.isoformat()}[/cyan]"
        )
    else:
        consola.print(
            f"  [dim]Sin revisión registrada para [bold]{cuenta}[/bold][/dim]"
        )

    if not click.confirm(f"  ¿Registrar revisión de '{cuenta}'?", default=True):
        return

    while True:
        raw = click.prompt(
            "  Fecha de la revisión",
            default=hoy.isoformat(),
        ).strip()
        try:
            fecha = date.fromisoformat(raw)
            break
        except ValueError:
            consola.print("  [red]Formato inválido. Usa YYYY-MM-DD (ej: 2026-03-31)[/red]")

    gestor.registrar_revision(cuenta, fecha)
    consola.print(
        f"  [green]✓ Revisión de '{cuenta}' registrada: {fecha.isoformat()}[/green]"
    )


# ---------------------------------------------------------------------------
# Grupo click
# ---------------------------------------------------------------------------

@click.group("añadir")
def cmd_añadir():
    """Añade entradas al presupuesto (Presupuesto o Real)."""


# ---------------------------------------------------------------------------
# Subcomando: presupuesto
# ---------------------------------------------------------------------------

@cmd_añadir.command("presupuesto")
def cmd_presupuesto():
    """Añade una nueva entrada de Presupuesto para uno o varios meses."""
    from rich.console import Console
    from rich.table import Table
    from rich import box
    from presupuesto.maestro import DatosMaestros
    from presupuesto.interactivo import pedir_categorizacion
    from presupuesto.escritor import EscritorDatos
    from presupuesto.categorizar import MovimientoCategorizado

    consola = Console()

    ruta_xlsx = _cargar_xlsx_config()
    consola.print("[dim]Leyendo datos del xlsx…[/dim]")
    meses_disponibles = _leer_meses_presupuesto(ruta_xlsx)
    cuentas           = _leer_cuentas(ruta_xlsx)

    if not meses_disponibles:
        consola.print("[yellow]No hay meses con entradas de Presupuesto en el xlsx.[/yellow]")
        raise SystemExit(1)

    # --- 1. Selección de meses ---
    meses_sel = _tui_seleccionar_meses(meses_disponibles)
    if not meses_sel:
        consola.print("[dim]Cancelado.[/dim]")
        return

    consola.print(
        f"\n  Meses seleccionados: "
        + ", ".join(f"[cyan]{mes} {año}[/cyan]" for año, mes in meses_sel)
    )

    # --- 2. Categorización ---
    datos_maestros = DatosMaestros(ruta_xlsx)
    resultado = pedir_categorizacion(datos_maestros, None)
    if resultado in ("salir", "saltar", "volver"):
        consola.print("[dim]Cancelado.[/dim]")
        return
    campos: dict = resultado

    # --- 3. Cuenta ---
    if not cuentas:
        consola.print("[red]No se encontraron cuentas en la hoja Claves.[/red]")
        raise SystemExit(1)
    seleccion_cuenta = _tui_seleccionar_cuenta(cuentas)
    if seleccion_cuenta is None:
        consola.print("[dim]Cancelado.[/dim]")
        return
    cuenta, banco, tipo_cuenta = seleccion_cuenta

    # --- 4. Importe ---
    consola.print(f"\n  Cuenta: [bold]{cuenta}[/bold]")
    while True:
        raw = click.prompt("\n  Importe (negativo = gasto, positivo = ingreso)").strip().replace(",", ".")
        try:
            importe = Decimal(raw).quantize(Decimal("0.01"))
            break
        except InvalidOperation:
            consola.print("  [red]Valor no válido. Usa formato numérico (ej: -45.50)[/red]")

    # --- 5. Previsualización ---
    consola.print()
    tabla = Table(
        title=f"Entradas a crear ({len(meses_sel)} mes(es))",
        box=box.SIMPLE_HEAD,
        show_lines=False,
        padding=(0, 1),
    )
    tabla.add_column("Mes",      no_wrap=True)
    tabla.add_column("Cat 1",    no_wrap=True)
    tabla.add_column("Cat 2",    no_wrap=True)
    tabla.add_column("Cat 3",    no_wrap=True, style="dim")
    tabla.add_column("Proveedor")
    tabla.add_column("Tipo gasto", style="dim")
    tabla.add_column("Cuenta",   no_wrap=True)
    tabla.add_column("Importe",  justify="right", no_wrap=True)

    color_imp = "red" if importe < 0 else "green"
    imp_str   = f"[{color_imp}]{importe:+.2f}[/{color_imp}]"

    for año, mes in meses_sel:
        tabla.add_row(
            f"{mes} {año}",
            campos.get("categoria1", ""),
            campos.get("categoria2", ""),
            campos.get("categoria3", ""),
            campos.get("proveedor",  ""),
            campos.get("tipo_gasto", ""),
            cuenta,
            imp_str,
        )

    consola.print(tabla)

    if not click.confirm("  ¿Confirmar escritura?", default=True):
        consola.print("[dim]Cancelado.[/dim]")
        return

    # --- 6. Escritura ---
    movimientos = []
    for año, mes in meses_sel:
        mov = MovimientoCategorizado(
            año=año,
            mes=mes,
            categoria1=campos.get("categoria1", ""),
            categoria2=campos.get("categoria2", ""),
            categoria3=campos.get("categoria3", ""),
            entidad=campos.get("entidad", ""),
            importe=importe,
            proveedor=campos.get("proveedor", ""),
            tipo_gasto=campos.get("tipo_gasto", ""),
            cuenta=cuenta,
            banco=banco or None,
            tipo_cuenta=tipo_cuenta or None,
            estado="Presupuesto",
            confianza="alta",
            fuente="añadir",
            requiere_confirmacion=False,
        )
        movimientos.append(mov)

    try:
        n = EscritorDatos(ruta_xlsx).escribir(movimientos)
        consola.print(f"\n  [green]✓ {n} entrada(s) escritas.[/green]")
    except Exception as e:
        consola.print(f"\n  [red]Error al escribir:[/red] {e}")
        raise SystemExit(1)


# ---------------------------------------------------------------------------
# Subcomando: movimiento
# ---------------------------------------------------------------------------

@cmd_añadir.command("movimiento")
def cmd_movimiento():
    """Añade un movimiento Real con TUI campo a campo."""
    from datetime import date
    from rich.console import Console
    from rich.table import Table
    from rich import box
    from presupuesto.maestro import DatosMaestros
    from presupuesto.tui_categorizar import TUICategorizacion
    from presupuesto.escritor import EscritorDatos
    from presupuesto.categorizar import MovimientoCategorizado

    consola = Console()

    ruta_xlsx = _cargar_xlsx_config()
    consola.print("[dim]Leyendo datos del xlsx…[/dim]")
    cuentas       = _leer_cuentas(ruta_xlsx)
    datos_maestros = DatosMaestros(ruta_xlsx)

    if not cuentas:
        consola.print("[red]No se encontraron cuentas en la hoja Claves.[/red]")
        raise SystemExit(1)

    # --- 1. Fecha ---
    año_hoy = date.today().year
    fecha = _tui_seleccionar_fecha(año_hoy)
    if fecha is None:
        consola.print("[dim]Cancelado.[/dim]")
        return
    año, mes = fecha
    consola.print(f"\n  Fecha: [cyan]{mes} {año}[/cyan]")

    # --- 2. Categorización ---
    # Creamos un MovimientoCategorizado vacío como sugerencia inicial
    sugerencia = MovimientoCategorizado(
        año=año, mes=mes,
        categoria1="", categoria2="", categoria3="",
        entidad="", importe=Decimal("0"), proveedor="",
        tipo_gasto="", cuenta="", banco=None, tipo_cuenta=None,
        estado="Real", confianza="ninguna", fuente="movimiento",
        requiere_confirmacion=False,
    )
    tui_cat = TUICategorizacion(sugerencia, datos_maestros)
    campos = tui_cat.run()

    if campos is None or campos in ("saltar", "salir", "volver"):
        consola.print("[dim]Cancelado.[/dim]")
        return

    # --- 3. Cuenta ---
    seleccion_cuenta = _tui_seleccionar_cuenta(cuentas)
    if seleccion_cuenta is None:
        consola.print("[dim]Cancelado.[/dim]")
        return
    cuenta, banco, tipo_cuenta = seleccion_cuenta
    consola.print(f"\n  Cuenta: [bold]{cuenta}[/bold]")

    # --- 4. Importe ---
    while True:
        raw = click.prompt("  Importe (negativo = gasto, positivo = ingreso)").strip().replace(",", ".")
        try:
            importe = Decimal(raw).quantize(Decimal("0.01"))
            break
        except InvalidOperation:
            consola.print("  [red]Valor no válido. Usa formato numérico (ej: -45.50)[/red]")

    # --- 5. Previsualización ---
    consola.print()
    tabla = Table(
        title="Movimiento Real a crear",
        box=box.SIMPLE_HEAD,
        show_lines=False,
        padding=(0, 1),
    )
    tabla.add_column("Campo",  style="dim", no_wrap=True)
    tabla.add_column("Valor",  no_wrap=False)

    color_imp = "red" if importe < 0 else "green"
    filas = [
        ("Fecha",       f"{mes} {año}"),
        ("Categoría 1", campos.get("categoria1", "") or "—"),
        ("Categoría 2", campos.get("categoria2", "") or "—"),
        ("Categoría 3", campos.get("categoria3", "") or "—"),
        ("Entidad",     campos.get("entidad",    "") or "—"),
        ("Proveedor",   campos.get("proveedor",  "") or "—"),
        ("Tipo gasto",  campos.get("tipo_gasto", "") or "—"),
        ("Cuenta",      cuenta),
        ("Banco",       banco or "—"),
        ("Importe",     f"[{color_imp}]{importe:+.2f}[/{color_imp}]"),
        ("Estado",      "[bold]Real[/bold]"),
    ]
    for campo, valor in filas:
        tabla.add_row(campo, valor)

    consola.print(tabla)

    if not click.confirm("  ¿Confirmar escritura?", default=True):
        consola.print("[dim]Cancelado.[/dim]")
        return

    # --- 6. Escritura ---
    mov = MovimientoCategorizado(
        año=año,
        mes=mes,
        categoria1=campos.get("categoria1", ""),
        categoria2=campos.get("categoria2", ""),
        categoria3=campos.get("categoria3", ""),
        entidad=campos.get("entidad", ""),
        importe=importe,
        proveedor=campos.get("proveedor", ""),
        tipo_gasto=campos.get("tipo_gasto", ""),
        cuenta=cuenta,
        banco=banco or None,
        tipo_cuenta=tipo_cuenta or None,
        estado="Real",
        confianza="alta",
        fuente="movimiento",
        requiere_confirmacion=False,
    )

    try:
        n = EscritorDatos(ruta_xlsx).escribir([mov])
        consola.print(f"\n  [green]✓ Movimiento escrito.[/green]")
    except Exception as e:
        consola.print(f"\n  [red]Error al escribir:[/red] {e}")
        raise SystemExit(1)

    # --- 7. Registrar revisión ---
    _registrar_revision_interactivo(consola, cuenta)
