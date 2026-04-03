"""Comando 'estado': muestra qué cuentas están al día y cuáles faltan por actualizar.

Cruza los marcadores de última importación (marcadores.json) con la lista de
cuentas definidas en la hoja Claves del xlsx.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import click
from rich.console import Console
from rich.table import Table
from rich import box

_RUTA_REVISIONES = Path("~/.config/presupuesto/revisiones.json").expanduser()


def _leer_cuentas_claves(ruta_xlsx: Path) -> list[tuple[str, str, str]]:
    """Devuelve [(cuenta, banco, tipo_cuenta)] desde la hoja Claves."""
    import openpyxl

    wb = openpyxl.load_workbook(str(ruta_xlsx), data_only=True, read_only=True)
    try:
        ws = wb["Claves"]
    except KeyError:
        wb.close()
        return []

    resultado = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        cuenta      = str(row[0]).strip()
        banco       = str(row[1] or "").strip()
        tipo_cuenta = str(row[2] or "").strip()
        if cuenta:
            resultado.append((cuenta, banco, tipo_cuenta))
    wb.close()
    return resultado


def _ultima_real(ruta_xlsx: Path) -> dict[str, date]:
    """Devuelve la fecha del último movimiento Real por cuenta, inferida de Año+Mes."""
    import openpyxl

    _MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
              "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    wb = openpyxl.load_workbook(str(ruta_xlsx), data_only=True, read_only=True)
    try:
        ws = wb["Datos"]
    except KeyError:
        wb.close()
        return {}

    ultimas: dict[str, date] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        estado = str(row[12] or "").strip()
        if estado != "Real":
            continue
        try:
            año = int(row[0])
            mes_str = str(row[1] or "").strip()
            mes_idx = _MESES.index(mes_str)            # lanza ValueError si no existe
        except (TypeError, ValueError):
            continue
        cuenta = str(row[9] or "").strip()
        if not cuenta:
            continue
        # Último día del mes como referencia
        siguiente_mes = date(año + (mes_idx == 11), (mes_idx + 1) % 12 + 1, 1)
        from datetime import timedelta
        fin_mes = siguiente_mes - timedelta(days=1)
        if cuenta not in ultimas or fin_mes > ultimas[cuenta]:
            ultimas[cuenta] = fin_mes

    wb.close()
    return ultimas


def _estado_marcador(marcador: date | None, hoy: date) -> tuple[str, str]:
    """Devuelve (icono, estilo_rich) según antigüedad del marcador."""
    if marcador is None:
        return "✗  Sin marcador", "red"
    dias = (hoy - marcador).days
    if dias <= 0:
        return "✓  Hoy",         "bold green"
    if hoy.year == marcador.year and hoy.month == marcador.month:
        return "✓  Este mes",    "green"
    # Mes anterior
    mes_ant = (hoy.month - 2) % 12 + 1
    año_ant  = hoy.year if hoy.month > 1 else hoy.year - 1
    if marcador.year == año_ant and marcador.month == mes_ant:
        return "⚠  Mes anterior", "yellow"
    return f"✗  Hace {dias}d",   "red"


@click.command("estado")
@click.option("--todos", is_flag=True, default=False,
              help="Incluir también las cuentas ya al día.")
def cmd_estado(todos: bool) -> None:
    """Muestra el estado de actualización de cada cuenta."""
    from presupuesto.config import cargar_config
    from presupuesto.duplicados import GestorRevisiones

    consola = Console()

    # --- Cargar xlsx ---
    config   = cargar_config()
    ruta_str = config.get("archivo_presupuesto", "")
    if not ruta_str:
        consola.print("[red]No hay ruta al xlsx configurada.[/red]")
        raise SystemExit(1)
    ruta_xlsx = Path(ruta_str).expanduser()
    if not ruta_xlsx.exists():
        consola.print(f"[red]No se encuentra:[/red] {ruta_xlsx}")
        raise SystemExit(1)

    hoy     = date.today()
    cuentas = _leer_cuentas_claves(ruta_xlsx)
    if not cuentas:
        consola.print("[yellow]No se encontraron cuentas en la hoja Claves.[/yellow]")
        return

    from presupuesto.duplicados import GestorRevisiones
    gestor       = GestorRevisiones()
    ultimas_real = _ultima_real(ruta_xlsx)

    # --- Clasificar ---
    pendientes: list[tuple] = []
    al_dia:     list[tuple] = []

    for cuenta, banco, tipo_cuenta in cuentas:
        marcador = gestor.obtener_revision(cuenta)
        ultima_r = ultimas_real.get(cuenta)
        estado_txt, estilo = _estado_marcador(marcador, hoy)

        fila = (cuenta, banco, tipo_cuenta, marcador, ultima_r, estado_txt, estilo)
        if "✓" in estado_txt:
            al_dia.append(fila)
        else:
            pendientes.append(fila)

    # --- Tabla ---
    def _hacer_tabla(filas: list[tuple], titulo: str) -> Table:
        tabla = Table(
            title=titulo,
            box=box.SIMPLE_HEAD,
            show_lines=False,
            padding=(0, 1),
        )
        tabla.add_column("Cuenta",        no_wrap=True)
        tabla.add_column("Banco",         no_wrap=True, style="dim")
        tabla.add_column("Tipo",          no_wrap=True, style="dim")
        tabla.add_column("Marcador",      no_wrap=True)
        tabla.add_column("Último Real",   no_wrap=True, style="dim")
        tabla.add_column("Estado",        no_wrap=True)

        for cuenta, banco, tipo_cuenta, marcador, ultima_r, estado_txt, estilo in filas:
            marcador_str  = marcador.isoformat()  if marcador  else "—"
            ultima_r_str  = ultima_r.isoformat()  if ultima_r  else "—"
            tabla.add_row(
                cuenta,
                banco,
                tipo_cuenta,
                marcador_str,
                ultima_r_str,
                f"[{estilo}]{estado_txt}[/{estilo}]",
            )
        return tabla

    if pendientes:
        consola.print(_hacer_tabla(pendientes, f"Cuentas pendientes de actualizar ({len(pendientes)})"))

    if todos and al_dia:
        consola.print(_hacer_tabla(al_dia, f"Cuentas al día ({len(al_dia)})"))

    if not pendientes:
        consola.print(f"\n  [bold green]✓ Todas las cuentas están al día.[/bold green]")
    elif not todos:
        consola.print(
            f"\n  [dim]{len(al_dia)} cuenta(s) al día. "
            f"Usa [bold]--todos[/bold] para verlas también.[/dim]"
        )
