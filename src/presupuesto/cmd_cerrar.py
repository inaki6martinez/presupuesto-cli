"""Comando 'cerrar': cierra el año creando el presupuesto del siguiente.

Operaciones:
1. Para cada mes del año siguiente sin entradas "Presupuesto":
   copia las entradas "Real" del mismo mes del año actual (excluyendo Excepcionales)
   como nuevas entradas "Presupuesto" del año siguiente.
2. Elimina todas las entradas "Presupuesto" del año actual.
"""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING

import click

if TYPE_CHECKING:
    pass

_MESES_ORD = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
              "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

_TIPO_EXCLUIR = "Excepcionales"

# Columnas de la hoja Datos (base 0)
_COL_AÑO       = 0
_COL_MES       = 1
_COL_CAT1      = 2
_COL_CAT2      = 3
_COL_CAT3      = 4
_COL_ENTIDAD   = 5
_COL_IMPORTE   = 6
_COL_PROVEEDOR = 7
_COL_TIPO_GASTO = 8
_COL_CUENTA    = 9
_COL_BANCO     = 10
_COL_TIPO_CUEN = 11
_COL_ESTADO    = 12


# ---------------------------------------------------------------------------
# Análisis del xlsx
# ---------------------------------------------------------------------------

def _analizar(ruta_xlsx: Path, cuenta: str = "") -> dict:
    """Lee la hoja Datos y devuelve:
    - real[año][mes] = lista de tuplas de fila (base 0)
    - presup[año] = lista de números de fila xlsx (base 1, min_row=2)
    - presup_meses[año] = set de meses con Presupuesto
    Si cuenta != "", filtra solo las filas de esa cuenta.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(ruta_xlsx), data_only=True, read_only=True)
    try:
        ws = wb["Datos"]
    except KeyError:
        wb.close()
        return {"real": {}, "presup": {}, "presup_meses": {}}

    real: dict[int, dict[str, list[tuple]]] = defaultdict(lambda: defaultdict(list))
    presup_filas: dict[int, list[int]] = defaultdict(list)
    presup_meses: dict[int, set[str]] = defaultdict(set)
    # fila_idx → mes, para poder filtrar borrado por mes
    presup_filas_mes: dict[int, dict[int, str]] = defaultdict(dict)

    cuenta_lower = cuenta.lower() if cuenta else ""

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[_COL_AÑO] is None:
            continue
        try:
            año = int(row[_COL_AÑO])
        except (TypeError, ValueError):
            continue
        mes    = str(row[_COL_MES]    or "").strip()
        estado = str(row[_COL_ESTADO] or "").strip()
        tipo   = str(row[_COL_TIPO_GASTO] or "").strip()
        fila_cuenta = str(row[_COL_CUENTA] or "").strip()

        # Filtro por cuenta si se especificó
        if cuenta_lower and fila_cuenta.lower() != cuenta_lower:
            continue

        if estado == "Real" and tipo != _TIPO_EXCLUIR:
            real[año][mes].append(tuple(row))
        elif estado == "Presupuesto":
            presup_filas[año].append(r_idx)
            presup_meses[año].add(mes)
            presup_filas_mes[año][r_idx] = mes

    wb.close()
    return {
        "real":             real,
        "presup_filas":     presup_filas,
        "presup_meses":     presup_meses,
        "presup_filas_mes": presup_filas_mes,
    }


def _plan(datos: dict, año_actual: int, mes_corte: int) -> dict:
    """Calcula qué meses crear y qué filas borrar.

    - crear_sig: meses ya terminados (< mes_corte) → Presupuesto año_actual+1
                 desde Real de año_actual.
    - crear_act: meses aún por venir (>= mes_corte) → Presupuesto año_actual
                 desde Real de año_actual-1 (si aún no tienen Presupuesto).
    - borrar:    filas Presupuesto de año_actual para meses ya terminados.
    """
    año_sig  = año_actual + 1
    año_ant  = año_actual - 1

    meses_terminados  = {m for i, m in enumerate(_MESES_ORD, start=1) if i < mes_corte}
    meses_pendientes  = {m for i, m in enumerate(_MESES_ORD, start=1) if i >= mes_corte}

    # 1. Presupuesto año siguiente: meses terminados sin presupuesto en año_sig
    ya_tiene_sig = datos["presup_meses"].get(año_sig, set())
    crear_sig: dict[str, list[tuple]] = {}
    for mes in _MESES_ORD:
        if mes not in meses_terminados or mes in ya_tiene_sig:
            continue
        filas = datos["real"].get(año_actual, {}).get(mes, [])
        if filas:
            crear_sig[mes] = filas

    # 2. Presupuesto año actual: meses pendientes sin presupuesto en año_actual
    ya_tiene_act = datos["presup_meses"].get(año_actual, set())
    crear_act: dict[str, list[tuple]] = {}
    for mes in _MESES_ORD:
        if mes not in meses_pendientes or mes in ya_tiene_act:
            continue
        filas = datos["real"].get(año_ant, {}).get(mes, [])
        if filas:
            crear_act[mes] = filas

    # 3. Borrar: Presupuesto año_actual (meses terminados) + todos los de año_ant
    borrar = [
        fila for fila, mes in datos.get("presup_filas_mes", {}).get(año_actual, {}).items()
        if mes in meses_terminados
    ]
    borrar += list(datos.get("presup_filas_mes", {}).get(año_ant, {}).keys())

    return {"crear_sig": crear_sig, "crear_act": crear_act, "borrar": borrar}


# ---------------------------------------------------------------------------
# TUI de previsualización y confirmación
# ---------------------------------------------------------------------------

def _tui_confirmar(año_actual: int, mes_corte: int, plan: dict, datos_: dict,
                   cuenta: str = "") -> bool:
    """Muestra resumen del plan y devuelve True si el usuario confirma."""
    from prompt_toolkit import Application
    from prompt_toolkit.application import get_app
    from prompt_toolkit.formatted_text import FormattedText
    from prompt_toolkit.key_binding import KeyBindings
    from prompt_toolkit.layout import Layout
    from prompt_toolkit.layout.containers import Window
    from prompt_toolkit.layout.controls import FormattedTextControl
    from prompt_toolkit.styles import Style

    año_sig = año_actual + 1
    año_ant = año_actual - 1
    crear_sig = plan["crear_sig"]
    crear_act = plan["crear_act"]
    borrar    = plan["borrar"]

    style = Style.from_dict({
        "titulo":  "bold",
        "titulo2": "bold #aaaaff",
        "ok":      "#00cc44",
        "nuevo":   "#55aaff bold",
        "nuevo2":  "#ffaa55 bold",
        "skip":    "#666666",
        "neg":     "#ff5555",
        "dim":     "#666666",
        "footer":  "#666666",
        "fkey":    "#aaaaaa bold",
    })

    state = {"resultado": None}

    def _render() -> FormattedText:
        try:
            size = get_app().output.get_size()
            w, h = size.columns, size.rows
        except Exception:
            w, h = 100, 40

        buf: list[tuple[str, str]] = []
        def t(st, s): buf.append((st, s))
        def nl(): buf.append(("", "\n"))

        hasta = _MESES_ORD[mes_corte - 2] if mes_corte > 1 else "—"
        titulo = f"  Cerrar {año_actual} hasta {hasta}"
        if cuenta:
            titulo += f"  [{cuenta}]"
        t("class:titulo", titulo)
        nl()
        t("class:dim", "─" * w)
        nl()

        # Sección 1: Presupuesto año siguiente
        t("class:titulo", f"  → Presupuesto {año_sig}  ")
        t("class:dim", f"(desde Real {año_actual})")
        nl()
        t("class:dim", f"  {'Mes':<6}  {'':^38}  {'Entradas':>8}")
        nl()

        presup_meses_sig = datos_["presup_meses"].get(año_sig, set())
        for i, mes in enumerate(_MESES_ORD, start=1):
            if i >= mes_corte:
                st, desc, n_str = "class:skip", f"· aún no terminado", ""
            elif mes in presup_meses_sig:
                st, desc, n_str = "class:ok", f"✓ ya tiene presupuesto", ""
            elif mes in crear_sig:
                st, desc = "class:nuevo", f"+ se creará desde {mes} {año_actual}"
                n_str = f"{len(crear_sig[mes]):>8}"
            else:
                st, desc, n_str = "class:skip", f"· sin datos Real en {mes} {año_actual}", ""
            t("class:dim", f"  {mes:<6}  ")
            t(st, f"{desc:<38}")
            t("class:dim", f"  {n_str}")
            nl()

        nl()

        # Sección 2: Presupuesto año actual (desde año anterior)
        if any(True for i, m in enumerate(_MESES_ORD, start=1) if i >= mes_corte):
            t("class:titulo2", f"  → Presupuesto {año_actual} pendiente  ")
            t("class:dim", f"(desde Real {año_ant})")
            nl()
            t("class:dim", f"  {'Mes':<6}  {'':^38}  {'Entradas':>8}")
            nl()

            presup_meses_act = datos_["presup_meses"].get(año_actual, set())
            for i, mes in enumerate(_MESES_ORD, start=1):
                if i < mes_corte:
                    continue
                if mes in presup_meses_act:
                    st, desc, n_str = "class:ok", f"✓ ya tiene presupuesto", ""
                elif mes in crear_act:
                    st, desc = "class:nuevo2", f"+ se creará desde {mes} {año_ant}"
                    n_str = f"{len(crear_act[mes]):>8}"
                else:
                    st, desc, n_str = "class:skip", f"· sin datos Real en {mes} {año_ant}", ""
                t("class:dim", f"  {mes:<6}  ")
                t(st, f"{desc:<38}")
                t("class:dim", f"  {n_str}")
                nl()
            nl()

        t("class:dim", "─" * w)
        nl()

        n_sig    = sum(len(v) for v in crear_sig.values())
        n_act    = sum(len(v) for v in crear_act.values())
        n_borrar = len(borrar)
        if n_sig:
            t("class:nuevo",  f"  + {n_sig} entradas nuevas Presupuesto {año_sig}")
            nl()
        if n_act:
            t("class:nuevo2", f"  + {n_act} entradas nuevas Presupuesto {año_actual}")
            nl()
        n_borrar_act = len([f for f, m in datos_["presup_filas_mes"].get(año_actual, {}).items()
                            if m in {mes for i, mes in enumerate(_MESES_ORD, start=1) if i < mes_corte}])
        n_borrar_ant = len(datos_["presup_filas_mes"].get(año_ant, {}))
        if n_borrar_act:
            t("class:neg", f"  - {n_borrar_act} entradas Presupuesto {año_actual} (meses terminados) se eliminarán")
            nl()
        if n_borrar_ant:
            t("class:neg", f"  - {n_borrar_ant} entradas Presupuesto {año_ant} (año pasado) se eliminarán")
            nl()

        nl()
        for k, desc in [("Enter", "Confirmar"), ("Esc", "Cancelar")]:
            t("class:fkey",   f" {k} ")
            t("class:footer", f"{desc}  ")

        return FormattedText(buf)

    kb = KeyBindings()

    @kb.add("enter")
    def _(e):
        state["resultado"] = True
        e.app.exit()

    @kb.add("escape")
    @kb.add("c-c")
    def _(e):
        state["resultado"] = False
        e.app.exit()

    app = Application(
        layout=Layout(Window(content=FormattedTextControl(text=_render, focusable=True))),
        key_bindings=kb,
        style=style,
        full_screen=True,
    )
    app.run()
    return state["resultado"] is True


# ---------------------------------------------------------------------------
# Escritura en xlsx
# ---------------------------------------------------------------------------

def _ejecutar(ruta_xlsx: Path, año_actual: int, plan: dict) -> tuple[int, int]:
    """Aplica el plan al xlsx. Devuelve (n_creadas, n_borradas)."""
    import openpyxl
    import shutil
    from datetime import datetime
    from presupuesto.escritor import detectar_formulas_cuenta, adaptar_formula_fila

    año_sig   = año_actual + 1
    crear_sig = plan["crear_sig"]
    crear_act = plan["crear_act"]
    borrar    = sorted(plan["borrar"], reverse=True)

    # Backup antes de modificar
    sufijo = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ruta_xlsx.parent / f"{ruta_xlsx.stem}_backup_{sufijo}{ruta_xlsx.suffix}"
    shutil.copy2(ruta_xlsx, backup)

    wb = openpyxl.load_workbook(str(ruta_xlsx))
    ws = wb["Datos"]

    # Detectar fórmulas de K y L antes de borrar filas (para replicarlas luego)
    primera_fila_datos = ws.max_row + 1  # tras borrar, las nuevas irán al final
    formula_k, formula_l = detectar_formulas_cuenta(ws, primera_fila_datos)

    def _append_presupuesto(plantilla: tuple, año: int) -> None:
        """Añade una fila de Presupuesto replicando las fórmulas de K y L."""
        nonlocal primera_fila_datos
        fila_num = ws.max_row + 1
        fila = list(plantilla)
        fila[_COL_AÑO]    = año
        fila[_COL_ESTADO] = "Presupuesto"
        ws.append(fila)
        # Sobrescribir K y L con fórmulas si las hay
        if formula_k:
            ws.cell(fila_num, _COL_BANCO + 1).value = adaptar_formula_fila(formula_k, fila_num)
        if formula_l:
            ws.cell(fila_num, _COL_TIPO_CUEN + 1).value = adaptar_formula_fila(formula_l, fila_num)

    # 1. Borrar filas Presupuesto del año actual (de abajo a arriba)
    for r_idx in borrar:
        ws.delete_rows(r_idx)

    # Re-detectar fórmulas después del borrado (las referencias ya se han ajustado)
    primera_fila_libre = ws.max_row + 1
    formula_k, formula_l = detectar_formulas_cuenta(ws, primera_fila_libre)

    # 2. Añadir Presupuesto del año siguiente (desde Real año actual)
    n_creadas = 0
    for mes in _MESES_ORD:
        if mes not in crear_sig:
            continue
        for plantilla in crear_sig[mes]:
            _append_presupuesto(plantilla, año_sig)
            n_creadas += 1

    # 3. Añadir Presupuesto pendiente del año actual (desde Real año anterior)
    for mes in _MESES_ORD:
        if mes not in crear_act:
            continue
        for plantilla in crear_act[mes]:
            _append_presupuesto(plantilla, año_actual)
            n_creadas += 1

    wb.save(str(ruta_xlsx))
    wb.close()

    return n_creadas, len(borrar)


# ---------------------------------------------------------------------------
# Comando click
# ---------------------------------------------------------------------------

@click.command("cerrar")
@click.option("--año", default=None, type=int,
              help="Año a cerrar (por defecto: año actual).")
@click.option("--mes", default=None, type=int,
              help="Mes de corte 1-12 (por defecto: mes actual). Solo se procesan meses anteriores.")
@click.option("--cuenta", default="", metavar="CUENTA",
              help="Limitar a una cuenta concreta (p.ej. 'Kutxabank').")
def cmd_cerrar(año, mes, cuenta):
    """Crea el presupuesto del año siguiente para los meses ya terminados.

    Ejecutado en junio, crea los presupuestos de enero a mayo del año siguiente
    y elimina las entradas Presupuesto de esos mismos meses del año actual.
    Con --cuenta, solo procesa las filas de esa cuenta.
    """
    from datetime import date
    from rich.console import Console
    from presupuesto.config import cargar_config

    consola = Console()

    config   = cargar_config()
    ruta_str = config.get("archivo_presupuesto", "")
    if not ruta_str:
        consola.print("[red]No hay ruta al xlsx configurada.[/red]")
        raise SystemExit(1)
    ruta_xlsx = Path(ruta_str).expanduser()
    if not ruta_xlsx.exists():
        consola.print(f"[red]No se encuentra:[/red] {ruta_xlsx}")
        raise SystemExit(1)

    hoy        = date.today()
    año_actual = año or hoy.year
    mes_corte  = mes or hoy.month   # meses estrictamente anteriores a este se procesan

    if mes_corte <= 1:
        consola.print("[yellow]No hay meses anteriores al mes 1 que procesar.[/yellow]")
        return

    if cuenta:
        consola.print(f"[dim]Filtrando por cuenta:[/dim] [bold]{cuenta}[/bold]")
    consola.print("[dim]Leyendo datos del xlsx…[/dim]")
    datos_ = _analizar(ruta_xlsx, cuenta=cuenta)
    plan   = _plan(datos_, año_actual, mes_corte)

    n_total = (sum(len(v) for v in plan["crear_sig"].values())
             + sum(len(v) for v in plan["crear_act"].values()))
    if n_total == 0 and not plan["borrar"]:
        consola.print("[yellow]No hay nada que hacer: todos los meses ya están presupuestados "
                      "o no tienen datos Real.[/yellow]")
        return

    if not _tui_confirmar(año_actual, mes_corte, plan, datos_, cuenta=cuenta):
        consola.print("[dim]Cancelado.[/dim]")
        return

    consola.print("\n[dim]Aplicando cambios…[/dim]")
    try:
        n_creadas, n_borradas = _ejecutar(ruta_xlsx, año_actual, plan)
    except Exception as e:
        consola.print(f"[red]Error:[/red] {e}")
        raise SystemExit(1)

    año_sig = año_actual + 1
    n_sig = sum(len(v) for v in plan["crear_sig"].values())
    n_act = sum(len(v) for v in plan["crear_act"].values())
    if n_sig:
        consola.print(f"[green]✓ {n_sig} entradas Presupuesto {año_sig} creadas.[/green]")
    if n_act:
        consola.print(f"[green]✓ {n_act} entradas Presupuesto {año_actual} creadas.[/green]")
    if n_borradas:
        consola.print(f"[green]✓ {n_borradas} entradas Presupuesto eliminadas.[/green]")
