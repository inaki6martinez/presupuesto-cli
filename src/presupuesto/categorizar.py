"""Motor de categorización de movimientos bancarios.

Implementa tres capas:
  1. Reglas exactas (reglas.json) — confianza alta.
  2. Similitud con historial (rapidfuzz) — confianza media o baja.
  3. Sin match — campos vacíos con sugerencias basadas en el contexto de cuenta.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path

from presupuesto.maestro import DatosMaestros
from presupuesto.parsers.base import MovimientoCrudo
from presupuesto.reglas import GestorReglas

# ---------------------------------------------------------------------------
# Contexto por cuenta
# ---------------------------------------------------------------------------

CONTEXTO_CUENTAS: dict[str, dict] = {
    "Cuenta Hipoteca": {
        "descripcion": "Casa y vivienda",
        "categoria1_probable": ["Vivienda", "Finanzas"],
        "tipo_gasto_probable": "Fijos",
    },
    "Cuenta Ocio": {
        "descripcion": "Ocio y salidas",
        "categoria1_probable": ["Ocio", "Alimentación"],
        "tipo_gasto_probable": "Discrecionales",
    },
    "Cuenta Nomina": {
        "descripcion": "Nómina y gastos personales",
        "categoria1_probable": ["Ingresos", "Gastos Personales", "Alimentación"],
        "tipo_gasto_probable": None,
    },
    "Cuenta Ahorro N26": {
        "descripcion": "Ocio y gastos discrecionales",
        "categoria1_probable": ["Ocio", "Gastos Personales"],
        "tipo_gasto_probable": "Discrecionales",
    },
    "Kutxabank": {
        "descripcion": "Peajes y gastos de transporte",
        "categoria1_probable": ["Transporte"],
        "tipo_gasto_probable": "Fijos",
    },
    "Ahorro colchon": {
        "descripcion": "Ahorro Trade Republic",
        "categoria1_probable": ["Ahorro", "Finanzas"],
        "tipo_gasto_probable": "Fijos",
    },
    "EPSV": {
        "descripcion": "Jubilación Indexa Capital",
        "categoria1_probable": ["Ahorro"],
        "tipo_gasto_probable": "Fijos",
    },
    "Fondos": {
        "descripcion": "Inversión Indexa Capital",
        "categoria1_probable": ["Ahorro", "Finanzas"],
        "tipo_gasto_probable": "Discrecionales",
    },
}

# ---------------------------------------------------------------------------
# Meses
# ---------------------------------------------------------------------------

_MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
          "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class MovimientoCategorizado:
    """Movimiento con todos los campos necesarios para escribir en la hoja 'Datos'."""
    # 13 columnas del xlsx (A→M)
    año: int
    mes: str
    categoria1: str
    categoria2: str
    categoria3: str
    entidad: str
    importe: Decimal
    proveedor: str
    tipo_gasto: str
    cuenta: str
    banco: str | None
    tipo_cuenta: str | None
    estado: str = "Real"
    # Metadatos (no se escriben en xlsx)
    confianza: str = "ninguna"        # "alta" | "media" | "baja" | "ninguna"
    fuente: str = ""                  # origen de la categorización (regla, historial, manual, sin match)
    requiere_confirmacion: bool = True
    concepto_original: str = ""
    n_originales: int = 1             # movimientos originales que forman este grupo (tras agrupar)


@dataclass
class _RegistroHistorial:
    """Entrada deduplicada del historial para la capa de similitud."""
    proveedor: str          # texto usado para el matching
    categoria1: str
    categoria2: str
    categoria3: str
    entidad: str
    tipo_gasto: str
    cuenta: str             # cuenta de origen (para el desempate por contexto)


# ---------------------------------------------------------------------------
# Categorizador
# ---------------------------------------------------------------------------

class Categorizador:
    """Categoriza movimientos crudos en tres capas."""

    def __init__(self, datos_maestros: DatosMaestros, gestor_reglas: GestorReglas):
        self._maestros = datos_maestros
        self._reglas   = gestor_reglas
        self._historial: list[_RegistroHistorial] = []

    # --- Carga del historial -----------------------------------------------

    def cargar_historial(self, ruta_xlsx: str | Path) -> int:
        """Carga los movimientos de la hoja 'Datos' y los deduplica por proveedor+categorización.

        Devuelve el número de registros únicos cargados.
        """
        import openpyxl

        ruta = Path(ruta_xlsx)
        if not ruta.exists():
            return 0

        wb = openpyxl.load_workbook(str(ruta), data_only=True)
        ws = wb["Datos"]

        vistos: set[tuple] = set()
        self._historial = []

        for r in range(2, ws.max_row + 1):
            proveedor  = str(ws.cell(r, 8).value or "").strip()
            categoria1 = str(ws.cell(r, 3).value or "").strip()
            categoria2 = str(ws.cell(r, 4).value or "").strip()
            categoria3 = str(ws.cell(r, 5).value or "").strip()
            entidad    = str(ws.cell(r, 6).value or "").strip()
            tipo_gasto = str(ws.cell(r, 9).value or "").strip()
            cuenta     = str(ws.cell(r, 10).value or "").strip()

            if not proveedor:
                continue

            clave = (proveedor.lower(), categoria1, categoria2, categoria3, entidad, tipo_gasto)
            if clave in vistos:
                continue
            vistos.add(clave)

            self._historial.append(_RegistroHistorial(
                proveedor=proveedor,
                categoria1=categoria1,
                categoria2=categoria2,
                categoria3=categoria3,
                entidad=entidad,
                tipo_gasto=tipo_gasto,
                cuenta=cuenta,
            ))

        wb.close()
        return len(self._historial)

    # --- API pública -------------------------------------------------------

    def categorizar(self, movimiento: MovimientoCrudo, cuenta: str) -> MovimientoCategorizado:
        """Categoriza un movimiento crudo pasando por las tres capas."""
        banco, tipo_cuenta = self._maestros.autocompletar_cuenta(cuenta)
        fecha   = movimiento.fecha
        año     = fecha.year
        mes     = _MESES[fecha.month - 1]
        base    = dict(
            año=año, mes=mes, importe=movimiento.importe,
            cuenta=cuenta, banco=banco, tipo_cuenta=tipo_cuenta,
            estado="Real", concepto_original=movimiento.concepto_original,
        )

        # Capa 1 — reglas
        resultado = self._capa_reglas(movimiento.concepto, base)
        if resultado:
            return resultado

        # Capa 2 — similitud historial
        resultado = self._capa_similitud(movimiento.concepto, cuenta, base)
        if resultado:
            return resultado

        # Capa 3 — sin match
        return self._capa_sin_match(cuenta, base)

    # --- Capas internas ----------------------------------------------------

    def _capa_reglas(self, concepto: str, base: dict) -> MovimientoCategorizado | None:
        resultado = self._reglas.buscar_match_con_patron(concepto)
        if resultado is None:
            return None
        campos, patron = resultado
        return MovimientoCategorizado(
            **base,
            categoria1=campos.get("categoria1", ""),
            categoria2=campos.get("categoria2", ""),
            categoria3=campos.get("categoria3", ""),
            entidad=campos.get("entidad", ""),
            proveedor=campos.get("proveedor", ""),
            tipo_gasto=campos.get("tipo_gasto", ""),
            confianza="alta",
            fuente=f"regla: {patron}",
            requiere_confirmacion=False,
        )

    def _capa_similitud(
        self, concepto: str, cuenta: str, base: dict
    ) -> MovimientoCategorizado | None:
        """Busca si algún proveedor del historial aparece literalmente en el concepto."""
        if not self._historial:
            return None

        concepto_lower = concepto.lower()
        candidatos = [
            r for r in self._historial
            if r.proveedor and r.proveedor.lower() in concepto_lower
        ]

        if not candidatos:
            return None

        elegido = self._desempatar(candidatos, cuenta)

        return MovimientoCategorizado(
            **base,
            categoria1=elegido.categoria1,
            categoria2=elegido.categoria2,
            categoria3=elegido.categoria3,
            entidad=elegido.entidad,
            proveedor=elegido.proveedor,
            tipo_gasto=elegido.tipo_gasto,
            confianza="media",
            fuente=f"historial: {elegido.proveedor}",
            requiere_confirmacion=True,
        )

    def _desempatar(
        self, candidatos: list[_RegistroHistorial], cuenta: str
    ) -> _RegistroHistorial:
        """Si hay varios candidatos empatados, prioriza el que cuadra con el contexto de cuenta."""
        if len(candidatos) == 1:
            return candidatos[0]

        contexto = CONTEXTO_CUENTAS.get(cuenta, {})
        cats_probables = contexto.get("categoria1_probable", [])

        if cats_probables:
            for candidato in candidatos:
                if candidato.categoria1 in cats_probables:
                    return candidato

        return candidatos[0]

    def _capa_sin_match(self, cuenta: str, base: dict) -> MovimientoCategorizado:
        """Sin match: campos vacíos con sugerencias del contexto de la cuenta."""
        contexto      = CONTEXTO_CUENTAS.get(cuenta, {})
        cats_probables = contexto.get("categoria1_probable", [])
        tipo_probable  = contexto.get("tipo_gasto_probable") or ""

        # Solo pre-rellenar categoría1 si es unívoca
        cat1_sugerida = cats_probables[0] if len(cats_probables) == 1 else ""

        return MovimientoCategorizado(
            **base,
            categoria1=cat1_sugerida,
            categoria2="",
            categoria3="",
            entidad="",
            proveedor="",
            tipo_gasto=tipo_probable,
            confianza="ninguna",
            fuente="sin match",
            requiere_confirmacion=True,
        )
