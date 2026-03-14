"""Escritura de movimientos categorizados en presupuesto.xlsx.

Abre el xlsx sin `data_only` para preservar fórmulas. Nunca crea un workbook
nuevo: siempre abre el existente y añade filas al final de la hoja 'Datos'.
Crea un backup automático antes de cualquier escritura.
"""

from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

if TYPE_CHECKING:
    from presupuesto.categorizar import MovimientoCategorizado


def leer_numero(valor) -> float | None:
    """Convierte el valor de una celda xlsx a float.

    Acepta:
    - Números literales (int, float, Decimal).
    - Fórmulas aritméticas simples como '=-106.25' o '=51.75+62.1'.
    Devuelve None si el valor está vacío o no se puede interpretar.
    """
    import re as _re
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    if s.startswith("="):
        s = s[1:].strip()
    # Solo evaluar si es aritmética pura: dígitos, operadores y punto decimal
    if _re.match(r'^[\d\s\.\+\-\*\/\(\)eE]+$', s):
        try:
            return float(eval(s, {"__builtins__": {}}))  # noqa: S307
        except Exception:
            pass
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def detectar_formulas_cuenta(ws, primera_libre: int) -> tuple[str | None, str | None]:
    """Detecta el patrón de fórmula de las columnas K (Banco) y L (Tipo cuenta).

    Busca en las filas existentes la primera celda K o L que contenga una
    fórmula (empieza por '='). Devuelve (formula_k, formula_l) con el patrón
    tal como aparece en la hoja, o None si no hay fórmulas.
    """
    formula_k: str | None = None
    formula_l: str | None = None
    for row in range(2, primera_libre):
        v_k = ws.cell(row, 11).value
        v_l = ws.cell(row, 12).value
        if formula_k is None and isinstance(v_k, str) and v_k.startswith("="):
            formula_k = v_k
        if formula_l is None and isinstance(v_l, str) and v_l.startswith("="):
            formula_l = v_l
        if formula_k and formula_l:
            break
    return formula_k, formula_l


def adaptar_formula_fila(formula: str, fila: int) -> str:
    """Reemplaza todas las referencias de fila en la fórmula por `fila`.

    Ejemplo: '=VLOOKUP(J5,Claves!$A:$C,2,0)' con fila=10
             → '=VLOOKUP(J10,Claves!$A:$C,2,0)'
    """
    return re.sub(r'\b([A-Z]+)(\d+)\b', lambda m: m.group(1) + str(fila), formula)


class EscritorDatos:
    """Escribe filas en la hoja 'Datos' de presupuesto.xlsx."""

    def __init__(self, ruta_xlsx: str | Path):
        self._ruta = Path(ruta_xlsx)
        if not self._ruta.exists():
            raise FileNotFoundError(f"No se encontró el archivo: {self._ruta}")

    def crear_backup(self) -> Path:
        """Crea una copia del xlsx con timestamp en el mismo directorio.

        Nombre del backup: presupuesto_backup_YYYYMMDD_HHMMSS.xlsx
        """
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre = f"{self._ruta.stem}_backup_{ts}{self._ruta.suffix}"
        ruta_backup = self._ruta.parent / nombre
        shutil.copy2(str(self._ruta), str(ruta_backup))
        return ruta_backup

    def escribir(
        self,
        movimientos: list[MovimientoCategorizado],
        crear_backup: bool = True,
    ) -> int:
        """Añade los movimientos como nuevas filas en la hoja 'Datos'.

        Columnas escritas (A→M, 13 campos):
        Año | Mes | Cat1 | Cat2 | Cat3 | Entidad | Importe | Proveedor |
        Tipo gasto | Cuenta | Banco | Tipo cuenta | Estado

        Args:
            movimientos:   lista de movimientos ya agrupados.
            crear_backup:  si True (defecto), hace copia antes de escribir.

        Returns el número de filas escritas.
        """
        if not movimientos:
            return 0

        if crear_backup:
            self.crear_backup()

        # Sin data_only para preservar fórmulas existentes
        wb = openpyxl.load_workbook(str(self._ruta))
        ws = wb["Datos"]

        # Localizar primera fila libre: buscar la última fila con datos reales
        primera_libre = 2  # mínimo: fila 2 (tras la cabecera)
        for r in range(ws.max_row, 1, -1):
            if any(ws.cell(r, c).value is not None for c in range(1, 14)):
                primera_libre = r + 1
                break

        # Detectar fórmulas de K y L para replicarlas en filas nuevas
        formula_k, formula_l = detectar_formulas_cuenta(ws, primera_libre)

        for i, m in enumerate(movimientos):
            fila = primera_libre + i
            ws.cell(fila,  1).value = m.año
            ws.cell(fila,  2).value = m.mes
            ws.cell(fila,  3).value = m.categoria1
            ws.cell(fila,  4).value = m.categoria2
            ws.cell(fila,  5).value = m.categoria3
            ws.cell(fila,  6).value = m.entidad
            ws.cell(fila,  7).value = float(m.importe)
            ws.cell(fila,  8).value = m.proveedor
            ws.cell(fila,  9).value = m.tipo_gasto
            ws.cell(fila, 10).value = m.cuenta
            # K y L: replicar fórmula si existe, si no usar el valor calculado
            ws.cell(fila, 11).value = (
                adaptar_formula_fila(formula_k, fila) if formula_k else m.banco
            )
            ws.cell(fila, 12).value = (
                adaptar_formula_fila(formula_l, fila) if formula_l else m.tipo_cuenta
            )
            ws.cell(fila, 13).value = m.estado

        wb.save(str(self._ruta))
        wb.close()

        return len(movimientos)
