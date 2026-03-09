"""Tests del módulo de expansión de cuotas hipotecarias."""

from __future__ import annotations

import dataclasses
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest

from presupuesto.categorizar import MovimientoCategorizado
from presupuesto.hipoteca import (
    buscar_cuota,
    es_cuota_hipoteca,
    expandir_hipoteca,
    expandir_hipotecas,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _maestros_mock():
    m = MagicMock()
    m.autocompletar_cuenta.return_value = ("BBVA", "Pasivo")
    return m


def _mov_hipoteca(**kwargs) -> MovimientoCategorizado:
    defaults = dict(
        año=2026, mes="Mar",
        categoria1="Vivienda", categoria2="Hipoteca", categoria3="",
        entidad="Piso", importe=Decimal("-843.88"), proveedor="",
        tipo_gasto="Fijos", cuenta="Cuenta Hipoteca",
        banco="BBVA", tipo_cuenta="Activos liquidos", estado="Real",
        confianza="alta", fuente="regla:amortizacion",
        requiere_confirmacion=False,
        concepto_original="28/02/2026 | Fee for loan amortization",
    )
    defaults.update(kwargs)
    return MovimientoCategorizado(**defaults)


def _xlsx_cuadro(ruta: Path, filas: list[tuple]) -> None:
    """Crea un xlsx mínimo con la hoja 'Cuadro hipteca'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuadro hipteca"
    ws.append(["Fecha", "Num cuota", "Cuota", "Intereses", "Amortización",
               "Amort. Acumulada", "Capital pendiente", "Amort. Puntual", "Tipo amort"])
    # fila 2-5 vacías (para simular las filas de resumen del xlsx real)
    for _ in range(4):
        ws.append([None] * 9)
    # cabecera real en fila 6 (ya escrita arriba como row 1 en este fixture,
    # pero buscar_cuota empieza en min_row=7 del xlsx real → aquí min_row=7 también)
    # Añadimos una fila extra para que los datos queden en fila 7+
    ws.insert_rows(1, amount=5)  # desplazar al mismo layout que el xlsx real
    for fecha, intereses, amortizacion in filas:
        ws.cell(ws.max_row + 1, 1).value = fecha  # type: ignore[assignment]
        row_n = ws.max_row
        ws.cell(row_n, 1).value = fecha
        ws.cell(row_n, 4).value = intereses
        ws.cell(row_n, 5).value = amortizacion
    wb.save(str(ruta))


def _xlsx_cuadro_simple(ruta: Path, filas_datos: list[tuple]) -> None:
    """Versión simplificada: crea xlsx con datos en filas 7+ directamente."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuadro hipteca"
    # 6 filas de cabecera/resumen (vacías)
    for _ in range(6):
        ws.append([None] * 9)
    for fecha, intereses, amortizacion in filas_datos:
        ws.append([fecha, None, None, float(intereses), float(amortizacion),
                   None, None, None, None])
    wb.save(str(ruta))


# ---------------------------------------------------------------------------
# es_cuota_hipoteca
# ---------------------------------------------------------------------------

def test_es_cuota_hipoteca_detecta_correctamente():
    assert es_cuota_hipoteca(_mov_hipoteca()) is True

def test_no_es_cuota_hipoteca_si_categoria_diferente():
    assert es_cuota_hipoteca(_mov_hipoteca(categoria1="Ahorro")) is False

def test_no_es_cuota_hipoteca_si_entidad_diferente():
    assert es_cuota_hipoteca(_mov_hipoteca(entidad="")) is False


# ---------------------------------------------------------------------------
# buscar_cuota
# ---------------------------------------------------------------------------

def test_buscar_cuota_encuentra_mes_correcto(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _xlsx_cuadro_simple(ruta, [
        (datetime(2026, 2, 1), 413.00, 430.88),
        (datetime(2026, 3, 1), 412.63, 431.25),
        (datetime(2026, 4, 1), 411.80, 432.08),
    ])
    resultado = buscar_cuota(2026, "Mar", ruta)
    assert resultado is not None
    intereses, amortizacion = resultado
    assert intereses == Decimal("412.63")
    assert amortizacion == Decimal("431.25")

def test_buscar_cuota_devuelve_none_si_mes_no_existe(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _xlsx_cuadro_simple(ruta, [(datetime(2026, 3, 1), 412.63, 431.25)])
    assert buscar_cuota(2026, "Feb", ruta) is None

def test_buscar_cuota_devuelve_none_si_archivo_no_existe(tmp_path):
    assert buscar_cuota(2026, "Mar", tmp_path / "no_existe.xlsx") is None

def test_buscar_cuota_devuelve_none_sin_hoja(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "OtraHoja"
    wb.save(str(ruta))
    assert buscar_cuota(2026, "Mar", ruta) is None


# ---------------------------------------------------------------------------
# expandir_hipoteca
# ---------------------------------------------------------------------------

def test_expandir_genera_tres_movimientos(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _xlsx_cuadro_simple(ruta, [(datetime(2026, 3, 1), 412.63, 431.25)])
    movs = expandir_hipoteca(_mov_hipoteca(), ruta, _maestros_mock())
    assert len(movs) == 3

def test_expandir_importes_correctos(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _xlsx_cuadro_simple(ruta, [(datetime(2026, 3, 1), 412.63, 431.25)])
    mov_int, mov_amort, mov_balance = expandir_hipoteca(_mov_hipoteca(), ruta, _maestros_mock())
    assert mov_int.importe    == Decimal("-412.63")
    assert mov_amort.importe  == Decimal("-431.25")
    assert mov_balance.importe == Decimal("431.25")

def test_expandir_categorias_intereses(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _xlsx_cuadro_simple(ruta, [(datetime(2026, 3, 1), 412.63, 431.25)])
    mov_int, _, _ = expandir_hipoteca(_mov_hipoteca(), ruta, _maestros_mock())
    assert mov_int.categoria1 == "Vivienda"
    assert mov_int.categoria2 == "Hipoteca"
    assert mov_int.tipo_gasto == "Fijos"
    assert mov_int.cuenta == "Cuenta Hipoteca"

def test_expandir_categorias_amortizacion(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _xlsx_cuadro_simple(ruta, [(datetime(2026, 3, 1), 412.63, 431.25)])
    _, mov_amort, _ = expandir_hipoteca(_mov_hipoteca(), ruta, _maestros_mock())
    assert mov_amort.categoria1 == "Ahorro"
    assert mov_amort.categoria2 == "Hipoteca"
    assert mov_amort.cuenta == "Cuenta Hipoteca"

def test_expandir_categorias_balance(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _xlsx_cuadro_simple(ruta, [(datetime(2026, 3, 1), 412.63, 431.25)])
    _, _, mov_balance = expandir_hipoteca(_mov_hipoteca(), ruta, _maestros_mock())
    assert mov_balance.categoria1 == "Finanzas"
    assert mov_balance.categoria2 == "Balance"
    assert mov_balance.cuenta     == "Hipoteca Piso"
    assert mov_balance.banco      == "BBVA"
    assert mov_balance.tipo_cuenta == "Pasivo"

def test_expandir_sin_cuota_devuelve_original(tmp_path):
    """Si no hay datos en el cuadro para ese mes, devuelve el movimiento original."""
    ruta = tmp_path / "pres.xlsx"
    _xlsx_cuadro_simple(ruta, [(datetime(2026, 2, 1), 413.00, 430.88)])
    mov = _mov_hipoteca(mes="Mar")
    resultado = expandir_hipoteca(mov, ruta, _maestros_mock())
    assert resultado == [mov]

def test_expandir_concepto_original_lleva_sufijo(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _xlsx_cuadro_simple(ruta, [(datetime(2026, 3, 1), 412.63, 431.25)])
    mov_int, mov_amort, mov_balance = expandir_hipoteca(_mov_hipoteca(), ruta, _maestros_mock())
    assert mov_int.concepto_original.endswith("[intereses]")
    assert mov_amort.concepto_original.endswith("[amortizacion]")
    assert mov_balance.concepto_original.endswith("[balance]")


# ---------------------------------------------------------------------------
# expandir_hipotecas (lista completa)
# ---------------------------------------------------------------------------

def test_expandir_hipotecas_solo_expande_hipoteca(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _xlsx_cuadro_simple(ruta, [(datetime(2026, 3, 1), 412.63, 431.25)])

    mov_hip = _mov_hipoteca()
    mov_otro = dataclasses.replace(
        _mov_hipoteca(),
        categoria1="Alimentación", categoria2="Supermercados", entidad="",
    )
    resultado = expandir_hipotecas([mov_otro, mov_hip, mov_otro], ruta, _maestros_mock())
    # mov_otro x2 + 3 de hipoteca = 5
    assert len(resultado) == 5
    assert resultado[0] is mov_otro
    assert resultado[4] is mov_otro
