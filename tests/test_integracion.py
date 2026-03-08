"""Tests de integración del flujo completo de importación.

Cubre:
1. Parsear fixture → categorizar con reglas → agrupar → escribir en xlsx.
2. Verificar detección de duplicados (importar el mismo fixture dos veces).
3. Verificar que el backup se crea antes de escribir.
4. Verificar que los valores escritos son válidos según el Maestro.
5. Verificar el marcador de última importación.
6. Tests con el xlsx real (se saltan si no está disponible).
"""

import shutil
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from presupuesto.agrupador import agrupar_movimientos
from presupuesto.categorizar import Categorizador
from presupuesto.duplicados import GestorMarcadores, detectar_duplicados
from presupuesto.escritor import EscritorDatos
from presupuesto.maestro import DatosMaestros
from presupuesto.parsers.n26 import ParserN26
from presupuesto.reglas import GestorReglas

# ---------------------------------------------------------------------------
# Rutas
# ---------------------------------------------------------------------------

XLSX_REAL   = Path("/mnt/c/Users/inaki.martinez/OneDrive/presupuesto/presupuesto.xlsx")
FIXTURE_N26 = Path(__file__).parent / "fixtures" / "n26_ejemplo.csv"
REGLAS_JSON = Path(__file__).parent.parent / "datos" / "reglas_iniciales.json"

CUENTA = "Cuenta Ahorro N26"

# Resultados esperados del fixture con las reglas_iniciales:
# 10 movimientos crudos → 5 grupos tras agrupar
_N_MOVS_CRUDOS  = 10
_N_GRUPOS       = 5


# ---------------------------------------------------------------------------
# Helper: xlsx mínimo con Maestro y Claves
# ---------------------------------------------------------------------------

def _crear_xlsx_test(ruta: Path) -> None:
    """Crea un xlsx mínimo con las hojas Datos (vacía), Maestro y Claves."""
    wb = openpyxl.Workbook()

    # Datos — solo cabecera
    ws_d = wb.active
    ws_d.title = "Datos"
    ws_d.append(["Año", "Mes", "Categoría 1", "Categoría 2", "Categoría 3",
                 "Entidad", "Importe", "Proveedor", "Tipo de Gasto",
                 "Cuentas", "Banco", "Tipo de cuenta", "Estado"])

    # Maestro — valores suficientes para cubrir lo que producen las reglas
    ws_m = wb.create_sheet("Maestro")
    ws_m.append(["Año", "Mes", "Cat1", "Cat2", "Cat3", "Entidad", "Proveedor",
                 "Tipo gasto", "Cuenta", "Banco", "Tipo cuenta"])
    for i, año in enumerate([2024, 2025, 2026, 2027], 2):
        ws_m.cell(i, 1, año)
    for i, mes in enumerate(
        ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"], 2
    ):
        ws_m.cell(i, 2, mes)
    cats1 = ["Alimentación", "Salud", "Ocio", "Transporte", "Finanzas", "Ahorro",
             "Ingresos", "Comunicaciones", "Gastos Personales", "Hogar",
             "Seguros", "Impuestos", "Educación", "Ropa", "Vivienda"]
    for i, c in enumerate(cats1, 2):
        ws_m.cell(i, 3, c)
    for i, tg in enumerate(["Fijos", "Discrecionales", "Opcionales", "Optimizable"], 2):
        ws_m.cell(i, 8, tg)
    cuentas = ["Cuenta Nomina", "Cuenta Ahorro N26", "Kutxabank",
               "Cuenta Hipoteca", "Cuenta Ocio", "Ahorro colchon", "EPSV", "Fondos"]
    for i, c in enumerate(cuentas, 2):
        ws_m.cell(i, 9, c)
    for i, b in enumerate(["Openbank", "N26", "Kutxabank", "BBVA",
                            "Trade republic", "Indexa Capital"], 2):
        ws_m.cell(i, 10, b)
    for i, tc in enumerate(["Activos liquidos", "Activos medio liquidos",
                             "Activos poco liquidos", "Pasivo"], 2):
        ws_m.cell(i, 11, tc)

    # Claves
    ws_c = wb.create_sheet("Claves")
    ws_c.append(["Cuenta", "Banco", "Tipo de cuenta"])
    for fila in [
        ("Cuenta Nomina",    "Openbank",        "Activos liquidos"),
        ("Cuenta Ahorro N26","N26",              "Activos liquidos"),
        ("Kutxabank",        "Kutxabank",        "Activos liquidos"),
        ("Cuenta Hipoteca",  "BBVA",             "Activos liquidos"),
        ("Cuenta Ocio",      "N26",              "Activos liquidos"),
        ("Ahorro colchon",   "Trade republic",   "Activos liquidos"),
        ("EPSV",             "Indexa Capital",   "Activos poco liquidos"),
        ("Fondos",           "Indexa Capital",   "Activos medio liquidos"),
    ]:
        ws_c.append(list(fila))

    wb.save(str(ruta))


def _parsear_y_categorizar(ruta_xlsx: Path) -> tuple:
    """Devuelve (movimientos_crudos, agrupados) para el fixture de N26."""
    parser = ParserN26()
    movs = parser.parsear(str(FIXTURE_N26))

    datos_maestros = DatosMaestros(ruta_xlsx)
    gestor_reglas  = GestorReglas(str(REGLAS_JSON))
    cat = Categorizador(datos_maestros, gestor_reglas)
    cat.cargar_historial(ruta_xlsx)

    categorizados = [cat.categorizar(m, CUENTA) for m in movs]
    agrupados = agrupar_movimientos(categorizados)
    return movs, agrupados


# ---------------------------------------------------------------------------
# 1. Flujo completo: parsear → categorizar → escribir
# ---------------------------------------------------------------------------

def test_flujo_completo_parse_y_escribir(tmp_path):
    """10 movimientos crudos del fixture N26 → 5 filas escritas en el xlsx."""
    ruta = tmp_path / "presupuesto.xlsx"
    _crear_xlsx_test(ruta)

    movs, agrupados = _parsear_y_categorizar(ruta)

    assert len(movs) == _N_MOVS_CRUDOS
    assert len(agrupados) == _N_GRUPOS

    escritor = EscritorDatos(ruta)
    n = escritor.escribir(agrupados, crear_backup=False)
    assert n == _N_GRUPOS

    # Verificar filas escritas
    wb = openpyxl.load_workbook(str(ruta), data_only=True)
    ws = wb["Datos"]
    assert ws.max_row == 1 + _N_GRUPOS   # 1 cabecera + N grupos
    wb.close()


def test_datos_escritos_correctos(tmp_path):
    """Los campos escritos corresponden a los movimientos del fixture."""
    ruta = tmp_path / "presupuesto.xlsx"
    _crear_xlsx_test(ruta)

    _, agrupados = _parsear_y_categorizar(ruta)
    EscritorDatos(ruta).escribir(agrupados, crear_backup=False)

    wb = openpyxl.load_workbook(str(ruta), data_only=True)
    ws = wb["Datos"]

    # Recoger todas las filas escritas
    filas = [[ws.cell(r, c).value for c in range(1, 14)]
             for r in range(2, ws.max_row + 1)]
    wb.close()

    # Todos los movimientos son de enero 2025 y de la cuenta correcta
    for fila in filas:
        assert fila[0] == 2025,       f"Año incorrecto: {fila[0]}"
        assert fila[1] == "Ene",      f"Mes incorrecto: {fila[1]}"
        assert fila[9] == CUENTA,     f"Cuenta incorrecta: {fila[9]}"
        assert fila[12] == "Real",    f"Estado incorrecto: {fila[12]}"
        assert fila[10] == "N26",     f"Banco incorrecto: {fila[10]}"
        assert fila[11] == "Activos liquidos", f"Tipo cuenta incorrecto: {fila[11]}"


def test_n_originales_refleja_agrupacion(tmp_path):
    """El grupo de 6 movimientos sin categoría tiene n_originales=6."""
    ruta = tmp_path / "presupuesto.xlsx"
    _crear_xlsx_test(ruta)

    _, agrupados = _parsear_y_categorizar(ruta)

    # El grupo más grande (sin categoría) debe tener n_originales=6
    max_grupo = max(agrupados, key=lambda m: m.n_originales)
    assert max_grupo.n_originales == 6
    assert sum(m.n_originales for m in agrupados) == _N_MOVS_CRUDOS


def test_importe_total_conservado(tmp_path):
    """La suma de importes agrupados debe igualar la suma de movimientos crudos."""
    ruta = tmp_path / "presupuesto.xlsx"
    _crear_xlsx_test(ruta)

    movs, agrupados = _parsear_y_categorizar(ruta)

    suma_crudos   = sum(m.importe for m in movs)
    suma_agrupados = sum(m.importe for m in agrupados)
    assert suma_crudos == suma_agrupados


# ---------------------------------------------------------------------------
# 2. Detección de duplicados (importar mismo fixture dos veces)
# ---------------------------------------------------------------------------

def test_segunda_importacion_detecta_todos_como_duplicados(tmp_path):
    """Importar el mismo fixture dos veces: la segunda detecta todos como duplicados."""
    ruta = tmp_path / "presupuesto.xlsx"
    _crear_xlsx_test(ruta)

    _, agrupados = _parsear_y_categorizar(ruta)

    # Primera importación
    EscritorDatos(ruta).escribir(agrupados, crear_backup=False)

    # Segunda pasada: todos deben ser detectados como duplicados
    duplicados = detectar_duplicados(agrupados, ruta)
    assert len(duplicados) == _N_GRUPOS
    assert all(fila >= 2 for _, fila in duplicados)  # fila 2 o posterior


def test_segunda_importacion_datos_diferentes_no_son_duplicados(tmp_path):
    """Movimientos con importe diferente no se detectan como duplicados."""
    ruta = tmp_path / "presupuesto.xlsx"
    _crear_xlsx_test(ruta)

    _, agrupados = _parsear_y_categorizar(ruta)
    EscritorDatos(ruta).escribir(agrupados, crear_backup=False)

    # Modificar importes ligeramente (más de ±0.01) → no deben ser duplicados
    import dataclasses
    agrupados_mod = [
        dataclasses.replace(m, importe=m.importe - Decimal("0.50"))
        for m in agrupados
    ]
    duplicados = detectar_duplicados(agrupados_mod, ruta)
    assert len(duplicados) == 0


# ---------------------------------------------------------------------------
# 3. Backup
# ---------------------------------------------------------------------------

def test_backup_se_crea_antes_de_escribir(tmp_path):
    """Se crea exactamente un archivo de backup antes de la escritura."""
    ruta = tmp_path / "presupuesto.xlsx"
    _crear_xlsx_test(ruta)

    _, agrupados = _parsear_y_categorizar(ruta)
    EscritorDatos(ruta).escribir(agrupados, crear_backup=True)

    backups = list(tmp_path.glob("*backup*.xlsx"))
    assert len(backups) == 1
    assert backups[0].exists()


def test_backup_contiene_datos_originales(tmp_path):
    """El backup conserva el estado anterior a la escritura."""
    ruta = tmp_path / "presupuesto.xlsx"
    _crear_xlsx_test(ruta)

    _, agrupados = _parsear_y_categorizar(ruta)
    escritor = EscritorDatos(ruta)
    ruta_backup = escritor.crear_backup()

    # Verificar que el backup existe y es un xlsx válido
    assert ruta_backup.exists()
    wb = openpyxl.load_workbook(str(ruta_backup), data_only=True)
    assert "Datos" in wb.sheetnames
    assert wb["Datos"].max_row == 1   # solo cabecera, sin datos aún
    wb.close()


# ---------------------------------------------------------------------------
# 4. Valores válidos según Maestro
# ---------------------------------------------------------------------------

def test_valores_escritos_validos_segun_maestro(tmp_path):
    """Los valores no vacíos de categoría1, tipo_gasto y cuenta están en el Maestro."""
    ruta = tmp_path / "presupuesto.xlsx"
    _crear_xlsx_test(ruta)

    _, agrupados = _parsear_y_categorizar(ruta)
    EscritorDatos(ruta).escribir(agrupados, crear_backup=False)

    datos_maestros = DatosMaestros(ruta)
    cats1_validas  = set(datos_maestros.categorias1)
    tipos_validos  = set(datos_maestros.tipos_gasto)
    cuentas_validas = set(datos_maestros.cuentas)

    wb = openpyxl.load_workbook(str(ruta), data_only=True)
    ws = wb["Datos"]
    for r in range(2, ws.max_row + 1):
        cat1       = ws.cell(r, 3).value or ""
        tipo_gasto = ws.cell(r, 9).value or ""
        cuenta     = ws.cell(r, 10).value or ""

        if cat1:
            assert cat1 in cats1_validas, f"Fila {r}: categoría1 inválida '{cat1}'"
        if tipo_gasto:
            assert tipo_gasto in tipos_validos, f"Fila {r}: tipo_gasto inválido '{tipo_gasto}'"
        if cuenta:
            assert cuenta in cuentas_validas, f"Fila {r}: cuenta inválida '{cuenta}'"
    wb.close()


# ---------------------------------------------------------------------------
# 5. Marcador de última importación
# ---------------------------------------------------------------------------

def test_marcador_actualizado_previene_reimportacion(tmp_path):
    """Tras actualizar el marcador, una segunda llamada a filtrar descarta todo."""
    ruta = tmp_path / "presupuesto.xlsx"
    _crear_xlsx_test(ruta)
    ruta_marcadores = tmp_path / "marcadores.json"

    movs, agrupados = _parsear_y_categorizar(ruta)
    EscritorDatos(ruta).escribir(agrupados, crear_backup=False)

    # Simular actualización del marcador con la fecha del último movimiento
    gestor = GestorMarcadores(ruta_marcadores)
    max_fecha = max(m.fecha for m in movs)
    gestor.actualizar_marcador(CUENTA, max_fecha)

    # Segunda importación: filtrar debe descartar todos
    gestor2 = GestorMarcadores(ruta_marcadores)
    aceptados, descartados = gestor2.filtrar_movimientos(movs, CUENTA)
    assert descartados == _N_MOVS_CRUDOS
    assert len(aceptados) == 0


def test_sin_marcador_pasan_todos_los_movimientos(tmp_path):
    """Sin marcador previo, ningún movimiento se filtra."""
    ruta_marcadores = tmp_path / "marcadores.json"
    gestor = GestorMarcadores(ruta_marcadores)

    movs = ParserN26().parsear(str(FIXTURE_N26))
    aceptados, descartados = gestor.filtrar_movimientos(movs, CUENTA)
    assert len(aceptados) == _N_MOVS_CRUDOS
    assert descartados == 0


def test_marcador_con_fecha_intermedia(tmp_path):
    """Marcador a fecha intermedia filtra solo los anteriores."""
    ruta_marcadores = tmp_path / "marcadores.json"
    gestor = GestorMarcadores(ruta_marcadores)
    # Marcador al 2025-01-15: deja pasar los 5 movimientos posteriores
    gestor.actualizar_marcador(CUENTA, date(2025, 1, 15))

    movs = ParserN26().parsear(str(FIXTURE_N26))
    aceptados, descartados = gestor.filtrar_movimientos(movs, CUENTA)

    # Movimientos del fixture: 4 anteriores o iguales al 15/01 (03,07,10,15)
    # y 6 posteriores (16,20,22,25,28,31)
    assert len(aceptados) == 6
    assert descartados == 4
    assert all(m.fecha > date(2025, 1, 15) for m in aceptados)


# ---------------------------------------------------------------------------
# 6. Tests con el xlsx real (se saltan si no está disponible)
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def xlsx_copia(tmp_path_factory):
    """Copia temporal del xlsx real. Se salta si el original no existe."""
    if not XLSX_REAL.exists():
        pytest.skip("presupuesto.xlsx real no disponible")
    copia = tmp_path_factory.mktemp("real") / "presupuesto_test.xlsx"
    shutil.copy2(str(XLSX_REAL), str(copia))
    return copia


def test_flujo_con_xlsx_real_escribe_filas(xlsx_copia):
    """Flujo completo sobre copia del xlsx real: las filas se añaden al final."""
    wb_antes = openpyxl.load_workbook(str(xlsx_copia), data_only=True)
    filas_antes = wb_antes["Datos"].max_row
    wb_antes.close()

    movs, agrupados = _parsear_y_categorizar(xlsx_copia)

    assert len(movs) == _N_MOVS_CRUDOS
    # Con el historial real hay más entradas únicas → más grupos (≤ movimientos crudos)
    assert 1 <= len(agrupados) <= _N_MOVS_CRUDOS
    n_grupos = len(agrupados)

    EscritorDatos(xlsx_copia).escribir(agrupados, crear_backup=False)

    wb_despues = openpyxl.load_workbook(str(xlsx_copia), data_only=True)
    filas_despues = wb_despues["Datos"].max_row
    wb_despues.close()

    assert filas_despues == filas_antes + n_grupos


def test_flujo_con_xlsx_real_preserva_hojas(xlsx_copia):
    """Después de escribir, las demás hojas del xlsx real siguen intactas."""
    _, agrupados = _parsear_y_categorizar(xlsx_copia)
    EscritorDatos(xlsx_copia).escribir(agrupados, crear_backup=False)

    wb = openpyxl.load_workbook(str(xlsx_copia), data_only=True)
    assert "Datos"   in wb.sheetnames
    assert "Maestro" in wb.sheetnames
    assert "Claves"  in wb.sheetnames
    wb.close()


def test_flujo_con_xlsx_real_valores_validos_maestro(xlsx_copia):
    """Los valores escritos sobre el xlsx real son válidos según su Maestro."""
    datos_maestros = DatosMaestros(xlsx_copia)
    cats1_validas  = set(datos_maestros.categorias1)
    tipos_validos  = set(datos_maestros.tipos_gasto)
    cuentas_validas = set(datos_maestros.cuentas)

    _, agrupados = _parsear_y_categorizar(xlsx_copia)
    EscritorDatos(xlsx_copia).escribir(agrupados, crear_backup=False)

    wb = openpyxl.load_workbook(str(xlsx_copia), data_only=True)
    ws = wb["Datos"]
    # Leer solo las últimas _N_GRUPOS filas (las recién escritas)
    primera_nueva = ws.max_row - _N_GRUPOS + 1
    for r in range(primera_nueva, ws.max_row + 1):
        cat1       = ws.cell(r, 3).value or ""
        tipo_gasto = ws.cell(r, 9).value or ""
        cuenta     = ws.cell(r, 10).value or ""

        if cat1:
            assert cat1 in cats1_validas,  f"Fila {r}: cat1 inválida '{cat1}'"
        if tipo_gasto:
            assert tipo_gasto in tipos_validos, f"Fila {r}: tipo_gasto inválido '{tipo_gasto}'"
        if cuenta:
            assert cuenta in cuentas_validas, f"Fila {r}: cuenta inválida '{cuenta}'"
    wb.close()


def test_flujo_con_xlsx_real_backup_creado(xlsx_copia):
    """La escritura sobre el xlsx real genera un backup en el mismo directorio."""
    directorio = xlsx_copia.parent
    backups_antes = list(directorio.glob("*backup*.xlsx"))

    _, agrupados = _parsear_y_categorizar(xlsx_copia)
    EscritorDatos(xlsx_copia).escribir(agrupados, crear_backup=True)

    backups_despues = list(directorio.glob("*backup*.xlsx"))
    assert len(backups_despues) == len(backups_antes) + 1


def test_duplicados_sobre_xlsx_real(xlsx_copia):
    """Importar el fixture sobre el xlsx real y re-detectar duplicados."""
    _, agrupados = _parsear_y_categorizar(xlsx_copia)
    EscritorDatos(xlsx_copia).escribir(agrupados, crear_backup=False)

    # Los mismos movimientos agrupados ahora deben detectarse como duplicados
    duplicados = detectar_duplicados(agrupados, xlsx_copia)
    assert len(duplicados) == len(agrupados)
