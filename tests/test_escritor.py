"""Tests del EscritorDatos."""

from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from presupuesto.categorizar import MovimientoCategorizado
from presupuesto.duplicados import GestorMarcadores
from presupuesto.escritor import EscritorDatos

XLSX_REAL = Path("/mnt/c/Users/inaki.martinez/OneDrive/presupuesto/presupuesto.xlsx")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _crear_xlsx(ruta: Path, filas_datos: list[tuple] | None = None) -> None:
    """Crea un xlsx mínimo con las hojas Datos, Maestro y Claves."""
    wb = openpyxl.Workbook()

    ws_datos = wb.active
    ws_datos.title = "Datos"
    ws_datos.append(["Año", "Mes", "Categoría 1", "Categoría 2", "Categoría 3",
                     "Entidad", "Importe", "Proveedor", "Tipo de Gasto",
                     "Cuentas", "Banco", "Tipo de cuenta", "Estado"])
    if filas_datos:
        for fila in filas_datos:
            ws_datos.append(list(fila))

    wb.create_sheet("Maestro")
    wb.create_sheet("Claves")
    wb.save(str(ruta))


def _mov(**kwargs) -> MovimientoCategorizado:
    defaults = dict(
        año=2026, mes="Mar",
        categoria1="Alimentación", categoria2="Supermercados", categoria3="",
        entidad="", importe=Decimal("-25.00"), proveedor="Eroski",
        tipo_gasto="Discrecionales", cuenta="Cuenta Nomina",
        banco="Openbank", tipo_cuenta="Activos liquidos", estado="Real",
    )
    defaults.update(kwargs)
    return MovimientoCategorizado(**defaults)


# ---------------------------------------------------------------------------
# Escritura básica
# ---------------------------------------------------------------------------

def test_escribe_filas_en_hoja_datos(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _crear_xlsx(ruta)
    escritor = EscritorDatos(ruta)
    n = escritor.escribir([_mov(), _mov(importe=Decimal("-10.00"))], crear_backup=False)

    assert n == 2
    wb = openpyxl.load_workbook(str(ruta), data_only=True)
    ws = wb["Datos"]
    assert ws.cell(2, 1).value == 2026
    assert ws.cell(2, 2).value == "Mar"
    assert ws.cell(2, 3).value == "Alimentación"
    assert ws.cell(2, 7).value == pytest.approx(-25.00)
    assert ws.cell(3, 7).value == pytest.approx(-10.00)
    wb.close()


def test_columnas_en_orden_correcto(tmp_path):
    """Verifica el orden A→M de los 13 campos."""
    ruta = tmp_path / "pres.xlsx"
    _crear_xlsx(ruta)
    m = _mov(
        año=2026, mes="Mar", categoria1="Ocio", categoria2="Entretenimiento",
        categoria3="Streaming", entidad="Netflix", importe=Decimal("-13.99"),
        proveedor="Netflix", tipo_gasto="Discrecionales", cuenta="Cuenta Ocio",
        banco="N26", tipo_cuenta="Activos liquidos", estado="Real",
    )
    escritor = EscritorDatos(ruta)
    escritor.escribir([m], crear_backup=False)

    wb = openpyxl.load_workbook(str(ruta), data_only=True)
    ws = wb["Datos"]
    fila = [ws.cell(2, c).value for c in range(1, 14)]
    wb.close()

    assert fila[0]  == 2026
    assert fila[1]  == "Mar"
    assert fila[2]  == "Ocio"
    assert fila[3]  == "Entretenimiento"
    assert fila[4]  == "Streaming"
    assert fila[5]  == "Netflix"
    assert fila[6]  == pytest.approx(-13.99)
    assert fila[7]  == "Netflix"
    assert fila[8]  == "Discrecionales"
    assert fila[9]  == "Cuenta Ocio"
    assert fila[10] == "N26"
    assert fila[11] == "Activos liquidos"
    assert fila[12] == "Real"


def test_aniade_tras_filas_existentes(tmp_path):
    """Los movimientos se añaden después de los datos ya existentes."""
    ruta = tmp_path / "pres.xlsx"
    _crear_xlsx(ruta, filas_datos=[
        (2025, "Dic", "Ocio", "", "", "", -50.0, "Spotify", "Discrecionales",
         "Cuenta Ocio", "N26", "Activos liquidos", "Real"),
    ])
    escritor = EscritorDatos(ruta)
    escritor.escribir([_mov()], crear_backup=False)

    wb = openpyxl.load_workbook(str(ruta), data_only=True)
    ws = wb["Datos"]
    assert ws.cell(2, 3).value == "Ocio"      # fila existente
    assert ws.cell(3, 3).value == "Alimentación"  # fila nueva
    wb.close()


def test_lista_vacia_devuelve_cero_y_no_modifica(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _crear_xlsx(ruta)
    mtime_antes = ruta.stat().st_mtime
    n = EscritorDatos(ruta).escribir([], crear_backup=False)
    assert n == 0
    # El archivo no se ha modificado
    assert ruta.stat().st_mtime == mtime_antes


def test_archivo_no_existe_lanza_error(tmp_path):
    with pytest.raises(FileNotFoundError):
        EscritorDatos(tmp_path / "no_existe.xlsx")


# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------

def test_backup_se_crea_antes_de_escribir(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _crear_xlsx(ruta)
    escritor = EscritorDatos(ruta)
    ruta_backup = escritor.crear_backup()

    assert ruta_backup.exists()
    assert "backup" in ruta_backup.name
    assert ruta_backup.suffix == ".xlsx"


def test_backup_automatico_al_escribir(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _crear_xlsx(ruta)
    EscritorDatos(ruta).escribir([_mov()])  # crear_backup=True por defecto

    backups = list(tmp_path.glob("*backup*.xlsx"))
    assert len(backups) == 1


def test_sin_backup_no_crea_fichero(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _crear_xlsx(ruta)
    EscritorDatos(ruta).escribir([_mov()], crear_backup=False)

    backups = list(tmp_path.glob("*backup*.xlsx"))
    assert len(backups) == 0


# ---------------------------------------------------------------------------
# Preservación del workbook
# ---------------------------------------------------------------------------

def test_otras_hojas_no_se_modifican(tmp_path):
    ruta = tmp_path / "pres.xlsx"
    _crear_xlsx(ruta)

    # Añadir contenido en Maestro y Claves antes de escribir
    wb = openpyxl.load_workbook(str(ruta))
    wb["Maestro"]["A1"] = "contenido_maestro"
    wb["Claves"]["A1"] = "contenido_claves"
    wb.save(str(ruta))

    EscritorDatos(ruta).escribir([_mov()], crear_backup=False)

    wb2 = openpyxl.load_workbook(str(ruta), data_only=True)
    assert wb2["Maestro"]["A1"].value == "contenido_maestro"
    assert wb2["Claves"]["A1"].value == "contenido_claves"
    wb2.close()


def test_formula_en_datos_se_preserva(tmp_path):
    """Una fórmula fuera de las 13 columnas de datos no se sobreescribe."""
    ruta = tmp_path / "pres.xlsx"
    _crear_xlsx(ruta)

    # Insertar fórmula en columna O (fuera de las 13 columnas de datos)
    wb = openpyxl.load_workbook(str(ruta))
    wb["Datos"]["O1"] = "=SUM(G:G)"
    wb.save(str(ruta))

    EscritorDatos(ruta).escribir([_mov()], crear_backup=False)

    wb2 = openpyxl.load_workbook(str(ruta))  # sin data_only para ver fórmulas
    assert wb2["Datos"]["O1"].value == "=SUM(G:G)"
    wb2.close()


# ---------------------------------------------------------------------------
# Integración con GestorMarcadores
# ---------------------------------------------------------------------------

def test_marcador_se_actualiza_tras_escritura_exitosa(tmp_path):
    """Flujo completo: escribir movimientos y luego actualizar el marcador."""
    ruta_xlsx = tmp_path / "pres.xlsx"
    _crear_xlsx(ruta_xlsx)
    ruta_marcadores = tmp_path / "marcadores.json"

    movimientos = [
        _mov(año=2026, mes="Mar"),
        _mov(año=2026, mes="Feb"),
    ]

    # Simula lo que hará el flujo principal: escribir + actualizar marcador
    escritor = EscritorDatos(ruta_xlsx)
    n_escritos = escritor.escribir(movimientos, crear_backup=False)
    assert n_escritos == 2

    gestor = GestorMarcadores(ruta_marcadores)
    gestor.actualizar_marcador("Cuenta Nomina", date(2026, 3, 31))

    gestor2 = GestorMarcadores(ruta_marcadores)
    assert gestor2.obtener_marcador("Cuenta Nomina") == date(2026, 3, 31)


# ---------------------------------------------------------------------------
# Tests con el xlsx real (se saltan si no está disponible)
# ---------------------------------------------------------------------------

def test_escribe_en_copia_del_xlsx_real(tmp_path):
    if not XLSX_REAL.exists():
        pytest.skip("presupuesto.xlsx real no disponible")

    import shutil
    copia = tmp_path / "presupuesto_test.xlsx"
    shutil.copy2(str(XLSX_REAL), str(copia))

    escritor = EscritorDatos(copia)
    movs = [_mov(), _mov(importe=Decimal("-50.00"), categoria1="Ocio")]
    n = escritor.escribir(movs, crear_backup=True)
    assert n == 2

    # Verificar que el backup existe
    backups = list(tmp_path.glob("*backup*.xlsx"))
    assert len(backups) == 1

    # Verificar que las filas se escribieron
    wb = openpyxl.load_workbook(str(copia), data_only=True)
    ws = wb["Datos"]
    # Última fila debe ser nuestra
    ultima = ws.max_row
    assert ws.cell(ultima, 3).value in ("Alimentación", "Ocio")
    # Las hojas del xlsx real siguen existiendo
    assert "Maestro" in wb.sheetnames
    assert "Claves" in wb.sheetnames
    wb.close()
